#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Catalog Intelligence Engine V5 for PriceBot.

Reads a pharmacy Excel/CSV file and produces conservative enrichment outputs:
- products_ready_for_upload.xlsx
- products_needs_review.xlsx
- catalog_quality_report.xlsx
- golden_text_tests_suggested.xlsx
- ocr_keywords_suggested.xlsx

No uncertain facts are invented. Rows with missing/ambiguous facts are marked
needs_review so a human can fix them before import.
"""
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Dict, List

try:
    import openpyxl
except Exception as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required. Install requirements.txt first.") from exc

ROOT = Path(__file__).resolve().parents[1]
import sys
sys.path.insert(0, str(ROOT))
from services.catalog_quality import REQUIRED_SCHEMA, analyze_products, normalize_text, split_tokens  # noqa: E402

MEDICINE_HINTS = {"tab", "tablet", "cap", "capsule", "syrup", "susp", "injection", "amp", "vial", "drops", "supp", "suppository", "mg", "mcg", "iu", "ml"}
COSMETIC_HINTS = {"cleanser", "cream", "lotion", "serum", "gel", "shampoo", "sunscreen", "toner", "mask", "scrub", "moisturizer", "balm", "غسول", "كريم", "لوشن", "سيروم"}
FORMS = ["tablet", "capsule", "syrup", "suspension", "injection", "drops", "cream", "ointment", "gel", "suppository", "pessary", "spray", "solution", "sachet", "cleanser", "face wash", "lotion", "serum", "shampoo", "conditioner", "sunscreen", "toner", "mask", "scrub", "oil", "moisturizer", "balm", "body wash", "deodorant"]
GENERIC_ALIAS = {"cream", "gel", "cleanser", "lotion", "غسول", "كريم", "لوشن", "syrup", "tablet", "capsule"}


def strength(text: str) -> str:
    m = re.search(r"\b\d+(?:\.\d+)?\s*(?:mg|mcg|g|iu|%)\s*(?:/\s*\d+\s*ml)?\b", text, re.I)
    return re.sub(r"\s+", "", m.group(0)).lower() if m else ""


def size(text: str) -> str:
    m = re.search(r"\b\d+(?:\.\d+)?\s*(?:ml|l|g|kg|tablets?|tabs?|capsules?|caps?)\b", text, re.I)
    return re.sub(r"\s+", "", m.group(0)).lower() if m else ""


def guess_form(text: str) -> str:
    low = normalize_text(text)
    for f in sorted(FORMS, key=len, reverse=True):
        if normalize_text(f) in low:
            return f
    arabic = {"غسول":"cleanser", "كريم":"cream", "لوشن":"lotion", "شراب":"syrup", "حبوب":"tablet", "اقراص":"tablet", "نقط":"drops", "قطره":"drops"}
    for k, v in arabic.items():
        if k in low:
            return v
    return ""


def guess_category(name: str, form: str, active: str) -> str:
    low = normalize_text(name)
    if active or any(x in low for x in MEDICINE_HINTS) or form in {"tablet", "capsule", "syrup", "suspension", "injection", "drops", "suppository", "pessary", "sachet"}:
        return "medicine"
    if any(x in low for x in COSMETIC_HINTS) or form in {"cleanser", "cream", "lotion", "serum", "gel", "shampoo", "sunscreen", "toner", "mask", "scrub", "oil", "moisturizer", "balm", "body wash", "deodorant"}:
        return "cosmetic"
    return "unknown"


def guess_family(name: str, brand: str, form: str, st: str, sz: str) -> str:
    value = str(name or "")
    for token in [brand, st, sz, form]:
        if token:
            value = re.sub(re.escape(str(token)), " ", value, flags=re.I)
    value = re.sub(r"\b\d+(?:\.\d+)?\s*(mg|mcg|g|ml|iu|%)\b", " ", value, flags=re.I)
    return " ".join(value.split())[:80]


def aliases_for(row: dict) -> str:
    base = [row.get("name", ""), f"{row.get('brand','')} {row.get('product_family','')}", f"{row.get('brand','')} {row.get('product_family','')} {row.get('form','')}"]
    result = []
    for a in base + split_tokens(row.get("aliases", "")):
        a = " ".join(str(a or "").split())
        n = normalize_text(a)
        if not a or n in GENERIC_ALIAS or len(n) < 3:
            continue
        if n not in {normalize_text(x) for x in result}:
            result.append(a)
    return " | ".join(result[:8])


def ocr_for(row: dict) -> str:
    parts = [row.get("brand"), row.get("product_family"), row.get("form"), row.get("strength"), row.get("size"), row.get("use_case"), row.get("skin_type"), row.get("name")]
    words = []
    for p in parts + split_tokens(row.get("ocr_keywords", "")) + split_tokens(row.get("image_ocr_keywords", "")):
        for token in str(p or "").replace("|", " ").split():
            if len(token) >= 2 and normalize_text(token) not in {normalize_text(x) for x in words}:
                words.append(token)
    return " ".join(words[:40])


def read_input(path: Path) -> List[dict]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))}
        if any(str(v or "").strip() for v in d.values()):
            rows.append(d)
    return rows


def map_row(raw: dict, idx: int) -> dict:
    def first(*keys):
        for k in keys:
            if k in raw and str(raw.get(k) or "").strip():
                return str(raw.get(k)).strip()
        return ""
    name = first("name", "product", "product_name", "اسم المنتج", "الصنف", "original_name")
    brand = first("brand", "company", "الشركة", "الماركة")
    form = first("form", "type", "الشكل", "النوع") or guess_form(name)
    active = first("active_ingredient", "المادة الفعالة")
    st = first("strength", "dose", "concentration") or strength(name)
    sz = first("size", "pack", "volume") or size(name)
    category = first("category") or guess_category(name, form, active)
    family = first("product_family") or guess_family(name, brand, form, st, sz)
    out = {k: "" for k in REQUIRED_SCHEMA}
    out.update({
        "product_id": first("product_id", "id") or f"AUTO-{idx:06d}",
        "name": name,
        "normalized_name": normalize_text(name),
        "brand": brand,
        "company": first("company") or brand,
        "product_family": family,
        "category": category,
        "active_ingredient": active,
        "form": form,
        "strength": st,
        "size": sz,
        "pack": first("pack"),
        "barcode": first("barcode"),
        "use_case": first("use_case"),
        "skin_type": first("skin_type"),
        "available": first("available") or "متوفر",
        "price": first("price", "السعر"),
        "currency": first("currency") or "LYD",
        "merchant_id": first("merchant_id") or "default",
        "review_status": first("review_status"),
        "review_notes": first("review_notes"),
    })
    out["aliases"] = aliases_for({**out, "aliases": first("aliases")})
    out["ocr_keywords"] = ocr_for({**out, "ocr_keywords": first("ocr_keywords"), "image_ocr_keywords": first("image_ocr_keywords")})
    if category == "unknown" or not form or not brand or (category == "medicine" and not st):
        out["review_status"] = "needs_review"
        reasons = []
        if category == "unknown": reasons.append("category unclear")
        if not brand: reasons.append("brand missing")
        if not form: reasons.append("form/type unclear")
        if category == "medicine" and not st: reasons.append("strength missing")
        out["review_notes"] = "; ".join(reasons)
    else:
        out["review_status"] = out["review_status"] or "approved"
    return out


def write_xlsx(path: Path, rows: List[dict], headers: List[str]) -> None:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = path.stem[:31]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(42, max(12, max(len(str(c.value or "")) for c in col[:200]) + 2))
    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("--out-dir", default=".")
    args = ap.parse_args()
    out_dir = Path(args.out_dir); out_dir.mkdir(parents=True, exist_ok=True)
    raw = read_input(Path(args.input))
    enriched = [map_row(r, i+1) for i, r in enumerate(raw)]
    report = analyze_products(enriched)
    ready = [r for r in enriched if r.get("review_status") != "needs_review" and not r.get("quality_issues")]
    needs = [r for r in enriched if r.get("review_status") == "needs_review" or r.get("quality_issues")]
    write_xlsx(out_dir / "products_ready_for_upload.xlsx", ready, REQUIRED_SCHEMA)
    write_xlsx(out_dir / "products_needs_review.xlsx", needs, REQUIRED_SCHEMA + ["quality_issues"])
    quality_rows = []
    for issue, count in sorted(report.get("issue_counts", {}).items()):
        quality_rows.append({"metric": issue, "count": count, "decision": report.get("decision"), "reason": " | ".join(report.get("reasons", []))})
    write_xlsx(out_dir / "catalog_quality_report.xlsx", quality_rows, ["metric", "count", "decision", "reason"])
    tests = [{"query": r.get("name"), "expected_decision": "EXACT_MATCH" if r.get("review_status") == "approved" else "ASK_CLARIFICATION", "expected_product_id": r.get("product_id"), "forbidden_product_id": "", "must_not_include_price": "false", "notes": "suggested by catalog_intelligence_v5"} for r in ready[:200]]
    write_xlsx(out_dir / "golden_text_tests_suggested.xlsx", tests, ["query", "expected_decision", "expected_product_id", "forbidden_product_id", "must_not_include_price", "notes"])
    ocr = [{"product_id": r.get("product_id"), "name": r.get("name"), "ocr_keywords": r.get("ocr_keywords"), "review_notes": r.get("review_notes")} for r in enriched]
    write_xlsx(out_dir / "ocr_keywords_suggested.xlsx", ocr, ["product_id", "name", "ocr_keywords", "review_notes"])
    print(f"CATALOG_INTELLIGENCE_V5_OK decision={report.get('decision')} total={len(enriched)} review={len(needs)} out={out_dir}")

if __name__ == "__main__":
    main()
