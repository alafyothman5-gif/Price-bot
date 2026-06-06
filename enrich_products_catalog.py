#!/usr/bin/env python3
"""Enrich a pharmacy product Excel/CSV into a safer PriceBot catalog.

The tool is deterministic. It extracts explicit evidence from the product name,
source columns, and local dictionaries. It does not invent medical facts, stock,
barcodes, or substitution groups. Anything uncertain is separated into the
review workbook with notes.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except Exception as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required. Run: pip install -r requirements.txt") from exc

OUTPUT_HEADERS = [
    "product_id", "name", "brand", "product_family", "category", "active_ingredient", "form", "strength", "size", "pack", "barcode",
    "aliases", "ocr_keywords", "use_case", "skin_type", "available", "price", "substitution_group_id", "review_status", "review_notes",
]

HEADER_CANDIDATES = {
    "name": {"name", "product", "item", "product_name", "product name", "item_name", "اسم الصنف", "اسم المنتج", "المنتج", "الصنف", "الاسم", "البيان"},
    "price": {"price", "السعر", "سعر", "sale_price", "selling_price", "retail_price", "سعر البيع", "final_price"},
    "quantity": {"quantity", "stock", "qty", "qoh", "available_quantity", "الكمية", "الرصيد", "المخزون", "العدد"},
    "available": {"available", "availability", "status", "الحالة", "التوفر", "توفر"},
    "barcode": {"barcode", "bar_code", "باركود", "الباركود"},
    "brand": {"brand", "company", "الشركة", "الماركة", "البراند", "المصنع", "الوكيل"},
    "category": {"category", "تصنيف", "الفئة", "القسم"},
    "form": {"form", "form_or_type", "type", "الشكل الدوائي", "الشكل", "النوع", "شكل"},
    "active_ingredient": {"active_ingredient", "active ingredient", "المادة الفعالة", "المادة"},
    "strength": {"strength", "concentration", "dose", "تركيز", "جرعة", "التركيز"},
    "size": {"size", "volume", "حجم", "الحجم"},
    "pack": {"pack", "package", "عبوة", "العبوة", "pack_size"},
    "aliases": {"aliases", "alias", "اسماء بديلة", "اسم بديل"},
    "ocr_keywords": {"ocr_keywords", "image_ocr_keywords", "keywords", "كلمات", "كلمات البحث"},
    "product_family": {"product_family", "family", "عائلة المنتج"},
    "use_case": {"use_case", "use", "purpose", "الاستخدام"},
    "skin_type": {"skin_type", "نوع البشرة"},
    "substitution_group_id": {"substitution_group_id", "substitution_group", "بدائل"},
}

ARABIC_MAP = {
    "أ": "ا", "إ": "ا", "آ": "ا", "ة": "ه", "ى": "ي", "ؤ": "و", "ئ": "ي", "ٱ": "ا",
    "ڤ": "ف", "ک": "ك", "ی": "ي", "گ": "ك", "چ": "ج", "پ": "ب",
}
ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹", "01234567890123456789")

MEDICINE_FORMS = {"tablet", "capsule", "syrup", "suspension", "injection", "drops", "suppository", "spray", "solution", "ointment"}
COSMETIC_FORMS = {"cleanser", "lotion", "serum", "cream", "gel", "shampoo", "conditioner", "sunscreen", "moisturizer", "toner", "balm", "mask", "soap", "foam", "oil"}
OVERLAP_FORMS = {"cream", "gel", "spray", "solution", "ointment"}
WEAK_KEYWORDS = {
    "face", "skin", "body", "hair", "cream", "gel", "lotion", "serum", "cleanser", "shampoo", "sunscreen", "oil", "soap", "foam",
    "كريم", "جل", "غسول", "شامبو", "لوشن", "سيروم", "بشرة", "وجه",
}
REQUEST_WORDS = {"متوفر", "السعر", "سعر", "كم", "عندكم", "please", "available", "price"}

STRENGTH_PATTERNS = [
    re.compile(r"(?P<full>(?P<n>\d+(?:[\.,]\d+)?)\s*(?P<u>mg|mcg|g|iu|units?|unit)\s*/\s*(?P<d>\d+(?:[\.,]\d+)?)\s*(?P<du>ml|mL|ML))", re.I),
    re.compile(r"(?P<full>(?P<n>\d+(?:[\.,]\d+)?)\s*(?P<u>mg|mcg|iu|units?|unit)\b)", re.I),
    re.compile(r"(?P<full>(?P<n>\d+(?:[\.,]\d+)?)\s*%)", re.I),
    re.compile(r"(?P<full>(?P<n>\d+(?:[\.,]\d+)?)\s*g\b)", re.I),
]
SIZE_PATTERN = re.compile(r"(?P<full>(?P<n>\d+(?:[\.,]\d+)?)\s*(?P<u>ml|mL|ML|l|L|g|gm|kg)\b)", re.I)
PACK_PATTERN = re.compile(r"(?P<full>(?P<n>\d+)\s*(?P<u>tabs?|tablets?|caps?|capsules?|ampoules?|amps?|amp|sachets?|strips?|suppositories?|suppository|bottles?|vials?)\b)", re.I)
UNCLEAR_STRENGTH_PATTERN = re.compile(r"(?<![A-Za-z0-9])(?P<n>\d{2,4})(?![A-Za-z0-9])")


def _dedupe_keep_order(values: Iterable[str]) -> List[str]:
    seen = set()
    out = []
    for value in values:
        v = str(value or "").strip()
        if not v:
            continue
        key = normalize_text(v)
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(v)
    return out


def normalize_text(value: Any) -> str:
    text = str(value or "").strip().translate(ARABIC_DIGITS)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    for src, dst in ARABIC_MAP.items():
        text = text.replace(src, dst)
    text = re.sub(r"[\u064b-\u065f]", "", text)
    text = text.replace("ـ", "")
    text = text.lower()
    text = re.sub(r"(?<=\d)\s+(?=(mg|mcg|g|ml|l|iu|%))", "", text, flags=re.I)
    text = re.sub(r"[^\w\s%/\.\+\-]+", " ", text, flags=re.UNICODE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_display_name(value: Any) -> str:
    text = str(value or "").strip().translate(ARABIC_DIGITS)
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"(?i)(\d+(?:[\.,]\d+)?)\s*(ml|mL|ML)\b", lambda m: f"{m.group(1).replace(',', '.')}ml", text)
    text = re.sub(r"(?i)(\d+(?:[\.,]\d+)?)\s*(mg|mcg|iu|g)\b", lambda m: f"{m.group(1).replace(',', '.')}{m.group(2).lower()}", text)
    text = re.sub(r"(?i)(\d+(?:[\.,]\d+)?)\s*%", lambda m: f"{m.group(1).replace(',', '.')}%", text)
    return text.strip()


def contains_phrase(norm_text: str, phrase: str) -> bool:
    nt = f" {normalize_text(norm_text)} "
    ph = normalize_text(phrase)
    return bool(ph and f" {ph} " in nt)


def load_json(path: Path, default: Dict[str, Any]) -> Dict[str, Any]:
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8") as fh:
        loaded = json.load(fh)
    return loaded if isinstance(loaded, dict) else default


def canonical_header(value: Any) -> str:
    raw = str(value or "").strip()
    norm = normalize_text(raw).replace("_", " ")
    low = raw.lower().strip()
    for canon, names in HEADER_CANDIDATES.items():
        normalized_names = {normalize_text(n).replace("_", " ") for n in names}
        lowered_names = {n.lower() for n in names}
        if norm in normalized_names or low in lowered_names:
            return canon
    return re.sub(r"\W+", "_", low).strip("_") or "unnamed"


def unique_headers(headers: Sequence[Any]) -> List[str]:
    seen: Dict[str, int] = {}
    out = []
    for h in headers:
        mapped = canonical_header(h)
        n = seen.get(mapped, 0)
        seen[mapped] = n + 1
        out.append(mapped if n == 0 else f"{mapped}__{n}")
    return out


def read_xlsx(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    best_rows: List[Tuple[Any, ...]] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > len(best_rows):
            best_rows = rows
        for idx, row in enumerate(rows[:20]):
            headers = unique_headers(row)
            if "name" in headers:
                return [dict(zip(headers, values)) for values in rows[idx + 1:] if any(str(c or "").strip() for c in values)]
    if not best_rows:
        return []
    headers = unique_headers(best_rows[0])
    if "name" not in headers:
        raise ValueError("No product name column found. Supported names include: name/product/item/product_name/اسم الصنف.")
    return [dict(zip(headers, values)) for values in best_rows[1:] if any(str(c or "").strip() for c in values)]


def read_xls(path: Path) -> List[Dict[str, Any]]:
    try:
        import xlrd  # type: ignore
    except Exception as exc:
        raise ValueError(".xls input requires xlrd. Run: pip install xlrd==2.0.1 or save the file as .xlsx.") from exc
    book = xlrd.open_workbook(str(path))
    best = None
    for sheet in book.sheets():
        rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        if best is None or len(rows) > len(best):
            best = rows
        for idx, row in enumerate(rows[:20]):
            headers = unique_headers(row)
            if "name" in headers:
                return [dict(zip(headers, values)) for values in rows[idx + 1:] if any(str(c or "").strip() for c in values)]
    if not best:
        return []
    headers = unique_headers(best[0])
    if "name" not in headers:
        raise ValueError("No product name column found in .xls file.")
    return [dict(zip(headers, values)) for values in best[1:] if any(str(c or "").strip() for c in values)]


def read_csv(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        rows = list(reader)
    if not rows:
        return []
    headers = unique_headers(rows[0])
    if "name" not in headers:
        raise ValueError("No product name column found in CSV file.")
    return [dict(zip(headers, row)) for row in rows[1:] if any(str(c or "").strip() for c in row)]


def read_table(path: Path) -> List[Dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    if suffix == ".xls":
        return read_xls(path)
    if suffix == ".csv":
        return read_csv(path)
    raise ValueError(f"Unsupported input format: {suffix}. Use .xlsx, .xlsm, .xls, or .csv")


@dataclass
class Dictionaries:
    brands: Dict[str, Any]
    forms: Dict[str, List[str]]
    active_ingredients: Dict[str, str]
    use_cases: Dict[str, str]
    skin_types: Dict[str, str]
    brand_aliases: List[Tuple[str, str, str]] = field(default_factory=list)  # normalized alias, canonical, display alias
    form_aliases: List[Tuple[str, str, str]] = field(default_factory=list)   # normalized alias, canonical, display alias

    @classmethod
    def load(cls, data_dir: Path) -> "Dictionaries":
        brands = load_json(data_dir / "brands_dictionary.json", {})
        forms = load_json(data_dir / "forms_dictionary.json", {})
        active = load_json(data_dir / "active_ingredients_dictionary.json", {})
        use_cases = load_json(data_dir / "use_case_dictionary.json", {})
        skin_types = load_json(data_dir / "skin_type_dictionary.json", {})
        obj = cls(brands=brands, forms=forms, active_ingredients=active, use_cases=use_cases, skin_types=skin_types)
        obj.brand_aliases = obj._build_brand_aliases()
        obj.form_aliases = obj._build_form_aliases()
        return obj

    def _build_brand_aliases(self) -> List[Tuple[str, str, str]]:
        aliases = []
        for key, value in self.brands.items():
            if isinstance(value, dict):
                canonical = str(value.get("canonical") or key).strip()
                raw_aliases = [key, canonical] + list(value.get("aliases") or [])
            else:
                canonical = str(value or key).strip()
                raw_aliases = [key, canonical]
            for alias in raw_aliases:
                norm = normalize_text(alias)
                if norm:
                    aliases.append((norm, canonical, str(alias)))
        aliases = _dedupe_tuple_aliases(aliases)
        return sorted(aliases, key=lambda t: len(t[0]), reverse=True)

    def _build_form_aliases(self) -> List[Tuple[str, str, str]]:
        aliases = []
        for canonical, raw_aliases in self.forms.items():
            for alias in [canonical] + list(raw_aliases or []):
                norm = normalize_text(alias)
                if norm:
                    aliases.append((norm, str(canonical), str(alias)))
        aliases = _dedupe_tuple_aliases(aliases)
        return sorted(aliases, key=lambda t: len(t[0]), reverse=True)


def _dedupe_tuple_aliases(items: Iterable[Tuple[str, str, str]]) -> List[Tuple[str, str, str]]:
    seen = set()
    out = []
    for item in items:
        key = (item[0], item[1])
        if key not in seen:
            seen.add(key)
            out.append(item)
    return out


def collect(row: Dict[str, Any], key: str) -> str:
    values = []
    for k, v in row.items():
        if k == key or str(k).startswith(f"{key}__"):
            text = "" if v is None else str(v).strip()
            if text and text.lower() not in {"none", "nan", "null"}:
                values.append(text)
    return " | ".join(values).strip()


def detect_brand(norm_name: str, raw_brand: str, d: Dictionaries) -> Tuple[str, List[str]]:
    notes: List[str] = []
    if raw_brand:
        raw_norm = normalize_text(raw_brand)
        for alias_norm, canonical, _display_alias in d.brand_aliases:
            if raw_norm == alias_norm or contains_phrase(raw_norm, alias_norm):
                return canonical, notes
        return clean_display_name(raw_brand), notes
    found = []
    for alias_norm, canonical, _display_alias in d.brand_aliases:
        if contains_phrase(norm_name, alias_norm):
            found.append(canonical)
    uniq = _dedupe_keep_order(found)
    if len(uniq) > 1:
        notes.append("multiple possible brands: " + ", ".join(uniq))
    return (uniq[0] if len(uniq) == 1 else ""), notes


def detect_forms(norm_name: str, raw_form: str, d: Dictionaries) -> Tuple[str, List[str]]:
    notes: List[str] = []
    source = normalize_text(raw_form) if raw_form else norm_name
    found = []
    for alias_norm, canonical, _display_alias in d.form_aliases:
        if contains_phrase(source, alias_norm):
            found.append(canonical)
    uniq = _dedupe_keep_order(found)
    # If a specific phrase maps to cleanser, ignore the generic gel inside "cleansing gel".
    if "cleanser" in uniq and "gel" in uniq:
        uniq = [x for x in uniq if x != "gel"]
    if "sunscreen" in uniq and "cream" in uniq:
        uniq = [x for x in uniq if x != "cream"]
    if len(uniq) > 1:
        notes.append("multiple possible forms: " + ", ".join(uniq))
    return (uniq[0] if len(uniq) == 1 else ""), notes


def extract_with_spans(patterns: Sequence[re.Pattern], text: str) -> Tuple[List[str], List[Tuple[int, int]]]:
    values: List[str] = []
    spans: List[Tuple[int, int]] = []
    occupied: List[Tuple[int, int]] = []
    for pattern in patterns:
        for m in pattern.finditer(text):
            span = m.span("full")
            if any(not (span[1] <= s[0] or span[0] >= s[1]) for s in occupied):
                continue
            raw = m.group("full")
            values.append(format_unit(raw))
            spans.append(span)
            occupied.append(span)
    return _dedupe_keep_order(values), spans


def format_unit(value: str) -> str:
    v = str(value or "").strip().replace(" ", "")
    v = v.replace(",", ".")
    v = re.sub(r"(?i)ML", "ml", v)
    v = re.sub(r"(?i)MG", "mg", v)
    v = re.sub(r"(?i)MCG", "mcg", v)
    v = re.sub(r"(?i)IU", "iu", v)
    v = re.sub(r"(?i)UNITS?", "unit", v)
    v = re.sub(r"(?i)GM", "g", v)
    return v


def extract_strength_and_size(name: str, category_hint: str) -> Tuple[str, str, str, List[str], List[Tuple[int, int]]]:
    notes: List[str] = []
    strength_values, strength_spans = extract_with_spans(STRENGTH_PATTERNS, name)
    size_values = []
    for m in SIZE_PATTERN.finditer(name):
        span = m.span("full")
        if any(not (span[1] <= s[0] or span[0] >= s[1]) for s in strength_spans):
            continue
        value = format_unit(m.group("full"))
        # In a medicine name, g without packaging context is more likely strength; STRENGTH_PATTERNS already caught it.
        if value.lower().endswith("g") and category_hint == "medicine":
            continue
        size_values.append(value)
    size_values = _dedupe_keep_order(size_values)
    unclear = ""
    if not strength_values and category_hint == "medicine":
        for m in UNCLEAR_STRENGTH_PATTERN.finditer(name):
            num = m.group("n")
            if num and num not in {re.sub(r"\D+", "", x) for x in size_values}:
                unclear = num
                notes.append("strength unit unclear")
                break
    strength = strength_values[0] if strength_values else unclear
    size = size_values[0] if size_values else ""
    return strength, size, " | ".join(notes), notes, strength_spans


def extract_pack(name: str) -> str:
    for m in PACK_PATTERN.finditer(name):
        n = m.group("n")
        u = m.group("u").lower()
        if u.startswith("tab"):
            unit = "tablets"
        elif u.startswith("cap"):
            unit = "capsules"
        elif u.startswith("amp"):
            unit = "ampoules"
        elif u.startswith("sachet"):
            unit = "sachets"
        elif u.startswith("strip"):
            unit = "strips"
        elif u.startswith("supp"):
            unit = "suppositories"
        elif u.startswith("bottle"):
            unit = "bottles"
        elif u.startswith("vial"):
            unit = "vials"
        else:
            unit = u
        return f"{n} {unit}"
    return ""


def detect_active_ingredient(norm_name: str, brand: str, raw_active: str, d: Dictionaries) -> Tuple[str, List[str]]:
    if raw_active:
        return normalize_active(raw_active), []
    lookup_parts = [normalize_text(brand), norm_name]
    for key, value in sorted(d.active_ingredients.items(), key=lambda kv: len(kv[0]), reverse=True):
        k = normalize_text(key)
        if k and any(contains_phrase(part, k) for part in lookup_parts):
            return str(value).strip().lower(), []
    return "", []


def normalize_active(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def detect_use_case(norm_name: str, raw_use_case: str, d: Dictionaries) -> str:
    if raw_use_case:
        return normalize_text(raw_use_case).replace(" ", "_")
    for phrase, value in sorted(d.use_cases.items(), key=lambda kv: len(kv[0]), reverse=True):
        if contains_phrase(norm_name, phrase):
            return str(value).strip()
    return ""


def detect_skin_type(norm_name: str, raw_skin_type: str, d: Dictionaries) -> str:
    if raw_skin_type:
        return normalize_text(raw_skin_type).replace(" ", "_")
    for phrase, value in sorted(d.skin_types.items(), key=lambda kv: len(kv[0]), reverse=True):
        if contains_phrase(norm_name, phrase):
            return str(value).strip()
    return ""


def classify_category(norm_name: str, brand: str, form: str, active: str, raw_category: str, use_case: str, skin_type: str) -> Tuple[str, List[str]]:
    notes: List[str] = []
    raw = normalize_text(raw_category)
    if raw in {"medicine", "دواء", "ادويه", "med", "drug"}:
        return "medicine", notes
    if raw in {"cosmetic", "cosmetics", "كوزمتك", "تجميل", "skin care", "skincare"}:
        return "cosmetic", notes
    has_strength = bool(extract_with_spans(STRENGTH_PATTERNS, norm_name)[0])
    medicine_signal = bool(active or form in MEDICINE_FORMS or has_strength or contains_phrase(norm_name, "antibiotic") or contains_phrase(norm_name, "painkiller"))
    cosmetic_signal = bool(form in COSMETIC_FORMS or use_case or skin_type or contains_phrase(norm_name, "spf") or contains_phrase(norm_name, "sunscreen"))
    if form in OVERLAP_FORMS and not active and not has_strength and not use_case and not skin_type:
        notes.append("category uncertain for overlapping form")
        return "other", notes
    if medicine_signal and not cosmetic_signal:
        return "medicine", notes
    if cosmetic_signal and not medicine_signal:
        return "cosmetic", notes
    if medicine_signal and cosmetic_signal:
        # Strength/active ingredient wins for medicinal creams/gels, otherwise review.
        if active or has_strength:
            return "medicine", notes
        notes.append("category uncertain: medicine/cosmetic signals")
        return "other", notes
    notes.append("category uncertain")
    return "other", notes


def remove_phrases(text: str, phrases: Iterable[str]) -> str:
    out = f" {text} "
    for phrase in sorted({p for p in phrases if p}, key=len, reverse=True):
        out = re.sub(rf"(?i)(?<!\w){re.escape(phrase)}(?!\w)", " ", out)
    out = re.sub(r"\s+", " ", out).strip(" -_/|,.\t")
    return out


def infer_product_family(name: str, brand: str, form: str, category: str, strength: str, size: str, pack: str, d: Dictionaries, raw_family: str = "") -> str:
    if raw_family:
        return clean_display_name(raw_family)
    family = clean_display_name(name)
    phrases: List[str] = []
    if brand:
        phrases.append(brand)
        for alias_norm, canonical, display_alias in d.brand_aliases:
            if canonical == brand:
                phrases.append(display_alias)
    if strength:
        phrases.append(strength)
    if size:
        phrases.append(size)
    if pack:
        phrases.append(pack)
    # Medicine family should not include dosage form. Cosmetic family often keeps cleanser/shampoo/sunscreen as product-line evidence.
    if category == "medicine" and form:
        phrases += [display for _norm, canon, display in d.form_aliases if canon == form]
    elif category == "cosmetic" and form in {"balm", "cream", "lotion", "serum", "gel", "moisturizer", "toner", "mask", "oil"}:
        phrases += [display for _norm, canon, display in d.form_aliases if canon == form]
    # Remove unit variants visible in the raw name.
    phrases += [p for p in re.findall(r"\d+(?:[\.,]\d+)?\s*(?:mg|mcg|g|ml|iu|%)\b", family, flags=re.I)]
    family = remove_phrases(family, phrases)
    family = re.sub(r"\s+", " ", family).strip(" -_/|,.\t")
    if not family and category == "medicine" and brand:
        family = brand
    return family


def clean_price(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    raw = raw.translate(ARABIC_DIGITS)
    match = re.findall(r"\d+(?:[\.,]\d+)*", raw)
    if not match:
        return ""
    val = match[-1]
    if "," in val and "." in val:
        val = val.replace(",", "")
    elif "," in val:
        if re.match(r"^\d+,\d{1,2}$", val):
            val = val.replace(",", ".")
        else:
            val = val.replace(",", "")
    try:
        num = float(val)
        return str(int(num)) if num.is_integer() else (f"{num:.4f}".rstrip("0").rstrip("."))
    except Exception:
        return val


def parse_quantity(value: str) -> Optional[float]:
    raw = str(value or "").strip().translate(ARABIC_DIGITS)
    if not raw:
        return None
    m = re.search(r"-?\d+(?:[\.,]\d+)?", raw)
    if not m:
        return None
    try:
        return float(m.group(0).replace(",", "."))
    except Exception:
        return None


def derive_available(row: Dict[str, Any], default_available: str) -> str:
    qty_text = collect(row, "quantity")
    qty = parse_quantity(qty_text)
    if qty is not None:
        return "true" if qty > 0 else "false"
    raw = normalize_text(collect(row, "available"))
    if raw:
        if any(x in raw for x in ["غير", "out", "no", "false", "نفد", "ناقص", "0"]):
            return "false"
        if any(x in raw for x in ["متوفر", "available", "yes", "true", "in stock"]):
            return "true"
    if default_available in {"true", "false"}:
        return default_available
    return ""


def safe_alias(alias: str, brand: str, family: str, form: str, strength: str, active: str) -> bool:
    norm = normalize_text(alias)
    if not norm:
        return False
    if norm in WEAK_KEYWORDS:
        return False
    brand_norm = normalize_text(brand)
    fam_tokens = set(normalize_text(family).split()) - WEAK_KEYWORDS
    has_identity = bool(brand_norm and contains_phrase(norm, brand_norm)) or bool(active and contains_phrase(norm, active))
    has_specific = bool(fam_tokens & set(norm.split())) or bool(form and contains_phrase(norm, form)) or bool(strength and contains_phrase(norm, strength))
    return has_identity and has_specific


def build_aliases(name: str, brand: str, family: str, form: str, strength: str, active: str, existing_aliases: str, d: Dictionaries) -> str:
    items: List[str] = []
    items.extend(re.split(r"[,،|;\n]+", existing_aliases or ""))
    if brand and family:
        items.append(f"{brand} {family}")
    if brand and family and form and form not in normalize_text(family):
        items.append(f"{brand} {family} {form}")
    if brand and form and strength:
        items.append(f"{brand} {form} {strength}")
    if active and form:
        items.append(f"{active} {form}")
    if form == "syrup" and brand:
        items.append(f"{brand} suspension")
    # Add Arabic brand + form aliases only when both identify the product.
    brand_ar = []
    for _norm, canonical, display in d.brand_aliases:
        if canonical == brand and re.search(r"[\u0600-\u06FF]", display):
            brand_ar.append(display)
    form_ar = []
    for _norm, canonical, display in d.form_aliases:
        if canonical == form and re.search(r"[\u0600-\u06FF]", display):
            form_ar.append(display)
    for b in brand_ar[:2]:
        for f in form_ar[:2]:
            items.append(f"{b} {f}")
            items.append(f"{f} {b}")
    # Include the cleaned full name as an alias if it is not brand-only/type-only.
    items.insert(0, name)
    good = [a.strip().lower() for a in items if safe_alias(a, brand, family, form, strength, active)]
    return "|".join(_dedupe_keep_order(good))


def build_ocr_keywords(name: str, brand: str, family: str, form: str, strength: str, size: str, active: str, existing_keywords: str) -> str:
    parts: List[str] = []
    parts.extend(re.split(r"[,،|;\n]+", existing_keywords or ""))
    if brand:
        parts.append(brand)
    for tok in normalize_text(family).split():
        if tok and tok not in WEAK_KEYWORDS and tok not in REQUEST_WORDS:
            parts.append(tok)
    if form:
        parts.append(form)
    if strength:
        parts.append(strength)
    if size:
        parts.append(size)
    if active:
        parts.append(active)
    # Keep only useful package/product evidence.
    cleaned = []
    for p in parts:
        p = normalize_text(p)
        if not p or p in REQUEST_WORDS:
            continue
        if p in WEAK_KEYWORDS and not (brand or family):
            continue
        cleaned.append(p)
    return "|".join(_dedupe_keep_order(cleaned))


def enrich_one(row: Dict[str, Any], idx: int, d: Dictionaries, default_available: str) -> Dict[str, str]:
    notes: List[str] = []
    source_name = collect(row, "name")
    name = clean_display_name(source_name)
    if not name:
        return {h: "" for h in OUTPUT_HEADERS} | {"product_id": f"P{idx:06d}", "review_status": "unsafe", "review_notes": "name missing"}
    norm_name = normalize_text(name)
    brand, brand_notes = detect_brand(norm_name, collect(row, "brand"), d)
    notes += brand_notes
    form, form_notes = detect_forms(norm_name, collect(row, "form"), d)
    notes += form_notes
    active, active_notes = detect_active_ingredient(norm_name, brand, collect(row, "active_ingredient"), d)
    notes += active_notes
    use_case = detect_use_case(norm_name, collect(row, "use_case"), d)
    skin_type = detect_skin_type(norm_name, collect(row, "skin_type"), d)
    category, category_notes = classify_category(norm_name, brand, form, active, collect(row, "category"), use_case, skin_type)
    notes += category_notes
    strength, size, _unclear_strength_note, strength_notes, _strength_spans = extract_strength_and_size(name, category)
    notes += strength_notes
    if collect(row, "strength"):
        strength = format_unit(collect(row, "strength"))
    if collect(row, "size"):
        size = format_unit(collect(row, "size"))
    pack = clean_display_name(collect(row, "pack")) or extract_pack(name)
    family = infer_product_family(name, brand, form, category, strength, size, pack, d, collect(row, "product_family"))
    price = clean_price(collect(row, "price"))
    available = derive_available(row, default_available)
    barcode = collect(row, "barcode")
    existing_aliases = collect(row, "aliases")
    existing_keywords = collect(row, "ocr_keywords")
    aliases = build_aliases(name, brand, family, form, strength, active, existing_aliases, d)
    ocr_keywords = build_ocr_keywords(name, brand, family, form, strength, size, active, existing_keywords)
    if not brand:
        notes.append("brand not detected")
    if category == "other":
        notes.append("category uncertain")
    if category in {"medicine", "cosmetic"} and not form:
        notes.append("form unclear")
    if category == "medicine" and not active:
        notes.append("active ingredient missing")
    if category == "cosmetic" and not form:
        notes.append("cosmetic type unclear")
    if category == "cosmetic" and form in OVERLAP_FORMS and not (use_case or skin_type or family):
        notes.append("cosmetic purpose unclear")
    if not family:
        notes.append("product family unclear")
    if not price:
        notes.append("price missing")
    if not aliases:
        notes.append("safe aliases missing")
    if not ocr_keywords or len(ocr_keywords.split("|")) < 2:
        notes.append("strong OCR keywords missing")
    product_id = clean_display_name(collect(row, "product_id")) or f"P{idx:06d}"
    status = "ready" if not notes else "needs_review"
    return {
        "product_id": product_id,
        "name": name,
        "brand": brand,
        "product_family": family,
        "category": category,
        "active_ingredient": active,
        "form": form,
        "strength": strength,
        "size": size,
        "pack": pack,
        "barcode": barcode,
        "aliases": aliases,
        "ocr_keywords": ocr_keywords,
        "use_case": use_case,
        "skin_type": skin_type,
        "available": available,
        "price": price,
        "substitution_group_id": clean_display_name(collect(row, "substitution_group_id")),
        "review_status": status,
        "review_notes": "; ".join(_dedupe_keep_order(notes)),
    }


def add_review(row: Dict[str, str], note: str, unsafe: bool = False) -> None:
    notes = _dedupe_keep_order((row.get("review_notes", "") + "; " + note).split(";"))
    row["review_notes"] = "; ".join(notes)
    if unsafe:
        row["review_status"] = "unsafe"
    elif row.get("review_status") == "ready":
        row["review_status"] = "needs_review"


def second_pass(rows: List[Dict[str, str]]) -> Dict[str, Any]:
    duplicate_candidates = 0
    norm_name_groups: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    signature_groups: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    med_family_groups: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    med_strength_groups: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    cos_size_groups: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in rows:
        norm_name = normalize_text(row.get("name"))
        norm_name_groups[norm_name].append(row)
        sig = "|".join(normalize_text(row.get(k)) for k in ["brand", "product_family", "form", "strength", "size"])
        signature_groups[sig].append(row)
        if row.get("category") == "medicine":
            fam_key = "|".join(normalize_text(row.get(k)) for k in ["brand", "product_family"])
            med_family_groups[fam_key].append(row)
            strength_key = "|".join(normalize_text(row.get(k)) for k in ["brand", "product_family", "form"])
            med_strength_groups[strength_key].append(row)
        if row.get("category") == "cosmetic":
            cos_key = "|".join(normalize_text(row.get(k)) for k in ["brand", "product_family", "form"])
            cos_size_groups[cos_key].append(row)
    for groups in [norm_name_groups, signature_groups]:
        for key, items in groups.items():
            if key and len(items) > 1:
                duplicate_candidates += len(items)
                for item in items:
                    add_review(item, "possible duplicate")
    for _key, items in med_family_groups.items():
        forms = {x.get("form") for x in items if x.get("form")}
        if len(forms) > 1:
            for item in items:
                if not item.get("form"):
                    add_review(item, "multiple possible forms")
    for _key, items in med_strength_groups.items():
        strengths = {x.get("strength") for x in items if x.get("strength")}
        if len(strengths) > 1:
            for item in items:
                if not item.get("strength"):
                    add_review(item, "multiple possible strengths")
    for _key, items in cos_size_groups.items():
        sizes = {x.get("size") for x in items if x.get("size")}
        if len(sizes) > 1:
            for item in items:
                if not item.get("size"):
                    add_review(item, "multiple possible sizes")
    return {"duplicate_candidates_count": duplicate_candidates}


def build_substitution_suggestions(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    groups: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in rows:
        if row.get("category") != "medicine":
            continue
        active = row.get("active_ingredient") or ""
        form = row.get("form") or ""
        strength = row.get("strength") or ""
        if active and form and strength:
            key = "|".join([normalize_text(active), normalize_text(form), normalize_text(strength)])
            groups[key].append(row)
    out = []
    n = 1
    for _key, items in sorted(groups.items()):
        if len(items) < 2:
            continue
        gid = f"SUG_MED_{n:04d}"
        n += 1
        for item in items:
            out.append({
                "suggested_substitution_group_id": gid,
                "product_id": item.get("product_id", ""),
                "name": item.get("name", ""),
                "active_ingredient": item.get("active_ingredient", ""),
                "form": item.get("form", ""),
                "strength": item.get("strength", ""),
                "review_required": "pharmacist approval required before use",
            })
    return out


def write_workbook(path: Path, rows: List[Dict[str, str]], headers: Sequence[str] = OUTPUT_HEADERS) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "A": 12, "B": 36, "C": 18, "D": 28, "E": 14, "F": 28, "G": 16, "H": 18, "I": 14, "J": 16, "K": 18,
        "L": 44, "M": 44, "N": 18, "O": 18, "P": 12, "Q": 12, "R": 24, "S": 16, "T": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_report(path: Path, rows: List[Dict[str, str]], extra: Dict[str, Any], assumptions: Dict[str, Any]) -> Dict[str, Any]:
    status = Counter(row.get("review_status") or "" for row in rows)
    report = {
        "total_products": len(rows),
        "ready_count": status.get("ready", 0),
        "needs_review_count": status.get("needs_review", 0),
        "unsafe_count": status.get("unsafe", 0),
        "missing_brand_count": sum(1 for row in rows if not row.get("brand")),
        "missing_form_count": sum(1 for row in rows if row.get("category") in {"medicine", "cosmetic"} and not row.get("form")),
        "missing_strength_count": sum(1 for row in rows if row.get("category") == "medicine" and not row.get("strength")),
        "duplicate_candidates_count": extra.get("duplicate_candidates_count", 0),
        "assumptions": assumptions,
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def enrich_catalog(input_path: Path, data_dir: Path, default_available: str = "blank") -> Tuple[List[Dict[str, str]], Dict[str, Any]]:
    d = Dictionaries.load(data_dir)
    source_rows = read_table(input_path)
    enriched = [enrich_one(row, i + 1, d, default_available) for i, row in enumerate(source_rows)]
    extra = second_pass(enriched)
    return enriched, extra


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Enrich pharmacy product catalog for PriceBot")
    parser.add_argument("--input", required=True, help="Original Excel/CSV file from the pharmacy")
    parser.add_argument("--out-ready", default="products_enriched_ready.xlsx")
    parser.add_argument("--out-review", default="products_needs_review.xlsx")
    parser.add_argument("--out-report", default="catalog_quality_report.json")
    parser.add_argument("--out-suggested-substitutions", default="suggested_substitution_groups.xlsx")
    parser.add_argument("--data-dir", default="data")
    parser.add_argument("--default-available", choices=["blank", "true", "false"], default="blank", help="Use true only if the pharmacy confirms the source file lists available products.")
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    data_dir = Path(args.data_dir)
    rows, extra = enrich_catalog(input_path, data_dir, args.default_available)
    ready_rows = [row for row in rows if row.get("review_status") == "ready"]
    review_rows = [row for row in rows if row.get("review_status") != "ready"]
    write_workbook(Path(args.out_ready), ready_rows)
    write_workbook(Path(args.out_review), review_rows)
    suggestions = build_substitution_suggestions(rows)
    if suggestions:
        write_workbook(Path(args.out_suggested_substitutions), suggestions, headers=list(suggestions[0].keys()))
    else:
        write_workbook(Path(args.out_suggested_substitutions), [], headers=["suggested_substitution_group_id", "product_id", "name", "active_ingredient", "form", "strength", "review_required"])
    report = write_report(Path(args.out_report), rows, extra, assumptions={"default_available": args.default_available, "ai_used": False, "medicine_substitution_groups_auto_approved": False})
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
