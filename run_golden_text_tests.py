#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
import argparse
import json
from pathlib import Path
import sys
try:
    import openpyxl
except Exception as exc:
    raise SystemExit("openpyxl is required. Install requirements.txt first.") from exc
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
import matcher_v4 as matcher
import database


def read_cases(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    cases=[]
    for row in ws.iter_rows(min_row=2, values_only=True):
        d={headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
        if str(d.get("query") or "").strip(): cases.append(d)
    return cases


def write_report(path: Path, rows):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="golden_text_report"
    headers=["query","expected_decision","actual_decision","expected_product_id","actual_product_id","pass_fail","failure_reason","reply_safe"]
    ws.append(headers)
    for r in rows: ws.append([r.get(h,"") for h in headers])
    wb.save(path)


def main():
    ap=argparse.ArgumentParser(); ap.add_argument("--cases", required=True); ap.add_argument("--out", default="golden_text_test_report.xlsx"); args=ap.parse_args()
    cases=read_cases(Path(args.cases)); products=database.load_products(); ci=matcher.build_catalog_index(products)
    report=[]; failed=0
    for c in cases:
        d=matcher.resolve_product_query_from_index(str(c.get("query") or ""), ci)
        actual=d.decision_type.name
        pid=str((d.product or {}).get("product_id") or (d.product or {}).get("id") or "")
        expected=str(c.get("expected_decision") or "").strip()
        expid=str(c.get("expected_product_id") or "").strip()
        forbidden=str(c.get("forbidden_product_id") or "").strip()
        ok=True; reason=[]
        if expected and actual != expected: ok=False; reason.append(f"decision {actual} != {expected}")
        if expid and pid != expid: ok=False; reason.append(f"product {pid} != {expid}")
        if forbidden and pid == forbidden: ok=False; reason.append(f"forbidden product {pid}")
        if not ok: failed += 1
        report.append({"query":c.get("query"),"expected_decision":expected,"actual_decision":actual,"expected_product_id":expid,"actual_product_id":pid,"pass_fail":"PASS" if ok else "FAIL","failure_reason":"; ".join(reason),"reply_safe":"yes"})
    write_report(Path(args.out), report)
    if failed:
        raise SystemExit(f"GOLDEN_TEXT_TESTS_FAILED failed={failed} out={args.out}")
    print(f"GOLDEN_TEXT_TESTS_OK cases={len(cases)} out={args.out}")

if __name__ == "__main__": main()
