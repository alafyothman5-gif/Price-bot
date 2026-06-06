#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Manual real-image Vision acceptance runner for PriceBot.

This script calls the same Vision function used by the bot. It is intentionally
not part of pytest because it needs live API keys and local test images.

Usage:
    python tools/run_real_vision_tests.py --cases vision_real_test_cases.xlsx --out vision_real_test_report.xlsx
"""
from __future__ import annotations

import argparse
import asyncio
import base64
import json
import os
import sys
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import app  # noqa: E402
import database  # noqa: E402
import matcher  # noqa: E402


CASE_COLUMNS = [
    "image_path",
    "expected_brand",
    "expected_product_family",
    "expected_form",
    "expected_strength",
    "expected_size",
    "expected_decision",
    "expected_product_id",
    "notes",
]

REPORT_COLUMNS = [
    "image_path",
    "vision_output_json",
    "matcher_decision",
    "matched_product_id",
    "expected_decision",
    "expected_product_id",
    "pass_fail",
    "failure_reason",
]


def _load_cases(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    missing = [c for c in CASE_COLUMNS if c not in headers]
    if missing:
        raise SystemExit(f"Missing columns in {path}: {', '.join(missing)}")
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
        if str(item.get("image_path") or "").strip():
            rows.append(item)
    return rows


def _write_report(path: Path, rows: List[Dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "vision_real_test_report"
    ws.append(REPORT_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in REPORT_COLUMNS])
    wb.save(path)


def _safe_expected(value: Any) -> str:
    return str(value or "").strip()


def _matched_product_id(decision) -> str:
    product = getattr(decision, "product", None) or {}
    return str(product.get("product_id") or product.get("id") or "").strip()


def _compare(case: Dict[str, Any], ai_data: Dict[str, Any], decision) -> tuple[str, str]:
    expected_decision = _safe_expected(case.get("expected_decision"))
    expected_product_id = _safe_expected(case.get("expected_product_id"))
    matcher_decision = getattr(getattr(decision, "decision_type", None), "name", str(getattr(decision, "decision_type", "")))
    matched_product_id = _matched_product_id(decision)
    failures: List[str] = []
    if expected_decision and matcher_decision != expected_decision:
        failures.append(f"decision expected {expected_decision}, got {matcher_decision}")
    if expected_product_id and matched_product_id != expected_product_id:
        failures.append(f"product_id expected {expected_product_id}, got {matched_product_id}")
    field_map = {
        "expected_brand": "brand",
        "expected_product_family": "product_family",
        "expected_form": "form",
        "expected_strength": "strength",
        "expected_size": "size",
    }
    for expected_col, ai_key in field_map.items():
        expected = _safe_expected(case.get(expected_col)).lower()
        if not expected:
            continue
        actual = _safe_expected(ai_data.get(ai_key) or ai_data.get("product_type") if ai_key == "form" else ai_data.get(ai_key)).lower()
        if expected and expected not in actual:
            failures.append(f"vision {ai_key} expected contains {expected!r}, got {actual!r}")
    return ("PASS" if not failures else "FAIL", "; ".join(failures))


async def _run(cases_path: Path, out_path: Path) -> int:
    if not app.AI_KEYS_LIST:
        print("No OpenRouter/AI key found. Set OPENROUTER_API_KEY or OPENROUTER_KEYS to run real Vision tests.")
        _write_report(out_path, [])
        return 0
    if app.http_client is None:
        import httpx
        app.http_client = httpx.AsyncClient(timeout=30.0)
    database.init_db()
    cases = _load_cases(cases_path)
    report_rows: List[Dict[str, Any]] = []
    for case in cases:
        image_path = Path(str(case.get("image_path") or "").strip())
        if not image_path.is_absolute():
            image_path = (cases_path.parent / image_path).resolve()
        if not image_path.exists():
            report_rows.append({
                "image_path": str(case.get("image_path") or ""),
                "vision_output_json": "{}",
                "matcher_decision": "",
                "matched_product_id": "",
                "expected_decision": _safe_expected(case.get("expected_decision")),
                "expected_product_id": _safe_expected(case.get("expected_product_id")),
                "pass_fail": "FAIL",
                "failure_reason": "image file not found",
            })
            continue
        try:
            raw_bytes = image_path.read_bytes()
            b64 = app.resize_image_b64(base64.b64encode(raw_bytes).decode("utf-8"))
            ai_data = await app.analyze_image_with_ai("vision_real_test", b64)
            ai_data = app.validate_ai_data(ai_data or {})
            decision = matcher.resolve_image_query_decision(ai_data)
            pass_fail, failure_reason = _compare(case, ai_data, decision)
            report_rows.append({
                "image_path": str(case.get("image_path") or ""),
                "vision_output_json": json.dumps(ai_data, ensure_ascii=False),
                "matcher_decision": getattr(decision.decision_type, "name", str(decision.decision_type)),
                "matched_product_id": _matched_product_id(decision),
                "expected_decision": _safe_expected(case.get("expected_decision")),
                "expected_product_id": _safe_expected(case.get("expected_product_id")),
                "pass_fail": pass_fail,
                "failure_reason": failure_reason,
            })
        except Exception as exc:
            report_rows.append({
                "image_path": str(case.get("image_path") or ""),
                "vision_output_json": "{}",
                "matcher_decision": "",
                "matched_product_id": "",
                "expected_decision": _safe_expected(case.get("expected_decision")),
                "expected_product_id": _safe_expected(case.get("expected_product_id")),
                "pass_fail": "FAIL",
                "failure_reason": repr(exc),
            })
    _write_report(out_path, report_rows)
    failed = sum(1 for row in report_rows if row.get("pass_fail") != "PASS")
    print(f"Real Vision report written: {out_path} | cases={len(report_rows)} | failed={failed}")
    if app.http_client is not None:
        try:
            await app.http_client.aclose()
        except Exception:
            pass
        app.http_client = None
    return 1 if failed else 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Run manual real-image Vision acceptance tests.")
    parser.add_argument("--cases", required=True, help="Path to vision_real_test_cases.xlsx")
    parser.add_argument("--out", required=True, help="Path to output vision_real_test_report.xlsx")
    args = parser.parse_args()
    cases_path = Path(args.cases).resolve()
    out_path = Path(args.out).resolve()
    if not cases_path.exists():
        print(f"Cases file not found: {cases_path}")
        return 2
    return asyncio.run(_run(cases_path, out_path))


if __name__ == "__main__":
    raise SystemExit(main())
