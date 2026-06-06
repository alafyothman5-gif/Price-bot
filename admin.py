import base64
import csv
import hashlib
import hmac
import html
import io
import os
import pathlib
import re
import secrets
import shutil
import time
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
from urllib.parse import quote

import openpyxl
try:
    import qrcode
except Exception:
    qrcode = None
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response

import database
import matcher
from services.catalog_quality import analyze_products, report_to_csv
from services.import_service import evaluate_catalog_before_import


load_dotenv()

ADMIN_KEY = os.getenv("ADMIN_PASSWORD") or os.getenv("PRICEBOT_ADMIN_KEY") or os.getenv("ADMIN_KEY")
SESSION_COOKIE = "pricebot_admin_session"
LOGIN_CSRF_COOKIE = "pricebot_login_csrf"
SESSION_SECRET = (
    os.getenv("ADMIN_SESSION_SECRET")
    or os.getenv("PRICEBOT_ADMIN_SESSION_SECRET")
    or os.getenv("PRICEBOT_ADMIN_SESSION_SALT")
    or "pricebot-admin-dev-secret-change-me"
)
SESSION_MAX_AGE_SECONDS = int(os.getenv("ADMIN_SESSION_MAX_AGE_SECONDS", str(8 * 60 * 60)))
LOGIN_RATE_LIMIT_MAX = int(os.getenv("ADMIN_LOGIN_RATE_LIMIT_MAX", "5"))
LOGIN_RATE_LIMIT_WINDOW_SECONDS = int(os.getenv("ADMIN_LOGIN_RATE_LIMIT_WINDOW_SECONDS", str(10 * 60)))
LOGIN_LOCKOUT_SECONDS = int(os.getenv("ADMIN_LOGIN_LOCKOUT_SECONDS", str(15 * 60)))
PRICEBOT_ENV = (os.getenv("PRICEBOT_ENV") or os.getenv("ENV") or "local").lower()
_login_failures: Dict[str, List[float]] = defaultdict(list)
_login_lock_until: Dict[str, float] = {}
router = APIRouter(prefix="/admin", tags=["Admin Panel"])


def admin_key_configured() -> bool:
    return bool(ADMIN_KEY)


def _is_production_env() -> bool:
    return PRICEBOT_ENV in {"prod", "production", "live"}


def _cookie_secure(request: Request) -> bool:
    return _is_production_env() or str(request.headers.get("x-forwarded-proto", "")).lower() == "https" or request.url.scheme == "https"


def _session_signature(ts: str) -> str:
    if not ADMIN_KEY:
        return ""
    payload = f"{ADMIN_KEY}:{ts}".encode("utf-8")
    return hmac.new(SESSION_SECRET.encode("utf-8"), payload, hashlib.sha256).hexdigest()


def _session_token() -> str:
    ts = str(int(time.time()))
    return f"{ts}:{_session_signature(ts)}"


def _is_authorized(request: Request) -> bool:
    if not ADMIN_KEY:
        return False
    cookie = request.cookies.get(SESSION_COOKIE, "")
    try:
        ts, sig = cookie.split(":", 1)
        if time.time() - int(ts) > SESSION_MAX_AGE_SECONDS:
            return False
        return bool(sig and hmac.compare_digest(sig, _session_signature(ts)))
    except Exception:
        return False


def _request_ip(request: Request) -> str:
    forwarded = str(request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
    return forwarded or (request.client.host if request.client else "unknown")


def _login_locked(request: Request) -> bool:
    ip = _request_ip(request)
    until = float(_login_lock_until.get(ip) or 0)
    if until and until > time.time():
        return True
    if until:
        _login_lock_until.pop(ip, None)
    return False


def _record_login_failure(request: Request) -> None:
    ip = _request_ip(request)
    now = time.time()
    recent = [t for t in _login_failures[ip] if now - t <= LOGIN_RATE_LIMIT_WINDOW_SECONDS]
    recent.append(now)
    _login_failures[ip] = recent
    if len(recent) >= LOGIN_RATE_LIMIT_MAX:
        _login_lock_until[ip] = now + LOGIN_LOCKOUT_SECONDS


def _clear_login_failures(request: Request) -> None:
    ip = _request_ip(request)
    _login_failures.pop(ip, None)
    _login_lock_until.pop(ip, None)


def _set_session_cookie(response: Response, request: Request) -> None:
    response.set_cookie(
        SESSION_COOKIE,
        _session_token(),
        httponly=True,
        secure=_cookie_secure(request),
        samesite="lax",
        max_age=SESSION_MAX_AGE_SECONDS,
    )


def _delete_session_cookie(response: Response) -> None:
    response.delete_cookie(SESSION_COOKIE)


def _new_login_csrf() -> str:
    return secrets.token_urlsafe(32)


def _validate_login_csrf(request: Request, token: str) -> bool:
    cookie = request.cookies.get(LOGIN_CSRF_COOKIE, "")
    return bool(cookie and token and hmac.compare_digest(cookie, token))


def _csrf_token(request: Request) -> str:
    session = request.cookies.get(SESSION_COOKIE, "")
    if not session:
        return ""
    return hmac.new(SESSION_SECRET.encode("utf-8"), f"csrf:{session}".encode("utf-8"), hashlib.sha256).hexdigest()


def csrf_field(request: Request) -> str:
    return f'<input type="hidden" name="csrf_token" value="{h(_csrf_token(request))}">'


def require_csrf(request: Request, token: str) -> None:
    expected = _csrf_token(request)
    if not expected or not token or not hmac.compare_digest(expected, token):
        raise HTTPException(status_code=403, detail="CSRF token invalid.")


def require_admin(request: Request):
    # Do not accept the admin key in the URL. Login through /admin/login so the
    # secret does not leak into browser history, proxy logs, or screenshots.
    if not ADMIN_KEY:
        raise HTTPException(status_code=503, detail="Admin password is not configured. Set ADMIN_PASSWORD.")
    if not _is_authorized(request):
        raise HTTPException(status_code=401, detail="غير مصرح بالدخول. افتح /admin/login")
    return True


def h(value) -> str:
    return html.escape(str(value or ""), quote=True)


def admin_path(path: str, **params) -> str:
    pairs = [f"{quote(str(k))}={quote(str(v))}" for k, v in params.items() if v is not None and v != ""]
    return f"{path}?{'&'.join(pairs)}" if pairs else path


def parse_float(value) -> float:
    text = str(value or "").replace("د.ل", " ").replace("دينار", " ").replace(",", ".")
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return 0.0
    try:
        return float(match.group(0))
    except Exception:
        return 0.0


def parse_dt(value):
    if not value:
        return None
    text = str(value)[:19]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            pass
    return None


def get_html_header(title: str = "لوحة الصيدلية") -> str:
    try:
        settings = database.get_merchant_settings()
        pharmacy_name = settings.get("pharmacy_name") or os.getenv("PHARMACY_NAME", "PriceBot")
    except Exception:
        pharmacy_name = os.getenv("PHARMACY_NAME", "PriceBot")
    return f"""
<!DOCTYPE html><html dir="rtl" lang="ar"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0"><title>{h(title)}</title>
<style>
:root{{--bg:#f3f6fb;--card:#fff;--ink:#172033;--muted:#637083;--line:#e1e7ef;--blue:#2563eb;--green:#16a34a;--orange:#f59e0b;--red:#dc2626;--purple:#7c3aed}}
*{{box-sizing:border-box}}body{{font-family:Tahoma,Arial,sans-serif;background:var(--bg);margin:0;color:var(--ink)}}.layout{{max-width:1280px;margin:auto;padding:18px}}
.topbar{{background:linear-gradient(135deg,#0f172a,#1e3a8a);color:#fff;border-radius:20px;padding:18px;box-shadow:0 12px 28px rgba(15,23,42,.18)}}.topbar h1{{margin:0 0 6px;font-size:24px}}.topbar p{{margin:0;opacity:.85}}
.nav{{display:flex;gap:8px;flex-wrap:wrap;margin-top:14px}}.nav a{{color:#fff;text-decoration:none;padding:9px 12px;border-radius:999px;background:rgba(255,255,255,.13)}}.nav a:hover{{background:rgba(255,255,255,.22)}}
.grid{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px;margin-top:16px}}.grid-2{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:16px}}
.card,.metric{{background:var(--card);border:1px solid var(--line);border-radius:18px;padding:16px;box-shadow:0 6px 20px rgba(15,23,42,.06)}}.metric-title{{font-size:13px;color:var(--muted)}}.metric-value{{font-size:27px;font-weight:800;margin-top:7px}}.metric-hint{{font-size:12px;color:var(--muted);margin-top:6px}}
.metric-blue{{border-top:4px solid var(--blue)}}.metric-green{{border-top:4px solid var(--green)}}.metric-orange{{border-top:4px solid var(--orange)}}.metric-red{{border-top:4px solid var(--red)}}.metric-purple{{border-top:4px solid var(--purple)}}
table{{width:100%;border-collapse:collapse;margin-top:12px;background:#fff;border-radius:12px;overflow:hidden}}th,td{{padding:10px;border-bottom:1px solid var(--line);text-align:right;vertical-align:top}}th{{background:#f8fafc;color:#334155}}input,select,textarea{{font:inherit;padding:10px;border:1px solid #cbd5e1;border-radius:10px;width:100%;margin:4px 0}}textarea{{min-height:80px}}.btn{{display:inline-block;padding:10px 14px;margin:4px;color:#fff;background:var(--blue);text-decoration:none;border-radius:10px;border:0;cursor:pointer}}.btn-success{{background:var(--green)}}.btn-danger{{background:var(--red)}}.btn-muted{{background:#64748b}}.muted{{color:var(--muted);font-size:13px}}.badge{{padding:5px 10px;border-radius:999px;color:#fff;font-size:12px}}.bg-pending{{background:var(--orange)}}.bg-completed{{background:var(--green)}}.bg-canceled{{background:var(--red)}}
.bar-row{{display:grid;grid-template-columns:130px 1fr 45px;gap:10px;align-items:center;margin:10px 0}}.bar-track{{background:#e2e8f0;border-radius:999px;height:10px;overflow:hidden}}.bar-fill{{background:linear-gradient(90deg,#2563eb,#16a34a);height:100%}}
@media(max-width:900px){{.grid,.grid-2{{grid-template-columns:1fr}}}}
</style></head><body><div class="layout"><div class="topbar"><h1>{h(title)}</h1><p>PriceBot — {h(pharmacy_name)}</p><div class="nav">
<a href="/admin/dashboard">الرئيسية</a><a href="/admin/products">المنتجات</a><a href="/admin/products/import">Import Wizard</a><a href="/admin/products/review">Review Queue</a><a href="/admin/products/duplicates">Duplicates</a><a href="/admin/catalog-quality">Catalog Quality</a><a href="/admin/orders">الطلبات</a><a href="/admin/failed-queries">Failed Queries</a><a href="/admin/vision-failures">Vision Failures</a><a href="/admin/learning-center">Learning</a><a href="/admin/quality-dashboard">Quality</a><a href="/admin/ai-usage">AI Usage</a><a href="/admin/settings">Settings</a><a href="/admin/security">Security</a><a href="/admin/logs">Logs</a><a href="/admin/merchants">Merchants</a><a href="/admin/customer-link">QR</a><a href="/admin/logout">خروج</a>
</div></div>
"""


def html_footer() -> str:
    return "</div></body></html>"


def metric_card(title: str, value: str, hint: str = "", tone: str = "blue") -> str:
    return f'<div class="metric metric-{tone}"><div class="metric-title">{h(title)}</div><div class="metric-value">{h(value)}</div><div class="metric-hint">{h(hint)}</div></div>'


def progress_bar(label: str, value: int, max_value: int) -> str:
    pct = 0 if max_value <= 0 else min(100, round((value / max_value) * 100))
    return f'<div class="bar-row"><div>{h(label)}</div><div class="bar-track"><div class="bar-fill" style="width:{pct}%"></div></div><div>{h(value)}</div></div>'


def order_product_name(order: dict) -> str:
    return str(order.get("product_name") or order.get("product") or order.get("message") or "").strip()


def safe_status(order: dict) -> str:
    return str(order.get("status") or "pending").lower().strip()


def order_stats(orders: List[dict]) -> dict:
    now = datetime.now(); today = now.date(); last_7 = now - timedelta(days=7); last_30 = now - timedelta(days=30)
    total=len(orders); pending=sum(1 for o in orders if safe_status(o)=="pending"); completed=sum(1 for o in orders if safe_status(o)=="completed"); canceled=sum(1 for o in orders if safe_status(o)=="canceled")
    today_count=week_count=month_count=0; product_counter=Counter(); product_revenue=defaultdict(float); phone_counter=Counter(); daily=Counter(); total_revenue=0.0
    for o in orders:
        product=order_product_name(o) or "غير محدد"; product_counter[product]+=1; product_revenue[product]+=parse_float(o.get("price")); phone_counter[str(o.get("phone") or "غير معروف")]+=1; total_revenue+=parse_float(o.get("price"))
        dt=parse_dt(o.get("created_at") or o.get("time"))
        if dt:
            if dt.date()==today: today_count+=1
            if dt>=last_7: week_count+=1
            if dt>=last_30: month_count+=1
            daily[dt.date().isoformat()]+=1
    return {"total":total,"pending":pending,"completed":completed,"canceled":canceled,"today":today_count,"week":week_count,"month":month_count,"completion_rate":round(completed/total*100,1) if total else 0,"cancel_rate":round(canceled/total*100,1) if total else 0,"total_revenue":total_revenue,"top_products":[(n,c,product_revenue[n]) for n,c in product_counter.most_common(10)],"top_customers":phone_counter.most_common(8),"last_7_days":[((today-timedelta(days=i)).isoformat(), daily.get((today-timedelta(days=i)).isoformat(),0)) for i in range(6,-1,-1)]}


def products_health(products: List[dict]) -> dict:
    brands=Counter(); forms=Counter(); norm=Counter(); missing_price=missing_brand=missing_aliases=missing_form=available=unavailable=0
    for p in products:
        if matcher.is_available(p.get("available","")): available+=1
        else: unavailable+=1
        if not str(p.get("price") or "").strip(): missing_price+=1
        brand=str(p.get("brand") or p.get("company") or "").strip()
        if brand: brands[brand]+=1
        else: missing_brand+=1
        form=str(p.get("form") or p.get("category") or p.get("category_guess") or "").strip()
        if form: forms[form]+=1
        else: missing_form+=1
        if not str(p.get("aliases") or "").strip() and not str(p.get("image_ocr_keywords") or "").strip(): missing_aliases+=1
        n=matcher.normalize_text(p.get("name", ""));
        if n: norm[n]+=1
    duplicates=sum(c-1 for c in norm.values() if c>1)
    return {"total":len(products),"available":available,"unavailable":unavailable,"missing_price":missing_price,"missing_brand":missing_brand,"missing_aliases":missing_aliases,"missing_form":missing_form,"duplicates":duplicates,"top_brands":brands.most_common(8),"top_forms":forms.most_common(8)}


@router.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    token = _new_login_csrf()
    content = get_html_header("تسجيل دخول الأدمن") + f'<div class="card"><form method="post" action="/admin/login"><input type="hidden" name="csrf_token" value="{h(token)}"><label>كلمة مرور الأدمن</label><input type="password" name="key" required><button class="btn btn-success" type="submit">دخول</button></form></div>' + html_footer()
    response = HTMLResponse(content)
    response.set_cookie(LOGIN_CSRF_COOKIE, token, httponly=True, secure=_cookie_secure(request), samesite="lax", max_age=600)
    return response


@router.post("/login")
async def login(request: Request, key: str = Form(...), csrf_token: str = Form("")):
    if not _validate_login_csrf(request, csrf_token):
        raise HTTPException(status_code=403, detail="CSRF token invalid.")
    if _login_locked(request):
        raise HTTPException(status_code=429, detail="محاولات دخول كثيرة. حاول لاحقاً.")
    if not ADMIN_KEY or not hmac.compare_digest(key, ADMIN_KEY):
        _record_login_failure(request)
        raise HTTPException(status_code=401, detail="غير مصرح.")
    _clear_login_failures(request)
    response = RedirectResponse(url="/admin", status_code=303)
    _set_session_cookie(response, request)
    response.delete_cookie(LOGIN_CSRF_COOKIE)
    return response


@router.get("/logout", response_class=HTMLResponse)
async def logout_page(request: Request, _: bool = Depends(require_admin)):
    return HTMLResponse(get_html_header("تسجيل خروج") + f'<div class="card"><form method="post" action="/admin/logout">{csrf_field(request)}<p>اضغط الزر لتسجيل الخروج.</p><button class="btn btn-danger" type="submit">تسجيل خروج</button></form></div>' + html_footer())


@router.post("/logout")
async def logout(request: Request, csrf_token: str = Form(""), _: bool = Depends(require_admin)):
    require_csrf(request, csrf_token)
    response = RedirectResponse(url="/admin/login", status_code=303)
    _delete_session_cookie(response)
    return response


@router.get("", response_class=HTMLResponse)
@router.get("/", response_class=HTMLResponse)
@router.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    require_admin(request)
    products = database.load_products(); orders = database.get_all_orders(); stats=order_stats(orders); health=products_health(products); ai=database.get_ai_usage_summary(30)
    content=get_html_header("لوحة التحكم")
    content += '<div class="grid">'+ metric_card("المنتجات", str(len(products)), "إجمالي قائمة الصيدلية") + metric_card("طلبات اليوم", str(stats["today"]), "آخر 24 ساعة", "green") + metric_card("قيد الانتظار", str(stats["pending"]), "تحتاج متابعة", "orange") + metric_card("AI صور 30 يوم", str(ai["total"]), f"نجاح {ai['success']} — Tokens {ai['tokens']}", "purple") + '</div>'
    content += '<div class="grid-2"><div class="card"><h2>أكثر المنتجات طلباً</h2><table><tr><th>المنتج</th><th>العدد</th><th>القيمة</th></tr>'
    for name,count,revenue in stats["top_products"][:8]: content += f'<tr><td>{h(name)}</td><td>{count}</td><td>{revenue:.2f}</td></tr>'
    content += '</table></div><div class="card"><h2>صحة ملف المنتجات</h2>' + progress_bar("بلا سعر", health["missing_price"], max(health["total"],1)) + progress_bar("بلا براند", health["missing_brand"], max(health["total"],1)) + progress_bar("بلا aliases/OCR", health["missing_aliases"], max(health["total"],1)) + progress_bar("تكرار محتمل", health["duplicates"], max(health["total"],1)) + '</div></div>'
    return HTMLResponse(content + html_footer())


@router.get("/analytics", response_class=HTMLResponse)
async def analytics(request: Request, _: bool = Depends(require_admin)):
    products=database.load_products(); orders=database.get_all_orders(); stats=order_stats(orders); health=products_health(products); ai=database.get_ai_usage_summary(30); failed=database.get_failed_queries(10)
    content=get_html_header("الإحصائيات")
    content += '<div class="grid">'+ metric_card("طلبات اليوم", str(stats["today"]), "اليوم", "green")+metric_card("آخر 7 أيام", str(stats["week"]), "أسبوع", "blue")+metric_card("آخر 30 يوم", str(stats["month"]), "شهر", "purple")+metric_card("معدل الإكمال", f"{stats['completion_rate']}%", "طلبات مكتملة", "green")+'</div>'
    content += '<div class="grid">'+metric_card("قيد الانتظار", str(stats["pending"]), "", "orange")+metric_card("مكتملة", str(stats["completed"]), "", "green")+metric_card("ملغاة", str(stats["canceled"]), f"إلغاء {stats['cancel_rate']}%", "red")+metric_card("قيمة الطلبات", f"{stats['total_revenue']:.2f}", "تقريبية", "blue")+'</div>'
    content += '<div class="grid-2"><div class="card"><h2>أكثر المنتجات طلباً</h2><table><tr><th>المنتج</th><th>العدد</th><th>قيمة تقريبية</th></tr>'
    for name,count,revenue in stats["top_products"]: content += f'<tr><td>{h(name)}</td><td>{count}</td><td>{revenue:.2f}</td></tr>'
    content += '</table></div><div class="card"><h2>AI Usage آخر 30 يوم</h2>'+metric_card("صور/محاولات", str(ai["total"]), "", "purple")+metric_card("نجاح", str(ai["success"]), "", "green")+metric_card("Tokens", str(ai["tokens"]), "", "blue")+metric_card("تكلفة تقريبية", f"{ai['cost']:.4f}$", "حسب env", "orange")+'</div></div>'
    content += '<div class="grid-2"><div class="card"><h2>أكثر الزبائن طلباً</h2><table><tr><th>الرقم</th><th>عدد الطلبات</th></tr>'
    for phone,count in stats["top_customers"]: content += f'<tr><td dir="ltr">{h(phone)}</td><td>{count}</td></tr>'
    content += '</table></div><div class="card"><h2>استعلامات فاشلة/غير متوفرة</h2><table><tr><th>الاستعلام</th><th>الحالة</th><th>العدد</th></tr>'
    for row in failed: content += f'<tr><td>{h(row.get("raw_query"))}</td><td>{h(row.get("status"))}</td><td>{h(row.get("count"))}</td></tr>'
    content += '</table></div></div>'
    return HTMLResponse(content+html_footer())


def product_matches(product: dict, query: str) -> bool:
    if not query: return True
    q_norm = matcher.normalize_text(query); identity = matcher.get_product_identity(product)
    return q_norm in identity




def normalize_public_phone(value: str) -> str:
    digits = re.sub(r"\D+", "", str(value or ""))
    if digits.startswith("00"):
        digits = digits[2:]
    if digits.startswith("0") and len(digits) == 10:
        digits = "218" + digits[1:]
    return digits


def configured_customer_link_phone() -> str:
    for key in ["BOT_WHATSAPP_NUMBER", "WHATSAPP_PUBLIC_NUMBER", "PUBLIC_WHATSAPP_NUMBER", "PHARMACY_WHATSAPP_NUMBER", "WHATSAPP_DISPLAY_PHONE"]:
        value = os.getenv(key)
        if value:
            return normalize_public_phone(value)
    return ""


def make_qr_data_uri(link: str) -> str:
    if not qrcode:
        return ""
    img = qrcode.make(link)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")


@router.get("/customer-link", response_class=HTMLResponse)
async def customer_link_page(request: Request, phone: str = "", text: str = "السلام عليكم", _: bool = Depends(require_admin)):
    number = normalize_public_phone(phone) or configured_customer_link_phone()
    content = get_html_header("رابط الزبائن و QR للطباعة")
    content += '<div class="card"><h2>رابط واتساب للزبائن</h2><p class="muted">ضع رقم واتساب البوت بصيغة دولية بدون +. مثال: 2189XXXXXXXX. الأفضل إضافته في .env باسم BOT_WHATSAPP_NUMBER.</p>'
    content += f'<form method="get" action="/admin/customer-link"><label>رقم واتساب البوت</label><input name="phone" dir="ltr" value="{h(number)}" placeholder="2189XXXXXXXX"><label>رسالة افتتاحية اختيارية</label><input name="text" value="{h(text)}"><button class="btn btn-success">إنشاء الرابط والـ QR</button></form></div>'
    if number:
        from urllib.parse import quote as _quote
        link = f"https://wa.me/{number}?text={_quote(text or 'السلام عليكم')}"
        qr = make_qr_data_uri(link)
        content += '<div class="card" id="printArea" style="text-align:center"><h2>صَوّر أو اطبع هذا الكود</h2>'
        content += f'<p><b>الرابط:</b></p><p dir="ltr" style="font-size:18px;word-break:break-all">{h(link)}</p>'
        if qr:
            content += f'<img src="{qr}" alt="QR" style="width:320px;max-width:100%;border:12px solid white;box-shadow:0 3px 18px #ddd">'
        else:
            content += '<p class="muted">مكتبة qrcode غير مثبتة. بعد تشغيل pip install -r requirements.txt سيظهر الكود.</p>'
        content += '<h1 style="margin:12px 0 4px">{h(database.get_merchant_settings().get("pharmacy_name") or "PriceBot")}</h1><p style="font-size:20px">امسح الكود للاستفسار عن السعر والتوفر عبر واتساب</p><button class="btn" onclick="window.print()">طباعة</button></div>'
    else:
        content += '<div class="card"><p>لم يتم ضبط رقم البوت. أدخل الرقم في النموذج أو أضف BOT_WHATSAPP_NUMBER داخل .env.</p></div>'
    content += '<style>@media print{.topbar,.card form,.btn{display:none!important}body{background:white}.layout{max-width:800px}.card{border:0;box-shadow:none}}</style>'
    return HTMLResponse(content + html_footer())


@router.get("/products", response_class=HTMLResponse)
async def manage_products(request: Request, q: str = "", page: int = 1, _: bool = Depends(require_admin)):
    all_products=database.load_products(); filtered=[p for p in all_products if product_matches(p,q)]; per_page=100; page=max(1,page); total_pages=max((len(filtered)+per_page-1)//per_page,1); page=min(page,total_pages); page_products=filtered[(page-1)*per_page:page*per_page]
    csrf = csrf_field(request)
    content=get_html_header("المنتجات")
    content += f'<div class="card"><h2>رفع Excel/CSV</h2><p class="muted">يفضل رفع ملف products_enriched_ready.xlsx الناتج من tools/enrich_products_catalog.py حتى تصل أعمدة brand/form/strength/size/product_family للبوت بشكل صحيح.</p><form action="/admin/upload" method="post" enctype="multipart/form-data">{csrf}<input type="file" name="file" accept=".csv,.xlsx,.xls" required><label><input type="checkbox" name="replace_all" value="yes" style="width:auto"> استبدال كامل بعد التحقق</label><br><label><input type="checkbox" name="force_confirm" value="yes" style="width:auto"> تأكيد صريح للاستبدال الكامل</label><br><button class="btn btn-success" type="submit">رفع وتحديث</button></form></div>'
    content += f'<div class="card"><form method="get" action="/admin/products"><input type="text" name="q" value="{h(q)}" placeholder="ابحث في كل المنتجات..."></form><p class="muted">النتائج: {len(filtered)} من {len(all_products)} — الصفحة {page}/{total_pages}</p><table><tr><th>ID</th><th>الاسم</th><th>السعر</th><th>البراند</th><th>الشكل</th><th>التوفر</th><th>إجراءات</th></tr>'
    for p in page_products:
        content += f'<tr><td>{h(p.get("id"))}</td><td>{h(p.get("name"))}</td><td>{h(p.get("price"))}</td><td>{h(p.get("brand") or p.get("company"))}</td><td>{h(p.get("form"))}</td><td>{h(p.get("available"))}</td><td><a class="btn" href="/admin/products/edit/{h(p.get("id"))}">تعديل</a><form action="/admin/products/delete/{h(p.get("id"))}" method="post" style="display:inline">{csrf}<button class="btn btn-danger" onclick="return confirm(\'تأكيد حذف المنتج؟\')">حذف</button></form></td></tr>'
    content += '</table>'
    if page>1: content += f'<a class="btn btn-muted" href="{admin_path("/admin/products",q=q,page=page-1)}">السابق</a>'
    if page<total_pages: content += f'<a class="btn btn-muted" href="{admin_path("/admin/products",q=q,page=page+1)}">التالي</a>'
    content += '</div>'
    return HTMLResponse(content+html_footer())


@router.get("/products/edit/{product_id}", response_class=HTMLResponse)
async def edit_product(product_id: int, request: Request, _: bool = Depends(require_admin)):
    p=database.get_product(product_id)
    if not p: raise HTTPException(status_code=404, detail="Product not found")
    fields=["product_id","name","normalized_name","price","currency","brand","company","category","subcategory","product_family","form","size","aliases","image_ocr_keywords","ocr_keywords","active_ingredient","use_case","skin_type","body_area","age_group","gender","medicine_route","strength","pack","barcode","available","requires_clarification","substitution_group_id","is_substitutable","image_refs","review_status","review_notes","last_updated","merchant_id","code","sku","item_code","product_code"]
    content=get_html_header("تعديل منتج") + f'<div class="card"><h2>تعديل #{product_id}</h2><form method="post" action="/admin/products/edit/{product_id}">{csrf_field(request)}'
    for f in fields:
        value=h(p.get(f,"")); textarea=f in {"aliases","image_ocr_keywords"}
        content += f'<label>{f}</label>' + (f'<textarea name="{f}">{value}</textarea>' if textarea else f'<input name="{f}" value="{value}">')
    content += '<button class="btn btn-success" type="submit">حفظ</button><a class="btn btn-muted" href="/admin/products">رجوع</a></form></div>'
    return HTMLResponse(content+html_footer())


@router.post("/products/edit/{product_id}")
async def save_product(product_id: int, request: Request, _: bool = Depends(require_admin)):
    form=await request.form(); require_csrf(request, str(form.get("csrf_token") or ""))
    payload = dict(form); payload.pop("csrf_token", None)
    validate_product_payload(payload)
    database.backup_database(); database.update_product(product_id, payload); matcher.invalidate_product_cache()
    database.log_audit("product_edit", "admin", "product", str(product_id), new_value=str({k: payload.get(k) for k in ["name", "price", "available", "form", "strength"]}), ip=request.client.host if request.client else "")
    return RedirectResponse(url=f"/admin/products/edit/{product_id}", status_code=303)


HEADER_ALIASES={
    "product_id":{"product_id","internal_id","معرف المنتج"},
    "name":{"name","product","item","product_name","product name","اسم المنتج","اسم الصنف","المنتج","الاسم","الصنف","canonical_name"},
    "price":{"price","السعر","سعر","final_price","box_price","strip_price","cost","سعر البيع"},
    "company":{"company","الشركة","المصنع","الوكيل"},
    "brand":{"brand","الماركة","البراند"},
    "category":{"category","تصنيف","الفئة","القسم"},
    "product_family":{"product_family","family","عائلة المنتج"},
    "aliases":{"aliases","اسماء بديلة","اسم بديل"},
    "image_ocr_keywords":{"image_ocr_keywords","ocr_keywords","keywords","كلمات","كلمات البحث"},
    "ocr_keywords":{"ocr_keywords"},
    "form":{"form","type","form_or_type","الشكل الدوائي","الشكل","النوع","شكل"},
    "size":{"size","volume","حجم","الحجم"},
    "available":{"available","availability","status","الحالة","التوفر","توفر","الكمية","qty"},
    "active_ingredient":{"active_ingredient","active ingredient","المادة الفعالة","المادة"},
    "use_case":{"use_case","use","purpose","الاستخدام"},
    "skin_type":{"skin_type","نوع البشرة"},
    "strength":{"strength","concentration","dose","strength_or_size","تركيز","جرعة"},
    "pack":{"pack","package","عبوة","العبوة"},
    "substitution_group_id":{"substitution_group_id","substitution_group","بدائل"},
    "review_status":{"review_status","status_review"},
    "review_notes":{"review_notes","notes_review"},
    "barcode":{"barcode","باركود"}, "subcategory":{"subcategory"}, "body_area":{"body_area"}, "age_group":{"age_group"},
    "gender":{"gender"}, "medicine_route":{"medicine_route"}, "requires_clarification":{"requires_clarification"},
    "currency":{"currency"}, "is_substitutable":{"is_substitutable"}, "image_refs":{"image_refs"}, "last_updated":{"last_updated"}, "merchant_id":{"merchant_id"},
    "code":{"code","كود"},"sku":{"sku"},"item_code":{"item_code"},"product_code":{"product_code"}
}


def map_header(header)->str:
    raw=str(header or "").strip(); norm=matcher.normalize_text(raw).replace("_"," "); low=raw.lower()
    for canon,aliases in HEADER_ALIASES.items():
        if norm in {matcher.normalize_text(a).replace("_"," ") for a in aliases} or low in {a.lower() for a in aliases}: return canon
    return low.replace(" ","_")


def unique_headers(headers: List[str])->List[str]:
    seen={}; result=[]
    for header in headers:
        mapped=map_header(header); count=seen.get(mapped,0); seen[mapped]=count+1; result.append(mapped if count==0 else f"{mapped}__{count}")
    return result


def collect(row: dict, key: str) -> str:
    vals=[]
    for k,v in row.items():
        if k==key or str(k).startswith(f"{key}__"):
            t="" if v is None else str(v).strip()
            if t and t.lower()!="none": vals.append(t)
    return " | ".join(vals)




def validate_product_payload(payload: dict) -> None:
    price = str(payload.get("price") or "").strip()
    if price:
        normalized = price.replace(",", ".")
        if not re.search(r"\d+(?:\.\d+)?", normalized):
            raise HTTPException(status_code=400, detail="price must be numeric")
    category = matcher.normalize_text(payload.get("category", ""))
    if category == "medicine":
        if not str(payload.get("form") or "").strip():
            raise HTTPException(status_code=400, detail="medicine requires form")
        if not str(payload.get("strength") or "").strip():
            raise HTTPException(status_code=400, detail="medicine requires strength")
    if category == "cosmetic" and not str(payload.get("form") or "").strip():
        raise HTTPException(status_code=400, detail="cosmetic requires form/type")

def product_from_row(row: dict)->dict:
    name=collect(row,"name")
    if not name: return {}
    company=collect(row,"company"); brand=collect(row,"brand") or company
    image_kw=collect(row,"image_ocr_keywords") or collect(row,"ocr_keywords")
    return {
        "product_id":collect(row,"product_id"),"name":name,"price":collect(row,"price"),"company":company,"brand":brand,
        "category":collect(row,"category"),"product_family":collect(row,"product_family"),"form":collect(row,"form"),"size":collect(row,"size"),
        "aliases":collect(row,"aliases"),"image_ocr_keywords":image_kw,"ocr_keywords":image_kw,
        "available":collect(row,"available"),"active_ingredient":collect(row,"active_ingredient"),"use_case":collect(row,"use_case"),"skin_type":collect(row,"skin_type"),
        "strength":collect(row,"strength"),"pack":collect(row,"pack"),"substitution_group_id":collect(row,"substitution_group_id"),
        "review_status":collect(row,"review_status"),"review_notes":collect(row,"review_notes"),
        "subcategory":collect(row,"subcategory"),"body_area":collect(row,"body_area"),"age_group":collect(row,"age_group"),"gender":collect(row,"gender"),
        "medicine_route":collect(row,"medicine_route"),"requires_clarification":collect(row,"requires_clarification"),"currency":collect(row,"currency") or "LYD",
        "is_substitutable":collect(row,"is_substitutable"),"image_refs":collect(row,"image_refs"),"last_updated":collect(row,"last_updated"),"merchant_id":collect(row,"merchant_id") or "default",
        "code":collect(row,"code"),"barcode":collect(row,"barcode"),"sku":collect(row,"sku"),"item_code":collect(row,"item_code"),"product_code":collect(row,"product_code")
    }


def parse_csv(content: bytes)->Tuple[List[dict],List[str]]:
    text=content.decode("utf-8-sig"); rows=list(csv.reader(io.StringIO(text)))
    if not rows: raise ValueError("الملف فارغ.")
    headers=unique_headers([str(c or "") for c in rows[0]]); parsed=[dict(zip(headers,row)) for row in rows[1:] if any(str(c or "").strip() for c in row)]
    return parsed, headers


def find_xlsx_table(content: bytes)->Tuple[List[dict],List[str]]:
    workbook=openpyxl.load_workbook(io.BytesIO(content),data_only=True,read_only=True)
    best=[]
    for sheet in workbook.worksheets:
        rows=list(sheet.iter_rows(values_only=True))
        for idx,row in enumerate(rows[:20]):
            headers=unique_headers([str(c or "") for c in row])
            if "name" in headers:
                return [dict(zip(headers,values)) for values in rows[idx+1:] if any(str(c or "").strip() for c in values)], headers
        if len(rows)>len(best): best=rows
    if best:
        headers=unique_headers([str(c or "") for c in best[0]]); return [dict(zip(headers,values)) for values in best[1:] if any(str(c or "").strip() for c in values)], headers
    raise ValueError("ملف Excel فارغ.")


def find_xls_table(content: bytes)->Tuple[List[dict],List[str]]:
    try:
        import xlrd
    except Exception:
        raise ValueError("صيغة .xls تحتاج xlrd. حوّل الملف إلى .xlsx ثم ارفعه.")
    workbook = xlrd.open_workbook(file_contents=content)
    best=[]
    for sheet in workbook.sheets():
        rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        for idx,row in enumerate(rows[:20]):
            headers=unique_headers([str(c or "") for c in row])
            if "name" in headers:
                return [dict(zip(headers,values)) for values in rows[idx+1:] if any(str(c or "").strip() for c in values)], headers
        if len(rows)>len(best): best=rows
    if best:
        headers=unique_headers([str(c or "") for c in best[0]]); return [dict(zip(headers,values)) for values in best[1:] if any(str(c or "").strip() for c in values)], headers
    raise ValueError("ملف Excel فارغ.")


def parse_upload(content: bytes, filename: str)->Tuple[List[dict],List[str]]:
    lower=(filename or "").lower()
    if lower.endswith(".csv"):
        rows,headers = parse_csv(content)
    elif lower.endswith(".xlsx"):
        rows,headers = find_xlsx_table(content)
    elif lower.endswith(".xls"):
        rows,headers = find_xls_table(content)
    else:
        raise ValueError("صيغة الملف غير مدعومة. ارفع CSV أو XLSX. إذا كان الملف XLS قديماً حوّله إلى .xlsx ثم ارفعه.")
    if "name" not in headers: raise ValueError("لم يتم العثور على عمود اسم المنتج.")
    products=[product_from_row(row) for row in rows]; products=[p for p in products if p]
    if not products: raise ValueError("لم يتم العثور على منتجات صالحة داخل الملف.")
    return products, headers


def upsert_products(products: List[dict], replace_all: bool)->int:
    import database as _database
    _database.init_db()
    import_columns=["product_id","name","normalized_name","price","currency","company","brand","category","subcategory","product_family","form","size","aliases","image_ocr_keywords","ocr_keywords","active_ingredient","use_case","skin_type","body_area","age_group","gender","medicine_route","available","strength","pack","barcode","requires_clarification","substitution_group_id","is_substitutable","image_refs","review_status","review_notes","last_updated","merchant_id","code","sku","item_code","product_code"]
    with database.get_db_connection() as conn:
        if replace_all: conn.execute("DELETE FROM products")
        changed=0
        for product in products:
            name=product["name"].strip(); normalized=matcher.normalize_text(name); product_id=str(product.get("product_id") or "").strip()
            existing=None
            if product_id:
                existing=conn.execute("SELECT id FROM products WHERE product_id=? LIMIT 1",(product_id,)).fetchone()
            if not existing:
                existing=conn.execute("SELECT id FROM products WHERE normalized_name=? OR name=? LIMIT 1",(normalized,name)).fetchone()
            payload={col:str(product.get(col,"") or "") for col in import_columns}
            payload["name"]=name; payload["normalized_name"]=normalized
            if not payload.get("image_ocr_keywords") and payload.get("ocr_keywords"):
                payload["image_ocr_keywords"]=payload["ocr_keywords"]
            if not payload.get("ocr_keywords") and payload.get("image_ocr_keywords"):
                payload["ocr_keywords"]=payload["image_ocr_keywords"]
            cols=list(payload.keys())
            if existing:
                set_clause=", ".join(f'"{c}"=?' for c in cols) + ", updated_at=CURRENT_TIMESTAMP"
                conn.execute(f"UPDATE products SET {set_clause} WHERE id=?", [payload[c] for c in cols]+[existing["id"]])
            else:
                col_clause=", ".join(f'"{c}"' for c in cols) + ", updated_at"
                placeholders=", ".join("?" for _ in cols) + ", CURRENT_TIMESTAMP"
                conn.execute(f"INSERT INTO products ({col_clause}) VALUES ({placeholders})", [payload[c] for c in cols])
            changed+=1
        conn.commit(); matcher.invalidate_product_cache(); return changed


def render_upload_error(message: str, status_code: int = 400)->HTMLResponse:
    return HTMLResponse(get_html_header("خطأ في الرفع")+f'<div class="card"><h2>تعذر الرفع</h2><p>{h(message)}</p><a class="btn" href="/admin/products/import">عودة إلى معالج الرفع</a></div>'+html_footer(), status_code=status_code)


def _product_review_key(product: dict) -> tuple:
    key = str(product.get("product_id") or product.get("source_serial") or "").strip().lower()
    if key:
        return ("product_id", key)
    return ("name", matcher.normalize_text(product.get("normalized_name") or product.get("name") or ""))


def split_ready_and_review(products: List[dict], report: dict) -> Tuple[List[dict], List[dict]]:
    review_keys = {_product_review_key(row) for row in report.get("review_rows", [])}
    ready, review = [], []
    for product in products:
        if _product_review_key(product) in review_keys:
            review.append(product)
        else:
            ready.append(product)
    return ready, review


def render_import_quality_report(report: dict, ready_count: int, review_count: int, duplicate_count: int, reasons: List[str], message: str, status_code: int = 400) -> HTMLResponse:
    reason_items = "".join(f"<li>{h(x)}</li>" for x in (reasons or ["لا توجد أسباب رفض صريحة"]))
    content = get_html_header("نتيجة فحص الكتالوج")
    content += f"""
    <div class='card'>
      <h2>Quality Gate Result: {h(report.get('decision'))}</h2>
      <div class='grid'>
        <div class='card metric'><span>Total rows</span><br><b>{h(report.get('total'))}</b></div>
        <div class='card metric'><span>Ready rows</span><br><b>{h(ready_count)}</b></div>
        <div class='card metric'><span>Review rows</span><br><b>{h(review_count)}</b></div>
        <div class='card metric'><span>Duplicate count</span><br><b>{h(duplicate_count)}</b></div>
      </div>
      <h3>Reject / warning reasons</h3><ul>{reason_items}</ul>
      <p class='muted'>{h(message)}</p>
      <p class='muted'>الصفوف التي تحتاج مراجعة لا تُرفع كمنتجات جاهزة. يتم تحويلها إلى Review Queue للتصحيح اليدوي.</p>
      <a class='btn' href='/admin/products/import'>عودة إلى Import Wizard</a>
    </div>
    """
    return HTMLResponse(content + html_footer(), status_code=status_code)


@router.post("/upload")
async def upload_file(request: Request, file: UploadFile = File(...), replace_all: str = Form(None), force_confirm: str = Form(None), force_confirm_text: str = Form(""), csrf_token: str = Form(""), _: bool = Depends(require_admin)):
    require_csrf(request, csrf_token)
    try:
        content=await file.read(); products,headers=parse_upload(content,file.filename or ""); current=database.count_products(); do_replace=replace_all=="yes"
        decision, reasons, report = evaluate_catalog_before_import(products)
        ready_products, review_products = split_ready_and_review(products, report)
        duplicate_count = int(report.get("duplicate_name_rows") or 0) + len(report.get("duplicate_aliases") or {})
        explicit_force = force_confirm == "yes" and str(force_confirm_text or "").strip() == "IMPORT_READY_ONLY"
        if do_replace:
            if not explicit_force:
                return render_import_quality_report(report, len(ready_products), len(review_products), duplicate_count, reasons, "الاستبدال الكامل يحتاج كتابة IMPORT_READY_ONLY مع تحديد خانة التأكيد.")
            if current>=100 and len(ready_products)<max(100,int(current*0.5)):
                return render_upload_error(f"عدد المنتجات الجاهزة في الملف ({len(ready_products)}) أقل بكثير من الموجود حالياً ({current}). تم رفض الاستبدال.")
        if decision == "REJECT" and not explicit_force:
            database.ensure_v19_tables()
            for item in review_products[:5000]:
                database.add_review_queue_item(item, str(item.get("quality_issues") or ";".join(reasons) or "needs_review"))
            return render_import_quality_report(report, len(ready_products), len(review_products), duplicate_count, reasons, f"هذا الملف يحتوي {len(review_products)} صف يحتاج مراجعة. لا ننصح برفعه كاملاً قبل التصحيح.")
        if not ready_products:
            for item in review_products[:5000]:
                database.add_review_queue_item(item, str(item.get("quality_issues") or ";".join(reasons) or "needs_review"))
            return render_import_quality_report(report, 0, len(review_products), duplicate_count, reasons, "لا توجد صفوف جاهزة آمنة للاستيراد.")
        database.backup_database()
        changed=upsert_products(ready_products,do_replace)
        for item in review_products[:5000]:
            database.add_review_queue_item(item, str(item.get("quality_issues") or ";".join(reasons) or "needs_review"))
        database.rebuild_review_queue_from_products(limit=5000)
        database.log_audit("product_upload", "admin", "catalog", "upload", new_value=f"changed={changed};ready={len(ready_products)};review={len(review_products)};decision={decision}", ip=request.client.host if request.client else "")
        print(f"PRODUCT_UPLOAD_OK: changed={changed} ready={len(ready_products)} review={len(review_products)} replace_all={do_replace} quality={decision} headers={headers}")
    except Exception as exc:
        print(f"PRODUCT_UPLOAD_ERROR: {exc}"); return render_upload_error(str(exc))
    return RedirectResponse(url="/admin/products", status_code=303)


@router.post("/products/delete/{product_id}")
async def delete_product(product_id: int, request: Request, csrf_token: str = Form(""), _: bool = Depends(require_admin)):
    require_csrf(request, csrf_token)
    database.backup_database()
    with database.get_db_connection() as conn:
        conn.execute("DELETE FROM products WHERE id=?",(product_id,)); conn.commit()
    database.log_audit("product_delete", "admin", "product", str(product_id), ip=request.client.host if request.client else "")
    matcher.invalidate_product_cache(); return RedirectResponse(url="/admin/products", status_code=303)


@router.get("/orders", response_class=HTMLResponse)
async def manage_orders(request: Request, _: bool = Depends(require_admin)):
    orders=database.get_all_orders(); content=get_html_header("الطلبات")+'<div class="card"><h2>طلبات الحجز</h2><table><tr><th>الرقم</th><th>الزبون</th><th>المنتج</th><th>السعر</th><th>التاريخ</th><th>الحالة</th><th>تحديث</th></tr>'
    for o in orders:
        status=safe_status(o); badge='bg-pending' if status=='pending' else 'bg-completed' if status=='completed' else 'bg-canceled'; text='قيد الانتظار' if status=='pending' else 'مكتمل' if status=='completed' else 'ملغى'
        content += f'<tr><td>#{h(o.get("id"))}</td><td dir="ltr">{h(o.get("phone"))}</td><td>{h(order_product_name(o))}</td><td>{h(o.get("price"))}</td><td dir="ltr">{h(o.get("created_at"))}</td><td><span class="badge {badge}">{text}</span></td><td><form action="/admin/orders/update/{h(o.get("id"))}" method="post">{csrf_field(request)}<select name="new_status"><option value="pending" {"selected" if status=="pending" else ""}>انتظار</option><option value="completed" {"selected" if status=="completed" else ""}>مكتمل</option><option value="canceled" {"selected" if status=="canceled" else ""}>إلغاء</option></select><button class="btn btn-success">حفظ</button></form></td></tr>'
    return HTMLResponse(content+'</table></div>'+html_footer())


@router.post("/orders/update/{order_id}")
async def update_order(order_id: int, request: Request, new_status: str = Form(...), csrf_token: str = Form(""), _: bool = Depends(require_admin)):
    require_csrf(request, csrf_token)
    database.update_order_status(order_id,new_status); return RedirectResponse(url="/admin/orders", status_code=303)


@router.get("/failed-queries", response_class=HTMLResponse)
async def failed_queries(request: Request, _: bool = Depends(require_admin)):
    rows=database.get_failed_queries(100); content=get_html_header("الاستعلامات الفاشلة")+'<div class="card"><h2>Fallback / Unavailable</h2><table><tr><th>الاستعلام</th><th>بعد التنظيف</th><th>المصدر</th><th>الحالة</th><th>العدد</th><th>آخر مرة</th></tr>'
    for r in rows: content += f'<tr><td>{h(r.get("raw_query"))}</td><td>{h(r.get("normalized_query"))}</td><td>{h(r.get("source"))}</td><td>{h(r.get("status"))}</td><td>{h(r.get("count"))}</td><td>{h(r.get("last_seen"))}</td></tr>'
    return HTMLResponse(content+'</table></div>'+html_footer())

@router.get("/misses", response_class=HTMLResponse)
async def query_misses(request: Request, _: bool = Depends(require_admin)):
    rows = database.get_query_misses(100)
    content = get_html_header("Misses / تحسين البحث")
    content += f'<div class="card"><h2>إضافة مرادف سريع</h2><p class="muted">مثال: source = سيرا في ، target = cerave. يشتغل فوراً بدون restart.</p><form method="post" action="/admin/addsynonym">{csrf_field(request)}<label>الكلمة التي يكتبها الزبون</label><input name="source" required><label>الكلمة الصحيحة في البحث</label><input name="target" required><button class="btn btn-success" type="submit">إضافة synonym</button></form></div>'
    content += '<div class="card"><h2>أكثر الاستفسارات الفاشلة</h2><table><tr><th>الاستعلام</th><th>بعد التنظيف</th><th>القرار</th><th>التكرار</th><th>آخر مرة</th><th>إضافة synonym</th></tr>'
    for row in rows:
        raw = h(row.get("raw_query")); clean = h(row.get("clean_query")); decision = h(row.get("decision")); freq = h(row.get("freq")); last = h(row.get("last_seen"))
        content += f'<tr><td>{raw}</td><td>{clean}</td><td>{decision}</td><td>{freq}</td><td dir="ltr">{last}</td><td><form method="post" action="/admin/addsynonym">{csrf_field(request)}<input name="source" value="{raw}"><input name="target" placeholder="target مثل cerave"><button class="btn">حفظ</button></form></td></tr>'
    content += '</table></div>'
    return HTMLResponse(content + html_footer())



@router.get("/catalog-quality", response_class=HTMLResponse)
async def catalog_quality(request: Request, _: bool = Depends(require_admin)):
    rows = matcher.get_catalog_quality_rows()
    issue_counts = Counter()
    for row in rows:
        for issue in str(row.get("issues") or "").split(";"):
            if issue:
                issue_counts[issue] += 1
    content = get_html_header("Catalog Quality V4")
    ready_count = sum(1 for row in rows if str(row.get("ready") or "") == "yes" or not row.get("issues"))
    review_count = len(rows) - ready_count
    completion = round((ready_count / max(len(rows), 1)) * 100, 1)
    content += '<div class="card"><h2>تقرير جودة الكتالوج</h2><p class="muted">هذا التقرير لا يغير البيانات. يستخدم لكشف الأعمدة الناقصة التي تؤثر على المطابقة: brand/category/form/strength/aliases/OCR/use_case/skin_type.</p><a class="btn" href="/admin/catalog-quality.csv">تحميل CSV</a><a class="btn btn-success" href="/admin/catalog-quality.xlsx">تحميل XLSX</a></div>'
    content += '<div class="grid">'
    content += metric_card("عدد المنتجات", str(len(rows)), "كل المنتجات المقروءة من SQLite", "blue")
    content += metric_card("جاهزة تقريباً", str(ready_count), f"اكتمال {completion}%", "green")
    content += metric_card("تحتاج مراجعة", str(review_count), "أي منتج به نقص مؤثر", "orange")
    content += metric_card("Missing brand", str(issue_counts.get("missing_brand", 0)), "منتجات بلا شركة واضحة", "red")
    content += '</div>'
    content += '<div class="grid">'
    content += metric_card("Missing category", str(issue_counts.get("missing_category", 0)), "دواء/كوزمتك/آخر", "orange")
    content += metric_card("Medicine missing form", str(issue_counts.get("medicine_missing_form", 0)), "أدوية بلا شكل", "red")
    content += metric_card("Medicine missing strength", str(issue_counts.get("medicine_missing_strength", 0)), "جرعات ناقصة", "red")
    content += metric_card("Medicine missing active", str(issue_counts.get("medicine_missing_active_ingredient", 0)), "مادة فعالة ناقصة", "orange")
    content += '</div><div class="grid">'
    content += metric_card("Cosmetic missing type", str(issue_counts.get("cosmetic_missing_product_type_form", 0)), "كوزمتك بلا نوع/form", "red")
    content += metric_card("Cosmetic missing use", str(issue_counts.get("cosmetic_missing_use_case", 0)), "استخدام ناقص", "purple")
    content += metric_card("Missing aliases", str(issue_counts.get("missing_aliases", 0)), "منتجات بلا aliases", "orange")
    content += metric_card("Missing OCR", str(issue_counts.get("missing_ocr_keywords", 0)), "منتجات بلا ocr_keywords", "orange")
    content += '</div><div class="grid">'
    content += metric_card("Duplicate aliases", str(issue_counts.get("duplicate_alias_or_ocr_keyword", 0)), "aliases/OCR مكررة", "red")
    content += metric_card("Duplicate names", str(issue_counts.get("duplicate_normalized_name", 0)), "أسماء normalized مكررة", "red")
    content += metric_card("Suspicious prices", str(issue_counts.get("suspicious_price", 0)), "أسعار غير منطقية", "orange")
    content += metric_card("Unavailable/unknown", str(issue_counts.get("unavailable_or_unknown_availability", 0)), "توفر غير واضح", "purple")
    content += '</div>'
    content += '<div class="card"><h2>أول 250 منتج يحتاج مراجعة</h2><table><tr><th>ID</th><th>الاسم</th><th>Brand</th><th>Form</th><th>Strength</th><th>Category</th><th>Use</th><th>Skin</th><th>Issues</th></tr>'
    shown = 0
    for row in rows:
        if not row.get("issues"):
            continue
        shown += 1
        if shown > 250:
            break
        content += f'<tr><td>{h(row.get("id"))}</td><td>{h(row.get("name"))}</td><td>{h(row.get("brand"))}</td><td>{h(row.get("form"))}</td><td>{h(row.get("strength"))}</td><td>{h(row.get("category"))}</td><td>{h(row.get("use_case"))}</td><td>{h(row.get("skin_type"))}</td><td>{h(row.get("issues"))}</td></tr>'
    content += '</table></div>'
    return HTMLResponse(content + html_footer())


@router.get("/catalog-quality.csv")
async def catalog_quality_csv(request: Request, _: bool = Depends(require_admin)):
    rows = matcher.get_catalog_quality_rows()
    output = io.StringIO()
    fieldnames = ["id", "name", "brand", "family", "active_ingredient", "form", "strength", "size", "category", "cosmetic_type", "use_case", "skin_type", "price", "available", "ready", "issues"]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    return Response(content=output.getvalue(), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=catalog_quality_report.csv"})




@router.get("/catalog-quality.xlsx")
async def catalog_quality_xlsx(request: Request, _: bool = Depends(require_admin)):
    rows = matcher.get_catalog_quality_rows()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "catalog_quality"
    fieldnames = ["id", "name", "brand", "family", "active_ingredient", "form", "strength", "size", "category", "cosmetic_type", "use_case", "skin_type", "price", "available", "ready", "issues"]
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(k, "") for k in fieldnames])
    for col in ws.columns:
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = min(45, max(12, max(len(str(c.value or "")) for c in col[:100]) + 2))
    out = io.BytesIO()
    wb.save(out)
    return Response(content=out.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=catalog_quality_report.xlsx"})

@router.get("/alias-learning", response_class=HTMLResponse)
async def alias_learning(request: Request, _: bool = Depends(require_admin)):
    misses = database.get_query_misses(100)
    suggestions = database.list_alias_suggestions(50, "")
    content = get_html_header("Alias Learning V4")
    content += '<div class="card"><h2>تعليم Aliases من الاستعلامات الفاشلة</h2><p class="muted">لا يتم اعتماد أي alias تلقائياً. أدخل ID المنتج الصحيح، ثم يتم إضافة الاستعلام كـ alias لذلك المنتج فقط.</p></div>'
    content += '<div class="card"><h2>Misses جاهزة للمراجعة</h2><table><tr><th>Raw</th><th>Clean</th><th>Decision</th><th>Freq</th><th>اعتماد كـ alias لمنتج</th></tr>'
    for row in misses:
        raw = row.get("raw_query") or ""
        clean = row.get("clean_query") or ""
        content += f'<tr><td>{h(raw)}</td><td>{h(clean)}</td><td>{h(row.get("decision"))}</td><td>{h(row.get("freq"))}</td><td><form method="post" action="/admin/alias-learning/approve">{csrf_field(request)}<input type="hidden" name="source_query" value="{h(raw)}"><input type="hidden" name="clean_query" value="{h(clean)}"><input name="product_id" placeholder="Product ID" required><button class="btn btn-success">اعتماد</button></form></td></tr>'
    content += '</table></div>'
    content += '<div class="card"><h2>سجل قرارات Alias Learning</h2><table><tr><th>ID</th><th>Query</th><th>Product</th><th>Status</th><th>Updated</th></tr>'
    for row in suggestions:
        content += f'<tr><td>{h(row.get("id"))}</td><td>{h(row.get("source_query"))}</td><td>{h(row.get("target_product_id"))} - {h(row.get("target_product_name"))}</td><td>{h(row.get("status"))}</td><td dir="ltr">{h(row.get("updated_at"))}</td></tr>'
    content += '</table></div>'
    return HTMLResponse(content + html_footer())


@router.post("/alias-learning/approve")
async def alias_learning_approve(request: Request, source_query: str = Form(...), clean_query: str = Form(""), product_id: int = Form(...), csrf_token: str = Form(""), _: bool = Depends(require_admin)):
    require_csrf(request, csrf_token)
    database.backup_database()
    database.approve_alias_learning(source_query, clean_query, int(product_id))
    matcher.invalidate_product_cache()
    return RedirectResponse(url="/admin/alias-learning", status_code=303)


@router.get("/synonyms", response_class=HTMLResponse)
async def synonyms_page(request: Request, _: bool = Depends(require_admin)):
    rows = database.list_synonyms(300)
    content = get_html_header("Dynamic Synonyms")
    content += f'<div class="card"><h2>إضافة مرادف</h2><form method="post" action="/admin/addsynonym">{csrf_field(request)}<label>Source</label><input name="source" required placeholder="مثال: سيرا في"><label>Target</label><input name="target" required placeholder="مثال: cerave"><button class="btn btn-success" type="submit">إضافة بدون restart</button></form></div>'
    content += '<div class="card"><h2>المرادفات الديناميكية</h2><table><tr><th>ID</th><th>Source</th><th>Target</th><th>Created</th></tr>'
    for row in rows:
        content += f'<tr><td>{h(row.get("id"))}</td><td>{h(row.get("source"))}</td><td>{h(row.get("target"))}</td><td dir="ltr">{h(row.get("created_at"))}</td></tr>'
    content += '</table></div>'
    return HTMLResponse(content + html_footer())


@router.post("/addsynonym")
async def add_synonym(request: Request, source: str = Form(...), target: str = Form(...), csrf_token: str = Form(""), _: bool = Depends(require_admin)):
    require_csrf(request, csrf_token)
    database.add_synonym(source, target)
    matcher.refresh_synonym_rules()
    return RedirectResponse(url="/admin/synonyms", status_code=303)


# ---------------- V19 Company-level admin pages ----------------
@router.get("/products/import", response_class=HTMLResponse)
async def import_wizard_page(request: Request, _: bool = Depends(require_admin)):
    content = get_html_header("Import Wizard V19.1")
    content += f"""<div class='card'><h2>معالج رفع المنتجات</h2><ol><li>Upload file</li><li>Preview columns</li><li>Map columns</li><li>Quality check</li><li>Review warnings</li><li>Confirm import</li><li>Import result</li></ol><p class='muted'>يعرض النظام total rows / ready rows / review rows / duplicate count / reject reasons قبل الكتابة. الصفوف التي تحتاج مراجعة لا تُرفع كمنتجات جاهزة؛ تدخل Review Queue فقط.</p><form action='/admin/upload' method='post' enctype='multipart/form-data'>{csrf_field(request)}<input type='file' name='file' accept='.csv,.xlsx,.xls' required><label><input type='checkbox' name='replace_all' value='yes' style='width:auto'> استبدال كامل للمنتجات الجاهزة فقط</label><label><input type='checkbox' name='force_confirm' value='yes' style='width:auto'> أفهم أن الملف فيه تحذيرات وأريد استيراد الصفوف الجاهزة فقط</label><label>اكتب IMPORT_READY_ONLY للتأكيد الصريح</label><input name='force_confirm_text' placeholder='IMPORT_READY_ONLY'><button class='btn btn-success'>فحص ثم استيراد الآمن فقط</button></form></div>"""
    return HTMLResponse(content + html_footer())


@router.get("/products/review", response_class=HTMLResponse)
@router.get("/review-products", response_class=HTMLResponse)
async def review_products(request: Request, _: bool = Depends(require_admin)):
    count = database.rebuild_review_queue_from_products(limit=1000)
    rows = database.get_review_queue(300)
    content = get_html_header("Review Queue V19")
    content += f'<div class="card"><h2>Review Queue</h2><p class="muted">تم بناء قائمة مراجعة من الكتالوج الحالي: {count} عنصر.</p></div>'
    content += '<div class="card"><table><tr><th>Product ID</th><th>Name</th><th>Reason</th><th>Actions</th></tr>'
    for r in rows:
        pid = h(r.get("product_id")); name = h(r.get("name")); reason = h(r.get("review_reason"))
        content += f'<tr><td>{pid}</td><td>{name}</td><td>{reason}</td><td><a class="btn" href="/admin/products/edit/{pid}">Edit</a><form method="post" action="/admin/learning-center/action" style="display:inline">{csrf_field(request)}<input type="hidden" name="action" value="ignore"><input type="hidden" name="product_id" value="{pid}"><button class="btn btn-muted">Ignore</button></form></td></tr>'
    content += '</table></div>'
    return HTMLResponse(content + html_footer())


@router.get("/products/duplicates", response_class=HTMLResponse)
@router.get("/duplicates", response_class=HTMLResponse)
async def duplicate_resolver(request: Request, _: bool = Depends(require_admin)):
    rows = database.get_duplicate_name_groups(150)
    content = get_html_header("Duplicate Resolver V19")
    content += '<div class="card"><h2>Duplicate Resolver</h2><p class="muted">لا يتم حذف أي شيء تلقائياً. راجع المجموعات وحدد هل هي أحجام/جرعات مختلفة أو تكرار حقيقي.</p><table><tr><th>Normalized name</th><th>Count</th><th>IDs</th><th>Rows</th><th>Decision</th></tr>'
    for r in rows:
        content += f'<tr><td>{h(r.get("normalized_name"))}</td><td>{h(r.get("count"))}</td><td>{h(r.get("ids"))}</td><td>{h(r.get("names"))}</td><td><button class="btn btn-muted">Keep variants</button><button class="btn btn-danger">Merge/Delete بعد المراجعة</button></td></tr>'
    content += '</table></div>'
    return HTMLResponse(content + html_footer())


@router.get("/quality-dashboard", response_class=HTMLResponse)
async def quality_dashboard(request: Request, _: bool = Depends(require_admin)):
    products = database.load_products(); report = analyze_products(products); ai = database.get_ai_usage_summary(30); failed = database.get_failed_queries(20)
    content = get_html_header("Quality Dashboard V19")
    content += '<div class="grid">'
    content += metric_card("Catalog readiness", f"{report.get('catalog_readiness_score',0)}%", report.get("decision", ""), "green" if report.get("decision") == "ACCEPT" else "orange")
    content += metric_card("Products needing review", str(len(report.get("review_rows", []))), "Review Queue", "orange")
    content += metric_card("AI calls 30d", str(ai.get("calls", 0)), f"cost≈{ai.get('cost',0)}", "purple")
    content += metric_card("Failed queries", str(len(failed)), "Top unresolved", "red")
    content += '</div><div class="card"><h2>Top failed queries</h2><table><tr><th>Query</th><th>Status</th><th>Count</th></tr>'
    for r in failed:
        content += f'<tr><td>{h(r.get("raw_query"))}</td><td>{h(r.get("status"))}</td><td>{h(r.get("count"))}</td></tr>'
    content += '</table></div>'
    return HTMLResponse(content + html_footer())


@router.get("/learning-center", response_class=HTMLResponse)
async def learning_center(request: Request, _: bool = Depends(require_admin)):
    failed = database.get_failed_queries(100)
    content = get_html_header("Human Learning Center V19")
    content += '<div class="card"><h2>تعليم البوت من الأخطاء</h2><p class="muted">كل إجراء يحتاج اختيار بشري. لا تتم إضافة aliases أو OCR تلقائياً.</p><table><tr><th>Query</th><th>Status</th><th>Actions</th></tr>'
    for r in failed:
        q = h(r.get("raw_query"));
        content += f'<tr><td>{q}</td><td>{h(r.get("status"))}</td><td><form method="post" action="/admin/learning-center/action">{csrf_field(request)}<input type="hidden" name="query" value="{q}"><select name="action"><option value="add_alias">Add as alias</option><option value="add_ocr">Add OCR keyword</option><option value="unavailable">Mark unavailable</option><option value="generic">Mark generic</option><option value="ignore">Ignore</option></select><input name="product_id" placeholder="Product ID"><button class="btn btn-success">Apply</button></form></td></tr>'
    content += '</table></div>'
    return HTMLResponse(content + html_footer())


@router.post("/learning-center/action")
async def learning_action(request: Request, action: str = Form(""), query: str = Form(""), product_id: str = Form("0"), csrf_token: str = Form(""), _: bool = Depends(require_admin)):
    require_csrf(request, csrf_token)
    if action == "add_alias" and product_id and product_id != "0":
        database.add_product_alias(int(product_id), query)
        matcher.invalidate_product_cache()
        database.log_audit("alias_added", "admin", "product", product_id, new_value=query, ip=request.client.host if request.client else "")
    else:
        database.log_audit(f"learning_{action}", "admin", "query", query, ip=request.client.host if request.client else "")
    return RedirectResponse("/admin/learning-center", status_code=303)


@router.get("/vision-failures", response_class=HTMLResponse)
async def vision_failures(request: Request, _: bool = Depends(require_admin)):
    rows = database.get_failed_queries(100)
    content = get_html_header("Vision Failures") + '<div class="card"><h2>Unclear / low confidence image cases</h2><table><tr><th>Query</th><th>Status</th><th>Count</th></tr>'
    for r in rows:
        if "image" in str(r.get("status") or "") or "unclear" in str(r.get("status") or ""):
            content += f'<tr><td>{h(r.get("raw_query"))}</td><td>{h(r.get("status"))}</td><td>{h(r.get("count"))}</td></tr>'
    content += '</table></div>'
    return HTMLResponse(content + html_footer())


@router.get("/vision-tests", response_class=HTMLResponse)
async def vision_tests_page(request: Request, _: bool = Depends(require_admin)):
    content = get_html_header("Real Vision Tests") + '<div class="card"><h2>Real Vision Test System</h2><p>ضع الصور خارج repo داخل /opt/pricebot/vision_test_images/ ثم شغّل:</p><pre>python tools/run_real_vision_tests.py --cases vision_real_test_cases.xlsx --out vision_real_test_report.xlsx</pre><p class="muted">هذه الصفحة تعرض الإرشادات، ولا تستدعي API تلقائياً.</p></div>'
    return HTMLResponse(content + html_footer())


@router.get("/golden-tests", response_class=HTMLResponse)
async def golden_tests_page(request: Request, _: bool = Depends(require_admin)):
    return HTMLResponse(get_html_header("Golden Text Tests") + '<div class="card"><h2>Golden Text Test System</h2><pre>python tools/run_golden_text_tests.py --cases golden_text_tests.xlsx --out golden_text_test_report.xlsx</pre><p class="muted">استخدمه قبل أي تعديل Matching مستقبلي.</p></div>' + html_footer())


@router.get("/ai-usage", response_class=HTMLResponse)
async def ai_usage_page(request: Request, _: bool = Depends(require_admin)):
    legacy = database.get_ai_usage_summary(30); rows = database.get_ai_usage_logs(200)
    content = get_html_header("AI Usage & Cost")
    content += '<div class="grid">' + metric_card("Legacy AI calls", str(legacy.get("calls",0)), "30 days", "purple") + metric_card("Estimated cost", str(legacy.get("cost",0)), "configured estimate", "purple") + metric_card("V19 logs", str(len(rows)), "ai_usage_logs", "blue") + '</div>'
    content += '<div class="card"><table><tr><th>Time</th><th>Model</th><th>Purpose</th><th>Images</th><th>Cost</th><th>Success</th></tr>'
    for r in rows:
        content += f'<tr><td>{h(r.get("created_at"))}</td><td>{h(r.get("model"))}</td><td>{h(r.get("purpose"))}</td><td>{h(r.get("image_count"))}</td><td>{h(r.get("cost_estimate"))}</td><td>{h(r.get("success"))}</td></tr>'
    content += '</table></div>'
    return HTMLResponse(content + html_footer())


@router.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request, _: bool = Depends(require_admin)):
    settings = database.get_merchant_settings()
    content = get_html_header("Settings") + f'<div class="card"><form method="post" action="/admin/settings">{csrf_field(request)}'
    for key in ["pharmacy_name", "city", "working_hours", "delivery_enabled", "whatsapp_number", "welcome_message", "currency"]:
        content += f'<label>{h(key)}</label><input name="{h(key)}" value="{h(settings.get(key,""))}">'
    content += '<button class="btn btn-success">حفظ</button></form></div>'
    return HTMLResponse(content + html_footer())


@router.post("/settings")
async def save_settings(request: Request, csrf_token: str = Form(""), pharmacy_name: str = Form(""), city: str = Form(""), working_hours: str = Form(""), delivery_enabled: str = Form("false"), whatsapp_number: str = Form(""), welcome_message: str = Form(""), currency: str = Form("LYD"), _: bool = Depends(require_admin)):
    require_csrf(request, csrf_token)
    for k, v in {"pharmacy_name":pharmacy_name,"city":city,"working_hours":working_hours,"delivery_enabled":delivery_enabled,"whatsapp_number":whatsapp_number,"welcome_message":welcome_message,"currency":currency}.items():
        database.set_merchant_setting(k, v)
    database.log_audit("settings_changed", "admin", "settings", "default", new_value="merchant settings", ip=request.client.host if request.client else "")
    return RedirectResponse("/admin/settings", status_code=303)


@router.get("/security", response_class=HTMLResponse)
async def security_page(request: Request, _: bool = Depends(require_admin)):
    content = get_html_header("Security")
    content += '<div class="card"><h2>Security checklist</h2><ul><li>CSRF لكل POST</li><li>Rate limit لتسجيل الدخول</li><li>Secure cookies في production/HTTPS</li><li>Webhook signature عند ضبط META_APP_SECRET</li><li>Debug endpoints مغلقة افتراضياً</li><li>Health endpoint آمن</li></ul></div>'
    return HTMLResponse(content + html_footer())


@router.get("/logs", response_class=HTMLResponse)
async def logs_page(request: Request, _: bool = Depends(require_admin)):
    rows = database.get_audit_logs(200)
    content = get_html_header("Audit Logs") + '<div class="card"><table><tr><th>Time</th><th>Event</th><th>Actor</th><th>Entity</th><th>New</th></tr>'
    for r in rows:
        content += f'<tr><td>{h(r.get("created_at"))}</td><td>{h(r.get("event_type"))}</td><td>{h(r.get("actor"))}</td><td>{h(r.get("entity_type"))} #{h(r.get("entity_id"))}</td><td>{h(r.get("new_value"))}</td></tr>'
    content += '</table></div>'
    return HTMLResponse(content + html_footer())


@router.get("/merchants", response_class=HTMLResponse)
async def merchants_page(request: Request, _: bool = Depends(require_admin)):
    database.ensure_v19_tables(); settings = database.get_merchant_settings()
    content = get_html_header("Super Admin Merchants") + '<div class="card"><h2>Merchants foundation</h2><table><tr><th>merchant_id</th><th>صيدلية</th><th>المدينة</th><th>الحالة</th><th>الخطة</th></tr>'
    content += f'<tr><td>default</td><td>{h(settings.get("pharmacy_name"))}</td><td>{h(settings.get("city"))}</td><td>{h(settings.get("status","active"))}</td><td>{h(settings.get("subscription_plan","pilot"))}</td></tr></table><p class="muted">Foundation جاهز للتوسع بدون SaaS كامل في V19.</p></div>'
    return HTMLResponse(content + html_footer())


@router.get("/products/{product_id}/images", response_class=HTMLResponse)
async def product_images_page(product_id: int, request: Request, _: bool = Depends(require_admin)):
    rows = database.list_product_images(product_id)
    content = get_html_header("Product Image Library")
    content += f"""<div class="card"><h2>صور المنتج #{product_id}</h2><p class="muted">V19.1 يدعم رفع صورة فعلي آمن للمنتج. يتم حفظ الصور داخل product_images/ خارج قاعدة البيانات، وتسجيل hash للربط مع cache/vision لاحقاً.</p><form method="post" action="/admin/products/{product_id}/images" enctype="multipart/form-data">{csrf_field(request)}<label>رفع صورة المنتج</label><input type="file" name="image_file" accept="image/png,image/jpeg,image/webp"><label>أو image_path موجود مسبقاً داخل السيرفر</label><input name="image_path" placeholder="اختياري إذا لم ترفع ملفاً"><label>image_type</label><select name="image_type"><option>front</option><option>side</option><option>barcode</option><option>package</option></select><label>ocr_text</label><textarea name="ocr_text"></textarea><button class="btn btn-success">إضافة صورة</button></form></div>"""
    content += '<div class="card"><table><tr><th>ID</th><th>Path</th><th>Type</th><th>Hash</th><th>OCR</th></tr>'
    for r in rows:
        content += f'<tr><td>{h(r.get("id"))}</td><td>{h(r.get("image_path"))}</td><td>{h(r.get("image_type"))}</td><td dir="ltr">{h(str(r.get("image_hash") or "")[:16])}</td><td>{h(r.get("ocr_text"))}</td></tr>'
    content += '</table></div>'
    return HTMLResponse(content + html_footer())


_ALLOWED_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
_ALLOWED_IMAGE_CONTENT_TYPES = {"image/jpeg", "image/png", "image/webp", "application/octet-stream"}


def _safe_product_image_dir(product_id: int) -> pathlib.Path:
    base = pathlib.Path(os.getenv("PRICEBOT_PRODUCT_IMAGES_DIR", "product_images")).resolve()
    target = base / str(int(product_id or 0))
    target.mkdir(parents=True, exist_ok=True)
    return target


def _safe_image_filename(product_id: int, original_name: str, data: bytes) -> str:
    ext = pathlib.Path(original_name or "image.jpg").suffix.lower()
    if ext not in _ALLOWED_IMAGE_EXTS:
        ext = ".jpg"
    digest = hashlib.sha256(data).hexdigest()[:24]
    return f"product_{int(product_id or 0)}_{digest}{ext}"


@router.post("/products/{product_id}/images")
async def product_images_add(product_id: int, request: Request, image_path: str = Form(""), image_type: str = Form("front"), ocr_text: str = Form(""), csrf_token: str = Form(""), image_file: UploadFile | None = File(None), _: bool = Depends(require_admin)):
    require_csrf(request, csrf_token)
    final_path = str(image_path or "").strip()
    image_hash = ""
    if image_file is not None and image_file.filename:
        content_type = str(image_file.content_type or "").lower()
        if content_type not in _ALLOWED_IMAGE_CONTENT_TYPES:
            raise HTTPException(status_code=400, detail="Unsupported image type")
        data = await image_file.read()
        if not data:
            raise HTTPException(status_code=400, detail="Empty image")
        if len(data) > 8 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Image too large")
        filename = _safe_image_filename(product_id, image_file.filename, data)
        target = _safe_product_image_dir(product_id) / filename
        target.write_bytes(data)
        final_path = str(target)
        image_hash = hashlib.sha256(data).hexdigest()
    if not final_path:
        raise HTTPException(status_code=400, detail="Upload an image file or provide image_path")
    safe_type = image_type if image_type in {"front", "side", "barcode", "package"} else "front"
    database.add_product_image(product_id, final_path, image_hash=image_hash, ocr_text=ocr_text, image_type=safe_type)
    database.log_audit("product_image_added", "admin", "product", str(product_id), new_value=final_path, ip=request.client.host if request.client else "")
    return RedirectResponse(f"/admin/products/{product_id}/images", status_code=303)
