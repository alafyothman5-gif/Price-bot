from __future__ import annotations

import base64
import csv
import html
import hashlib
import io
import json
import os
import re
import shutil
import sqlite3
import tempfile
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from fastapi import FastAPI, File, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, PlainTextResponse, RedirectResponse
from openpyxl import load_workbook
import uvicorn

# =============================================================================
# PriceBot WhatsApp Webhook + Admin Panel
# -----------------------------------------------------------------------------
# Data files are intentionally CSV files so the owner can back them up and edit
# them easily. Secrets are never stored in this file; they must be in .env.
# =============================================================================

APP_DIR = Path(os.getenv("PRICEBOT_DATA_DIR", Path(__file__).resolve().parent))
ENV_FILE = Path(os.getenv("PRICEBOT_ENV_FILE", APP_DIR / ".env"))
PRODUCTS_FILE = Path(os.getenv("PRODUCTS_FILE", APP_DIR / "products.csv"))
ORDERS_FILE = Path(os.getenv("ORDERS_FILE", APP_DIR / "orders.csv"))
MEDIA_DIR = Path(os.getenv("PRICEBOT_MEDIA_DIR", APP_DIR / "media"))
MEMORY_FILE = Path(os.getenv("PRICEBOT_MEMORY_FILE", APP_DIR / "memory.json"))
DB_FILE = Path(os.getenv("PRICEBOT_DB_FILE", APP_DIR / "pricebot.db"))

VERIFY_TOKEN = os.getenv("VERIFY_TOKEN", "pricebot_verify_2026")
ADMIN_KEY = os.getenv("PRICEBOT_ADMIN_KEY", os.getenv("ADMIN_KEY", "PriceBotAdmin2026"))
ADMIN_NOTIFY_PHONE = os.getenv("ADMIN_NOTIFY_PHONE", "")
GRAPH_API_VERSION = os.getenv("GRAPH_API_VERSION", "v20.0")

PRODUCT_FIELDS = [
    "name", "aliases", "active_ingredient", "brand", "company", "form", "strength", "pack",
    "price", "available", "notes", "image"
]
ORDER_FIELDS = ["time", "phone", "product", "price", "available", "notes", "message", "status"]

TOKEN_ENV_NAMES = [
    "WHATSAPP_ACCESS_TOKEN",
    "WHATSAPP_TOKEN",
    "WA_ACCESS_TOKEN",
    "WA_TOKEN",
    "WHATSAPP_API_TOKEN",
    "WHATSAPP_BEARER_TOKEN",
    "WHATSAPP_PERMANENT_TOKEN",
    "META_WHATSAPP_ACCESS_TOKEN",
    "META_WHATSAPP_TOKEN",
    "META_ACCESS_TOKEN",
    "META_TOKEN",
    "FACEBOOK_ACCESS_TOKEN",
    "FB_ACCESS_TOKEN",
    "GRAPH_API_TOKEN",
    "CLOUD_API_TOKEN",
    "PRICEBOT_TOKEN",
    "ACCESS_TOKEN",
]

PHONE_ID_ENV_NAMES = [
    "WHATSAPP_PHONE_NUMBER_ID",
    "WA_PHONE_NUMBER_ID",
    "PHONE_NUMBER_ID",
    "META_PHONE_NUMBER_ID",
    "META_WHATSAPP_PHONE_NUMBER_ID",
]

app = FastAPI(title="PriceBot", version="3.0.0")
LAST_PRODUCT: Dict[str, dict] = {}
PENDING_SUGGESTION: Dict[str, dict] = {}
PENDING_OPTIONS: Dict[str, dict] = {}


def load_dotenv_file() -> None:
    """Load .env values into os.environ without overwriting already set vars."""
    if not ENV_FILE.exists():
        return
    try:
        for raw_line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value
    except Exception as exc:
        print(f"ENV LOAD WARNING: {exc}", flush=True)


def read_env_file_value(names: Iterable[str]) -> str:
    """Read one of the requested names directly from .env, preferring top names."""
    names = list(names)
    values: Dict[str, str] = {}
    if ENV_FILE.exists():
        try:
            for raw_line in ENV_FILE.read_text(encoding="utf-8").splitlines():
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                values[key.strip()] = value.strip().strip('"').strip("'")
        except Exception as exc:
            print(f"ENV READ WARNING: {exc}", flush=True)
    for name in names:
        if values.get(name):
            return values[name]
    for name in names:
        if os.getenv(name):
            return os.getenv(name, "")
    return ""


def clean_token(token: str) -> str:
    token = (token or "").strip()
    if token.lower().startswith("bearer "):
        token = token[7:].strip()
    return token.strip().strip('"').strip("'")


def get_access_token() -> str:
    return clean_token(read_env_file_value(TOKEN_ENV_NAMES))


def get_phone_number_id() -> str:
    return read_env_file_value(PHONE_ID_ENV_NAMES).strip()


def get_config(name: str, default: str = "") -> str:
    """Read optional business settings from environment or .env without exposing secrets."""
    value = os.getenv(name, "").strip()
    if value:
        return value
    if ENV_FILE.exists():
        try:
            for raw_line in ENV_FILE.read_text(encoding="utf-8").splitlines():
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, raw_value = line.split("=", 1)
                if key.strip() == name:
                    return raw_value.strip().strip('"').strip("'")
        except Exception as exc:
            print(f"CONFIG READ WARNING: {exc}", flush=True)
    return default


def business_name() -> str:
    return get_config("PHARMACY_NAME", "صيدلية بدر البشرية")


def business_city() -> str:
    return get_config("PHARMACY_CITY", "أجدابيا")


def business_hours() -> str:
    return get_config("PHARMACY_HOURS", "24 ساعة")


def delivery_enabled() -> bool:
    value = get_config("DELIVERY_AVAILABLE", "no").strip().lower()
    return value in {"1", "true", "yes", "y", "نعم", "متوفر"}


def delivery_text() -> str:
    if delivery_enabled():
        return get_config("DELIVERY_TEXT", "التوصيل متوفر")
    return get_config("DELIVERY_TEXT", "التوصيل غير متوفر حالياً")


def admin_notify_phone() -> str:
    """Optional WhatsApp number that receives internal admin alerts."""
    return normalize_phone(get_config("ADMIN_NOTIFY_PHONE", ""))


def public_base_url() -> str:
    return get_config("PUBLIC_BASE_URL", "https://46.101.148.246.sslip.io").rstrip("/")


def normalize_phone(phone: str) -> str:
    """Keep digits only, convert 00 prefix, and remove + for WhatsApp Cloud API."""
    phone = str(phone or "").strip()
    phone = phone.replace("+", "").replace(" ", "").replace("-", "")
    phone = re.sub(r"\D+", "", phone)
    if phone.startswith("00"):
        phone = phone[2:]
    return phone


# =============================================================================
# Optional real AI integration with multi-provider fallback
# =============================================================================
def config_bool(name: str, default: bool = False) -> bool:
    raw = get_config(name, "1" if default else "0").strip().lower()
    return raw in {"1", "true", "yes", "y", "on", "نعم", "مفعل"}


def split_secret_list(raw: str) -> List[str]:
    """Parse keys from .env values or admin textareas without exposing them."""
    raw = (raw or "").replace("\r", "\n")
    parts: List[str] = []
    for chunk in re.split(r"[\n,;|]+", raw):
        value = clean_token(chunk)
        if value and value not in parts:
            parts.append(value)
    return parts


def join_secret_list(raw: str) -> str:
    return "||".join(split_secret_list(raw))


def ai_provider_order() -> List[str]:
    raw = get_config("AI_PROVIDER_ORDER", "gemini,openrouter,groq")
    supported = {"openrouter", "gemini", "groq", "openai", "custom"}
    out: List[str] = []
    for item in re.split(r"[,;|\n]+", raw):
        name = item.strip().lower()
        if name in supported and name not in out:
            out.append(name)
    return out or ["gemini", "openrouter", "groq"]


def provider_key_env(provider: str) -> str:
    return {
        "openrouter": "AI_OPENROUTER_KEYS",
        "gemini": "AI_GEMINI_KEYS",
        "groq": "AI_GROQ_KEYS",
        "openai": "AI_OPENAI_KEYS",
        "custom": "AI_CUSTOM_KEYS",
    }.get(provider, "AI_API_KEY")


def ai_keys_for_provider(provider: str) -> List[str]:
    keys = split_secret_list(get_config(provider_key_env(provider), ""))
    # Backward compatibility with the old single-key setting.
    legacy_provider = get_config("AI_PROVIDER", "openrouter").strip().lower() or "openrouter"
    legacy_key = clean_token(get_config("AI_API_KEY", ""))
    if legacy_key and provider == legacy_provider and legacy_key not in keys:
        keys.append(legacy_key)
    return keys


def any_ai_key_saved() -> bool:
    return any(ai_keys_for_provider(provider) for provider in ai_provider_order())


def ai_enabled() -> bool:
    return config_bool("AI_ENABLED", False) and any_ai_key_saved()


def ai_provider() -> str:
    # Compatibility for old UI/health; the actual call uses ai_provider_order().
    order = ai_provider_order()
    return order[0] if order else "openrouter"


def ai_default_model_for(provider: str) -> str:
    if provider == "gemini":
        return "gemini-2.5-flash-lite"
    if provider == "groq":
        return "llama-3.1-8b-instant"
    if provider == "openai":
        return "gpt-4o-mini"
    return "openrouter/free"


def ai_model_for(provider: str) -> str:
    specific = {
        "openrouter": "AI_OPENROUTER_MODEL",
        "gemini": "AI_GEMINI_MODEL",
        "groq": "AI_GROQ_MODEL",
        "openai": "AI_OPENAI_MODEL",
        "custom": "AI_CUSTOM_MODEL",
    }.get(provider, "AI_MODEL")
    return get_config(specific, "").strip() or get_config("AI_MODEL", "").strip() or ai_default_model_for(provider)


def ai_model() -> str:
    return ai_model_for(ai_provider())


def ai_api_key() -> str:
    # Compatibility helper: first key in the first available provider.
    for provider in ai_provider_order():
        keys = ai_keys_for_provider(provider)
        if keys:
            return keys[0]
    return ""


def masked_count_for(provider: str) -> str:
    count = len(ai_keys_for_provider(provider))
    return f"{count} مفتاح" if count else "لا يوجد"


def ai_status_text() -> str:
    counts = [f"{provider}: {masked_count_for(provider)}" for provider in ai_provider_order()]
    if ai_enabled():
        return "مفعل — " + " | ".join(counts)
    return "غير مفعل — " + " | ".join(counts)


def ai_default_model() -> str:
    return ai_default_model_for(ai_provider())


def ai_endpoint_for(provider: str) -> str:
    if provider == "gemini":
        return ""
    if provider == "openai":
        return get_config("AI_OPENAI_BASE_URL", get_config("AI_BASE_URL", "https://api.openai.com/v1/chat/completions"))
    if provider == "groq":
        return get_config("AI_GROQ_BASE_URL", "https://api.groq.com/openai/v1/chat/completions")
    if provider == "custom":
        return get_config("AI_CUSTOM_BASE_URL", get_config("AI_BASE_URL", "")).strip()
    return get_config("AI_OPENROUTER_BASE_URL", get_config("AI_BASE_URL", "https://openrouter.ai/api/v1/chat/completions"))


def ai_provider_display_name(provider: str) -> str:
    return {
        "openrouter": "OpenRouter",
        "gemini": "Gemini",
        "groq": "Groq",
        "openai": "OpenAI",
        "custom": "Custom",
    }.get(provider, provider)


load_dotenv_file()


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalize(text: str) -> str:
    text = (text or "").strip().lower()
    replacements = {
        "أ": "ا",
        "إ": "ا",
        "آ": "ا",
        "ٱ": "ا",
        "ى": "ي",
        "ة": "ه",
        "ؤ": "و",
        "ئ": "ي",
        "ـ": "",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    text = re.sub(r"[\u064b-\u065f\u0670]", "", text)  # Arabic tashkeel
    text = re.sub(r"[^\w\s\u0600-\u06FF]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text




# =============================================================================
# SQLite database layer
# -----------------------------------------------------------------------------
# The bot now stores products, orders, memory/cache, corrections, and AI counters
# in a single local SQLite database: pricebot.db. Existing CSV/JSON files are
# imported automatically on first use and then kept as backup/export mirrors.
# =============================================================================
def db_connect() -> sqlite3.Connection:
    DB_FILE.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_FILE), timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def ensure_sqlite_db() -> None:
    with db_connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                aliases TEXT DEFAULT '',
                active_ingredient TEXT DEFAULT '',
                brand TEXT DEFAULT '',
                company TEXT DEFAULT '',
                form TEXT DEFAULT '',
                strength TEXT DEFAULT '',
                pack TEXT DEFAULT '',
                price TEXT DEFAULT '',
                available TEXT DEFAULT '',
                notes TEXT DEFAULT '',
                image TEXT DEFAULT '',
                created_at TEXT DEFAULT '',
                updated_at TEXT DEFAULT ''
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_products_name ON products(name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_products_active ON products(active_ingredient)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_products_strength ON products(strength)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                time TEXT DEFAULT '',
                phone TEXT DEFAULT '',
                product TEXT DEFAULT '',
                price TEXT DEFAULT '',
                available TEXT DEFAULT '',
                notes TEXT DEFAULT '',
                message TEXT DEFAULT '',
                status TEXT DEFAULT 'new'
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_orders_status ON orders(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_orders_phone ON orders(phone)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS memory_entries (
                category TEXT NOT NULL,
                key TEXT NOT NULL,
                value_json TEXT NOT NULL,
                created_at TEXT DEFAULT '',
                updated_at TEXT DEFAULT '',
                hits INTEGER DEFAULT 0,
                PRIMARY KEY(category, key)
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_memory_category ON memory_entries(category)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS meta (
                key TEXT PRIMARY KEY,
                value TEXT DEFAULT ''
            )
            """
        )
        conn.commit()


def db_count(table: str) -> int:
    ensure_sqlite_db()
    with db_connect() as conn:
        return int(conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0])


def _csv_row_get(row: dict, *names: str, default: str = '') -> str:
    for name in names:
        val = row.get(name)
        if val is not None and str(val).strip():
            return str(val).strip()
    normalized = {normalize(str(k)): v for k, v in row.items()}
    for name in names:
        val = normalized.get(normalize(name))
        if val is not None and str(val).strip():
            return str(val).strip()
    return default


def _product_from_csv_row(row: dict) -> dict:
    return {
        "name": _csv_row_get(row, "name", "product", "product_name", "اسم", "الاسم", "اسم المنتج"),
        "aliases": _csv_row_get(row, "aliases", "alias", "أسماء بديلة", "اسماء بديلة", "بدائل"),
        "active_ingredient": _csv_row_get(row, "active_ingredient", "ingredient", "generic", "المادة الفعالة", "ماده فعاله", "المادة", "ماده"),
        "brand": _csv_row_get(row, "brand", "trade_name", "ماركة", "الماركة", "براند", "اسم تجاري"),
        "company": _csv_row_get(row, "company", "manufacturer", "origin", "country", "الشركة", "الشركه", "المنشأ", "بلد", "البلد"),
        "form": _csv_row_get(row, "form", "dosage_form", "shape", "الشكل", "الشكل الدوائي", "نوع", "النوع"),
        "strength": _csv_row_get(row, "strength", "concentration", "dose", "التركيز", "جرعة", "عيار"),
        "pack": _csv_row_get(row, "pack", "package", "pack_size", "العبوة", "عبوة", "التعبئة"),
        "price": _csv_row_get(row, "price", "سعر", "السعر"),
        "available": _csv_row_get(row, "available", "availability", "stock", "توفر", "التوفر", "الحالة", default="متوفر") or "متوفر",
        "notes": _csv_row_get(row, "notes", "note", "ملاحظات", "ملاحظة"),
        "image": _csv_row_get(row, "image", "image_url", "photo", "صورة", "رابط الصورة"),
    }


def read_products_from_csv_file() -> List[dict]:
    if not PRODUCTS_FILE.exists():
        return []
    products: List[dict] = []
    try:
        with PRODUCTS_FILE.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                item = _product_from_csv_row(row)
                if item.get("name"):
                    products.append(item)
    except Exception as exc:
        print(f"PRODUCT CSV IMPORT WARNING: {exc}", flush=True)
    return products


def read_orders_from_csv_file() -> List[dict]:
    if not ORDERS_FILE.exists():
        return []
    try:
        with ORDERS_FILE.open("r", encoding="utf-8-sig", newline="") as f:
            return [{field: row.get(field, '') for field in ORDER_FIELDS} for row in csv.DictReader(f)]
    except Exception as exc:
        print(f"ORDER CSV IMPORT WARNING: {exc}", flush=True)
        return []


def insert_product_rows(conn: sqlite3.Connection, products: List[dict]) -> None:
    now = now_str()
    rows = []
    for item in products:
        name = str(item.get("name", "")).strip()
        if not name:
            continue
        rows.append(tuple(str(item.get(field, "")).strip() for field in PRODUCT_FIELDS) + (now, now))
    if rows:
        conn.executemany(
            f"INSERT INTO products ({','.join(PRODUCT_FIELDS)},created_at,updated_at) VALUES ({','.join(['?'] * (len(PRODUCT_FIELDS)+2))})",
            rows,
        )


def insert_order_rows(conn: sqlite3.Connection, orders: List[dict]) -> None:
    rows = []
    for row in orders:
        rows.append(tuple(str(row.get(field, "")).strip() for field in ORDER_FIELDS))
    if rows:
        conn.executemany(
            f"INSERT INTO orders ({','.join(ORDER_FIELDS)}) VALUES ({','.join(['?'] * len(ORDER_FIELDS))})",
            rows,
        )


def migrate_legacy_files_to_db() -> None:
    ensure_sqlite_db()
    with db_connect() as conn:
        if int(conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]) == 0:
            products = read_products_from_csv_file()
            if products:
                insert_product_rows(conn, products)
                print(f"DB MIGRATE: imported {len(products)} products from CSV", flush=True)
        if int(conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0]) == 0:
            orders = read_orders_from_csv_file()
            if orders:
                insert_order_rows(conn, orders)
                print(f"DB MIGRATE: imported {len(orders)} orders from CSV", flush=True)
        if int(conn.execute("SELECT COUNT(*) FROM memory_entries").fetchone()[0]) == 0 and MEMORY_FILE.exists():
            try:
                data = json.loads(MEMORY_FILE.read_text(encoding="utf-8"))
                if isinstance(data, dict):
                    _write_memory_to_db(conn, data, clear_existing=False)
                    print("DB MIGRATE: imported memory.json", flush=True)
            except Exception as exc:
                print(f"MEMORY JSON IMPORT WARNING: {exc}", flush=True)
        conn.commit()


def mirror_products_to_csv(products: List[dict]) -> None:
    PRODUCTS_FILE.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix="products_", suffix=".csv", dir=str(PRODUCTS_FILE.parent))
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        with tmp_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=PRODUCT_FIELDS)
            writer.writeheader()
            for item in products:
                if str(item.get("name", "")).strip():
                    writer.writerow({field: str(item.get(field, "")).strip() for field in PRODUCT_FIELDS})
        tmp_path.replace(PRODUCTS_FILE)
    finally:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)


def mirror_orders_to_csv(orders: List[dict]) -> None:
    ORDERS_FILE.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix="orders_", suffix=".csv", dir=str(ORDERS_FILE.parent))
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        with tmp_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=ORDER_FIELDS)
            writer.writeheader()
            for row in orders:
                writer.writerow({field: row.get(field, "") for field in ORDER_FIELDS})
        tmp_path.replace(ORDERS_FILE)
    finally:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)


def _write_memory_to_db(conn: sqlite3.Connection, data: dict, clear_existing: bool = True) -> None:
    if clear_existing:
        conn.execute("DELETE FROM memory_entries")
    now = now_str()
    rows = []
    for category in ["query_cache", "product_alias_memory", "image_cache", "admin_corrections", "ai_usage"]:
        section = data.get(category, {}) if isinstance(data, dict) else {}
        if not isinstance(section, dict):
            continue
        for key, value in section.items():
            rows.append((category, str(key), json.dumps(value, ensure_ascii=False), now, now, 0))
    if rows:
        conn.executemany(
            "INSERT OR REPLACE INTO memory_entries(category,key,value_json,created_at,updated_at,hits) VALUES (?,?,?,?,?,?)",
            rows,
        )


def db_stats() -> dict:
    ensure_sqlite_db()
    migrate_legacy_files_to_db()
    with db_connect() as conn:
        return {
            "db_file": str(DB_FILE),
            "db_size": DB_FILE.stat().st_size if DB_FILE.exists() else 0,
            "products": int(conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]),
            "orders": int(conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0]),
            "memory_entries": int(conn.execute("SELECT COUNT(*) FROM memory_entries").fetchone()[0]),
        }

# =============================================================================
# Persistent memory/cache to reduce AI API usage
# =============================================================================

def empty_memory() -> dict:
    return {
        "version": 3,
        "storage": "sqlite",
        "query_cache": {},
        "product_alias_memory": {},
        "image_cache": {},
        "admin_corrections": {},
        "ai_usage": {},
    }


def memory_get_entry(category: str, key: str) -> object:
    ensure_sqlite_db()
    if not key:
        return None
    with db_connect() as conn:
        row = conn.execute(
            "SELECT value_json, hits FROM memory_entries WHERE category=? AND key=?",
            (category, key),
        ).fetchone()
        if not row:
            return None
        try:
            value = json.loads(row["value_json"])
        except Exception:
            value = row["value_json"]
        conn.execute(
            "UPDATE memory_entries SET hits=COALESCE(hits,0)+1, updated_at=? WHERE category=? AND key=?",
            (now_str(), category, key),
        )
        conn.commit()
        if isinstance(value, dict):
            value["hits"] = int(row["hits"] or 0) + 1
            value["last_hit"] = now_str()
        return value


def memory_put_entry(category: str, key: str, value: object) -> None:
    ensure_sqlite_db()
    if not key:
        return
    now = now_str()
    with db_connect() as conn:
        conn.execute(
            """
            INSERT INTO memory_entries(category,key,value_json,created_at,updated_at,hits)
            VALUES (?,?,?,?,?,0)
            ON CONFLICT(category,key) DO UPDATE SET
              value_json=excluded.value_json,
              updated_at=excluded.updated_at
            """,
            (category, key, json.dumps(value, ensure_ascii=False), now, now),
        )
        conn.commit()


def memory_delete_category(category: str) -> None:
    ensure_sqlite_db()
    with db_connect() as conn:
        conn.execute("DELETE FROM memory_entries WHERE category=?", (category,))
        conn.commit()


def load_memory() -> dict:
    ensure_sqlite_db()
    migrate_legacy_files_to_db()
    data = empty_memory()
    try:
        with db_connect() as conn:
            for row in conn.execute("SELECT category,key,value_json,hits FROM memory_entries ORDER BY updated_at DESC LIMIT 2000"):
                category = row["category"]
                if category not in data or not isinstance(data[category], dict):
                    continue
                try:
                    value = json.loads(row["value_json"])
                except Exception:
                    value = row["value_json"]
                if isinstance(value, dict):
                    value.setdefault("hits", int(row["hits"] or 0))
                data[category][row["key"]] = value
    except Exception as exc:
        print(f"MEMORY DB LOAD WARNING: {exc}", flush=True)
    return data


def save_memory(data: dict) -> None:
    """Admin-level save only. Runtime message cache uses direct SQLite upserts."""
    ensure_sqlite_db()
    with db_connect() as conn:
        _write_memory_to_db(conn, data, clear_existing=True)
        conn.commit()
    # Keep JSON mirror only when admin explicitly changes memory, not on every message.
    try:
        MEMORY_FILE.parent.mkdir(parents=True, exist_ok=True)
        MEMORY_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as exc:
        print(f"MEMORY JSON MIRROR WARNING: {exc}", flush=True)


def memory_query_key(text: str) -> str:
    q = extract_product_query(text or "") if "extract_product_query" in globals() else (text or "")
    q = normalize(q or text or "")
    return q[:180]


def find_product_by_name_or_alias(name: str) -> Optional[dict]:
    n = normalize(name or "")
    if not n:
        return None
    for item in load_products():
        keys = [item.get("name", "")] + split_aliases(item.get("aliases", ""))
        for key in keys:
            if normalize(key) == n:
                return item
    item = find_product(name)
    return item


def memory_get_admin_correction(text: str) -> Optional[dict]:
    key = memory_query_key(text)
    product_name = memory_get_entry("admin_corrections", key) or memory_get_entry("product_alias_memory", key)
    if isinstance(product_name, dict):
        product_name = product_name.get("product_name") or product_name.get("value")
    if product_name:
        item = find_product_by_name_or_alias(str(product_name))
        if item:
            return item
    return None


def memory_get_query_response(text: str) -> dict:
    entry = memory_get_entry("query_cache", memory_query_key(text))
    return entry if isinstance(entry, dict) else {}


def memory_remember_product_query(text: str, item: dict, source: str = "local") -> None:
    key = memory_query_key(text)
    product_name = str(item.get("name", "")).strip()
    if not key or not product_name:
        return
    memory_put_entry("query_cache", key, {
        "type": "product",
        "product_name": product_name,
        "source": source,
        "created_at": now_str(),
    })
    memory_put_entry("product_alias_memory", key, product_name)


def memory_remember_options_query(text: str, items: List[dict], source: str = "local") -> None:
    key = memory_query_key(text)
    names = [str(item.get("name", "")).strip() for item in items if str(item.get("name", "")).strip()]
    if not key or not names:
        return
    memory_put_entry("query_cache", key, {
        "type": "options",
        "product_names": names[:8],
        "source": source,
        "created_at": now_str(),
    })


def memory_items_from_names(names: List[str]) -> List[dict]:
    out = []
    seen = set()
    for name in names:
        item = find_product_by_name_or_alias(name)
        if not item:
            continue
        k = normalize(item.get("name", ""))
        if k and k not in seen:
            out.append(item)
            seen.add(k)
    return out


def memory_lookup_image(sha: str) -> dict:
    entry = memory_get_entry("image_cache", sha)
    return entry if isinstance(entry, dict) else {}


def memory_remember_image_product(sha: str, item: dict, source: str = "gemini_vision") -> None:
    if not sha or not item:
        return
    memory_put_entry("image_cache", sha, {
        "type": "product",
        "product_name": item.get("name", ""),
        "source": source,
        "created_at": now_str(),
    })


def memory_remember_image_review(sha: str, reason: str = "review") -> None:
    if not sha:
        return
    memory_put_entry("image_cache", sha, {
        "type": "review",
        "reason": reason,
        "source": "image_review",
        "created_at": now_str(),
    })


def memory_record_ai_usage(kind: str, provider: str = "unknown") -> None:
    day = datetime.now().strftime("%Y-%m-%d")
    key = f"{day}:{kind}:{provider or 'unknown'}"
    entry = memory_get_entry("ai_usage", key)
    count = 0
    if isinstance(entry, dict):
        count = int(entry.get("count", 0) or 0)
    memory_put_entry("ai_usage", key, {
        "day": day,
        "kind": kind,
        "provider": provider or "unknown",
        "count": count + 1,
        "updated_at": now_str(),
    })


def memory_stats() -> dict:
    ensure_sqlite_db()
    today = datetime.now().strftime("%Y-%m-%d")
    with db_connect() as conn:
        def c(cat):
            return int(conn.execute("SELECT COUNT(*) FROM memory_entries WHERE category=?", (cat,)).fetchone()[0])
        usage_rows = conn.execute("SELECT key,value_json FROM memory_entries WHERE category='ai_usage' AND key LIKE ?", (today + ":%",)).fetchall()
    usage = {}
    for row in usage_rows:
        try:
            value = json.loads(row["value_json"])
            label = f"{value.get('kind','text')}:{value.get('provider','unknown')}"
            usage[label] = int(value.get("count", 0) or 0)
        except Exception:
            pass
    return {
        "query_cache": c("query_cache"),
        "product_alias_memory": c("product_alias_memory"),
        "image_cache": c("image_cache"),
        "admin_corrections": c("admin_corrections"),
        "ai_usage_today": usage,
    }


def split_aliases(aliases: str) -> List[str]:
    if not aliases:
        return []
    parts = re.split(r"[|,،;؛\n]+", aliases)
    return [part.strip() for part in parts if part and part.strip()]


def ensure_csv_file(path: Path, fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()


def ensure_products_file() -> None:
    ensure_csv_file(PRODUCTS_FILE, PRODUCT_FIELDS)


def ensure_orders_file() -> None:
    ensure_csv_file(ORDERS_FILE, ORDER_FIELDS)


def _row_get(row: dict, *names: str, default: str = "") -> str:
    return _csv_row_get(row, *names, default=default)


def product_dict_from_db_row(row: sqlite3.Row) -> dict:
    item = {field: str(row[field] or "") for field in PRODUCT_FIELDS}
    keyword_parts = [
        item["name"], item["aliases"], item["active_ingredient"], item["brand"],
        item["company"], item["form"], item["strength"], item["pack"], item["notes"],
    ]
    keywords: List[str] = []
    for part in keyword_parts:
        if not part:
            continue
        if part == item["aliases"]:
            keywords.extend(split_aliases(part))
        else:
            keywords.append(part)
    item["keywords"] = [x for x in keywords if str(x).strip()]
    return item


def load_products() -> List[dict]:
    ensure_sqlite_db()
    migrate_legacy_files_to_db()
    with db_connect() as conn:
        rows = conn.execute(f"SELECT {','.join(PRODUCT_FIELDS)} FROM products ORDER BY id ASC").fetchall()
    return [product_dict_from_db_row(row) for row in rows]


def save_products(products: List[dict]) -> None:
    ensure_sqlite_db()
    with db_connect() as conn:
        conn.execute("DELETE FROM products")
        insert_product_rows(conn, products)
        conn.commit()
    mirror_products_to_csv(products)


def backup_file(path: Path) -> Optional[Path]:
    if not path.exists():
        return None
    backup = path.with_name(f"{path.name}.bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    shutil.copy2(path, backup)
    return backup


def product_key(name: str) -> str:
    return normalize(name)


def upsert_product(new_item: dict) -> Tuple[List[dict], str]:
    products = load_products()
    key = product_key(new_item.get("name", ""))
    for item in products:
        if product_key(item.get("name", "")) == key:
            item.update(new_item)
            save_products(products)
            return products, "updated"
    products.append(new_item)
    save_products(products)
    return products, "added"



def extract_product_query(text: str) -> str:
    """Turn natural customer wording into a product-focused query without removing useful form/strength words."""
    q = normalize(text)
    if not q:
        return ""

    phrase_replacements = [
        "السلام عليكم", "السلام", "مرحبا", "اهلا", "اهلين", "لو سمحت", "من فضلك", "بالله",
        "كم سعر", "شن سعر", "شنو سعر", "بكم", "قداش", "السعر", "سعر",
        "عندكم", "موجود", "موجوده", "متوفر", "متوفره", "هل يوجد", "هل عندكم",
        "اريد", "نريد", "نبي", "ابي", "ابغى", "ممكن", "عطيني", "احتاج",
        "الصيدليه", "الصيدلية", "بدر", "البشرية", "البشريه",
        "do you have", "have", "price", "how much", "need", "want", "please",
    ]
    for phrase in sorted(phrase_replacements, key=len, reverse=True):
        q = q.replace(normalize(phrase), " ")

    # Keep useful identifiers like 250/500/400, extra/advance, شراب/تحاميل/كريم.
    q = re.sub(r"\s+", " ", q).strip()
    return q


def search_tokens(text: str) -> List[str]:
    q = normalize(text or "")
    # normalize common strengths so 400 matches 400mg / 400 مجم / 400 ملغ.
    q = re.sub(r"(\d+)\s*(mg|مجم|ملغ|مغ|ملجم|g|جم|mcg|ميكرو)", r"\1", q)
    raw = [t for t in q.split() if len(t) > 1]
    out: List[str] = []
    for t in raw:
        if t not in out:
            out.append(t)
        # Also extract digits from mixed tokens like 400mg.
        m = re.search(r"\d+", t)
        if m and m.group(0) not in out:
            out.append(m.group(0))
    return out


def product_full_search_text(item: dict) -> str:
    parts = []
    for field in ["name", "aliases", "active_ingredient", "brand", "company", "form", "strength", "pack", "notes"]:
        val = str(item.get(field, "") or "").strip()
        if val:
            parts.append(val)
    return " ".join(parts)


def match_score(query: str, keyword: str) -> float:
    q = normalize(query)
    k = normalize(keyword)
    if not q or not k:
        return 0.0
    if q == k:
        return 1.0

    q_tokens = search_tokens(q)
    k_tokens = search_tokens(k)
    q_set, k_set = set(q_tokens), set(k_tokens)

    # If the customer query is more specific and fully contained in the product text,
    # prefer that product over a shorter generic one. Example: بروفين 400 > بروفين.
    if q_set and k_set:
        shared = len(q_set & k_set)
        if q_set <= k_set:
            return 0.99
        if k_set <= q_set:
            # keyword is generic inside a longer query; useful but not the best if a specific option exists.
            return 0.88 if len(k_set) < len(q_set) else 0.96
        overlap = shared / max(len(q_set), len(k_set), 1)
        if overlap >= 0.75:
            return 0.93
        if overlap >= 0.50:
            return 0.78

    if q in k and len(q) >= 3:
        return 0.94
    if k in q and len(k) >= 3:
        return 0.86

    return SequenceMatcher(None, q, k).ratio()


def ranked_products(text: str) -> List[Tuple[float, dict]]:
    products = load_products()
    candidates = [text, extract_product_query(text)]
    seen_candidates = set()
    candidates = [c for c in candidates if c and not (normalize(c) in seen_candidates or seen_candidates.add(normalize(c)))]

    ranked: List[Tuple[float, dict]] = []
    for item in products:
        best_score = 0.0
        full_text = product_full_search_text(item)
        for candidate in candidates:
            best_score = max(best_score, match_score(candidate, full_text))
            # Also compare against individual strong fields for exact aliases/trade names.
            for keyword in item.get("keywords", [item.get("name", "")]):
                best_score = max(best_score, match_score(candidate, keyword))
        if best_score > 0:
            ranked.append((best_score, item))
    ranked.sort(key=lambda x: x[0], reverse=True)
    return ranked


def category_suggestions(text: str, limit: int = 4) -> List[dict]:
    q = normalize(text)
    products = load_products()
    hints = {
        "صداع": ["بنادول", "فيفادول", "بروفين", "باراسيتامول", "panadol", "paracetamol"],
        "مسكن": ["بنادول", "فيفادول", "بروفين", "كتافلام", "ديكلوفيناك"],
        "الم": ["بنادول", "فيفادول", "بروفين", "كتافلام"],
        "حراره": ["بنادول", "فيفادول", "باراسيتامول", "panadol"],
        "مضاد": ["أموكسيل", "اموكسيل", "أوجمنتين", "اوجمنتين", "أزيثرومايسين"],
        "حموضه": ["جلوسيد", "اوميبرازول", "omeprazole"],
        "معده": ["جلوسيد", "موتيليوم", "سماكتا"],
        "اسهال": ["سماكتا"],
        "ضغط": ["أملور", "كونكور", "كابوتين", "لازيكس"],
    }
    wanted: List[str] = []
    for key, names in hints.items():
        if normalize(key) in q:
            wanted.extend(names)
    if not wanted:
        return []

    result: List[dict] = []
    for item in products:
        haystack = normalize(" ".join([item.get("name", ""), item.get("aliases", ""), item.get("notes", "")]))
        if any(normalize(name) in haystack for name in wanted):
            result.append(item)
        if len(result) >= limit:
            break
    return result


def find_product(text: str) -> Optional[dict]:
    ranked = ranked_products(text)
    if ranked and ranked[0][0] >= 0.76:
        return ranked[0][1]
    return None


def suggested_products(text: str, limit: int = 4) -> List[dict]:
    category = category_suggestions(text, limit=limit)
    if category:
        return category[:limit]
    ranked = [item for score, item in ranked_products(text) if score >= 0.48]
    # De-duplicate by normalized product name.
    out: List[dict] = []
    seen = set()
    for item in ranked:
        key = normalize(item.get("name", ""))
        if key and key not in seen:
            out.append(item)
            seen.add(key)
        if len(out) >= limit:
            break
    return out



def json_from_model_text(content: str) -> dict:
    """Extract a JSON object from an LLM answer safely."""
    content = (content or "").strip()
    if not content:
        return {}
    content = content.replace("```json", "```").strip()
    if "```" in content:
        parts = content.split("```")
        # Prefer the largest fenced block.
        content = max(parts, key=len).strip()
    start = content.find("{")
    end = content.rfind("}")
    if start >= 0 and end > start:
        content = content[start:end + 1]
    try:
        data = json.loads(content)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def compact_catalog_for_ai(text: str, limit: int = 80) -> List[dict]:
    """Send a compact product catalog to the AI without leaking orders or secrets."""
    products = load_products()
    ranked = ranked_products(text)
    chosen: List[dict] = []
    seen = set()

    # Start with likely local matches.
    for score, item in ranked:
        key = normalize(item.get("name", ""))
        if key and key not in seen:
            chosen.append(item)
            seen.add(key)
        if len(chosen) >= min(limit, 30):
            break

    # Add category suggestions.
    for item in category_suggestions(text, limit=20):
        key = normalize(item.get("name", ""))
        if key and key not in seen:
            chosen.append(item)
            seen.add(key)

    # If still little context, add beginning of catalog.
    for item in products:
        key = normalize(item.get("name", ""))
        if key and key not in seen:
            chosen.append(item)
            seen.add(key)
        if len(chosen) >= limit:
            break

    compact = []
    for item in chosen[:limit]:
        compact.append({
            "name": item.get("name", ""),
            "aliases": item.get("aliases", ""),
            "active_ingredient": item.get("active_ingredient", ""),
            "company": item.get("company", ""),
            "form": item.get("form", ""),
            "strength": item.get("strength", ""),
            "pack": item.get("pack", ""),
            "notes": item.get("notes", ""),
        })
    return compact


def call_ai_once(provider: str, key: str, model: str, user_text: str, catalog: List[dict]) -> dict:
    """Call one provider/key. Exceptions are handled by the caller."""
    system_prompt = (
        "أنت مساعد واتساب لصيدلية في ليبيا. مهمتك فقط فهم رسالة الزبون وتحويلها إلى JSON. "
        "لا تعطِ تشخيصاً طبياً ولا جرعات ولا علاجاً. لا تخترع منتجات غير موجودة في الكتالوج. "
        "لو الزبون يسأل عن سعر/توفر منتج، استخرج اسم المنتج المقصود. "
        "لو يسأل عن جرعة/استعمال/هل يناسب حامل أو طفل، اجعل intent=medical_advice. "
        "لو قال نعم أو حجز، intent=reservation_yes. لو قال لا أو إلغاء، intent=reservation_no. "
        "لو كانت تحية فقط، intent=greeting. "
        "أرجع JSON فقط بهذا الشكل: "
        "{\"intent\":\"product_lookup|greeting|reservation_yes|reservation_no|medical_advice|unknown\","
        "\"product_query\":\"\",\"matched_product_names\":[],\"suggested_category\":\"\"}"
    )
    user_prompt = json.dumps(
        {
            "business": business_name(),
            "city": business_city(),
            "message": user_text,
            "catalog": catalog,
        },
        ensure_ascii=False,
    )

    if provider == "gemini":
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{urllib.parse.quote(model, safe='')}:generateContent?key={urllib.parse.quote(key)}"
        payload = {
            "contents": [
                {"role": "user", "parts": [{"text": system_prompt + "\n\n" + user_prompt}]}
            ],
            "generationConfig": {"temperature": 0.1, "maxOutputTokens": 350},
        }
        headers = {"Content-Type": "application/json"}
    else:
        url = ai_endpoint_for(provider)
        if not url:
            raise RuntimeError(f"AI endpoint is missing for provider {provider}")
        payload = {
            "model": model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "temperature": 0.1,
            "max_tokens": 350,
        }
        headers = {
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "HTTP-Referer": get_config("AI_HTTP_REFERER", "https://pricebot.local"),
            "X-Title": get_config("AI_APP_TITLE", "PriceBot"),
        }

    req = urllib.request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        method="POST",
        headers=headers,
    )
    with urllib.request.urlopen(req, timeout=15) as response:
        raw = response.read().decode("utf-8", errors="ignore")
        data = json.loads(raw)
        if provider == "gemini":
            content = ""
            for cand in data.get("candidates", []):
                for part in cand.get("content", {}).get("parts", []):
                    if part.get("text"):
                        content += part.get("text", "")
        else:
            content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
        return json_from_model_text(content)


def call_ai_json(user_text: str) -> dict:
    """Call real external AI APIs with provider/key fallback."""
    if not ai_enabled():
        return {}

    catalog = compact_catalog_for_ai(user_text)
    last_error = ""
    for provider in ai_provider_order():
        keys = ai_keys_for_provider(provider)
        if not keys:
            continue
        model = ai_model_for(provider)
        for index, key in enumerate(keys, 1):
            try:
                parsed = call_ai_once(provider, key, model, user_text, catalog)
                if parsed:
                    print(f"AI OK: provider={provider} key_index={index} model={model} parsed={parsed}", flush=True)
                    memory_record_ai_usage("text", provider)
                    return parsed
                last_error = f"{provider} key {index}: empty parse"
                print("AI PARSE EMPTY:", last_error, flush=True)
            except urllib.error.HTTPError as exc:
                body = exc.read().decode("utf-8", errors="ignore")[:700]
                last_error = f"{provider} key {index}: HTTP {exc.code} {body}"
                print("AI HTTP ERROR:", last_error, flush=True)
                # Continue to next key/provider on quota/auth/rate-limit/model errors.
                continue
            except Exception as exc:
                last_error = f"{provider} key {index}: {exc}"
                print("AI EXCEPTION:", last_error, flush=True)
                continue
    print("AI FALLBACK EXHAUSTED:", last_error, flush=True)
    return {}


def product_by_ai_names(ai_data: dict) -> Optional[dict]:
    names = []
    if ai_data.get("product_query"):
        names.append(str(ai_data.get("product_query", "")))
    for name in ai_data.get("matched_product_names") or []:
        if isinstance(name, str):
            names.append(name)
    for name in names:
        item = find_product(name)
        if item:
            return item
        # Strong exact/alias scan as backup.
        n = normalize(name)
        for product in load_products():
            keys = [product.get("name", "")] + split_aliases(product.get("aliases", ""))
            if any(normalize(k) == n for k in keys):
                return product
    return None


def suggestions_from_ai(ai_data: dict, user_text: str, limit: int = 4) -> List[dict]:
    items = []
    seen = set()
    names = []
    for name in ai_data.get("matched_product_names") or []:
        if isinstance(name, str):
            names.append(name)
    if ai_data.get("product_query"):
        names.append(str(ai_data.get("product_query")))
    if ai_data.get("suggested_category"):
        names.append(str(ai_data.get("suggested_category")))
    for name in names:
        for score, product in ranked_products(name):
            if score < 0.45:
                continue
            key = normalize(product.get("name", ""))
            if key and key not in seen:
                items.append(product)
                seen.add(key)
            if len(items) >= limit:
                return items
    for product in suggested_products(user_text, limit=limit):
        key = normalize(product.get("name", ""))
        if key and key not in seen:
            items.append(product)
            seen.add(key)
        if len(items) >= limit:
            break
    return items

def is_greeting(text: str) -> bool:
    q = normalize(text)
    greetings = ["السلام عليكم", "سلام", "مرحبا", "اهلا", "اهلين", "هاي", "hi", "hello"]
    return any(normalize(g) in q for g in greetings) and len(q) <= 45






def _safe_text_func(func_name: str, default: str) -> str:
    try:
        fn = globals().get(func_name)
        if callable(fn):
            val = str(fn()).strip()
            return val or default
    except Exception:
        pass
    return default


def pharmacy_name() -> str:
    return _safe_text_func("business_name", "صيدلية بدر البشرية")


def pharmacy_city() -> str:
    return _safe_text_func("business_city", "أجدابيا")


def pharmacy_hours() -> str:
    return _safe_text_func("business_hours", "24 ساعة")


def pharmacy_delivery() -> str:
    return _safe_text_func("delivery_text", "التوصيل غير متوفر حالياً")


def is_negative_response_query(query: str) -> bool:
    q = normalize(query or "")
    return q in {normalize(x) for x in ["لا", "لا شكرا", "الغاء", "إلغاء", "cancel", "no"]}


def is_positive_response_query(query: str) -> bool:
    q = normalize(query or "")
    return q in {normalize(x) for x in ["نعم", "اي", "تمام", "yes", "ok", "اوكي"]}


def _product_terms(item: dict) -> list:
    terms = []
    name = str(item.get("name", "")).strip()
    if name:
        terms.append(name)
    aliases = str(item.get("aliases", "") or "")
    for part in re.split(r"[,،;؛|\n]+", aliases):
        part = part.strip()
        if part:
            terms.append(part)
    clean = []
    seen = set()
    for t in terms:
        nt = normalize(t)
        if nt and nt not in seen:
            clean.append(nt)
            seen.add(nt)
    return clean


def exact_product_in_text(text: str):
    q = normalize(text or "")
    if not q:
        return None
    try:
        products = load_products()
    except Exception:
        products = []
    for item in products:
        for term in _product_terms(item):
            if len(term) >= 2 and (q == term or term in q):
                return item
    return None



def product_label(item: dict) -> str:
    parts = [str(item.get("name", "")).strip()]
    extras = []
    for field in ["form", "strength", "pack", "company"]:
        val = str(item.get(field, "")).strip()
        if val and normalize(val) not in normalize(" ".join(parts + extras)):
            extras.append(val)
    return " - ".join([p for p in [parts[0], " / ".join(extras)] if p])


def product_group_key(item: dict) -> str:
    active = normalize(item.get("active_ingredient", ""))
    strength = normalize(item.get("strength", ""))
    form = normalize(item.get("form", ""))
    if not active:
        return ""
    return "|".join([active, strength, form])


def related_same_ingredient(item: dict, limit: int = 8) -> List[dict]:
    key = product_group_key(item)
    if not key:
        return []
    out = []
    seen = set()
    for p in load_products():
        if product_group_key(p) == key:
            name_key = normalize(p.get("name", ""))
            if name_key and name_key not in seen:
                out.append(p)
                seen.add(name_key)
        if len(out) >= limit:
            break
    return out


def query_product_matches(text: str, limit: int = 8) -> List[dict]:
    q = extract_product_query(text) or text
    qn = normalize(q)
    if not qn:
        return []
    ranked = ranked_products(q)
    out = []
    seen = set()
    if ranked:
        top = ranked[0][0]
        threshold = 0.72 if top >= 0.90 else 0.62
        for score, item in ranked:
            if score < threshold:
                continue
            key = normalize(item.get("name", ""))
            if key and key not in seen:
                out.append(item)
                seen.add(key)
            if len(out) >= limit:
                break
    return out


def should_ask_options(text: str, matches: List[dict]) -> bool:
    if len(matches) < 2:
        return False
    q = normalize(extract_product_query(text) or text)
    if not q:
        return False
    # Ask if the customer's text is generic and matches several variants/companies.
    return True


def option_choice_from_text(text: str, items: List[dict]) -> Optional[dict]:
    q = normalize(text or "")
    if not q:
        return None
    m = re.search(r"\d+", q)
    if m:
        idx = int(m.group(0)) - 1
        if 0 <= idx < len(items):
            return items[idx]
    ranked = []
    for item in items:
        candidates = [product_label(item), item.get("name", ""), item.get("company", ""), item.get("brand", ""), item.get("form", ""), item.get("strength", ""), item.get("pack", ""), item.get("aliases", "")]
        best = 0.0
        for c in candidates:
            if not c:
                continue
            best = max(best, match_score(q, str(c)))
        ranked.append((best, item))
    ranked.sort(key=lambda x: x[0], reverse=True)
    if ranked and ranked[0][0] >= 0.58:
        return ranked[0][1]
    return None



def asks_medical_advice(text: str) -> bool:
    """
    Strict safety gate:
    The bot must not recommend treatment, dosage, or products for symptoms.
    Product price/availability is allowed only when the user asks about a product without dosage/treatment advice.
    """
    q = normalize(text or "")
    if not q:
        return False

    dose_or_treatment_words = [
        "جرعه", "جرعة", "جرعات", "كم حبه", "كم حبة", "كم مره", "كم مرة",
        "استعمل", "استخدم", "طريقة الاستخدام", "ينفع", "ينفعني", "عادي",
        "شن ناخذ", "شن نأخذ", "ماذا اخذ", "ماذا آخذ", "نبي علاج", "ابي علاج",
        "علاج", "دواء ل", "حاجه ل", "حاجة ل", "شن ندير", "كيف ندير",
        "حامل", "حمل", "مرضع", "رضاعه", "رضاعة", "طفل", "رضيع",
        "سكر", "ضغط", "حساسيه", "حساسية", "اعراض", "أعراض", "تشخيص",
    ]
    if any(normalize(w) in q for w in dose_or_treatment_words):
        return True

    # If the user clearly names a product and only asks price/availability, answer product info.
    if exact_product_in_text(text or ""):
        return False

    symptom_words = [
        "صداع", "راس", "رأس", "الم راس", "ألم راس", "وجع راس",
        "حراره", "حرارة", "سخونه", "سخونية",
        "كحه", "كحة", "سعال", "زكام", "رشح", "انفلونزا",
        "اسهال", "إسهال", "مغص", "معده", "معدة", "قيء", "ترجيع",
        "حموضه", "حموضة", "حرقان", "التهاب",
        "الم", "ألم", "وجع", "طفح", "حكة", "حكه", "دوخه", "دوخة",
    ]
    return any(normalize(w) in q for w in symptom_words)


def build_welcome_reply() -> str:
    return (
        f"شكراً لتواصلكم مع {pharmacy_name()} 🌿\n\n"
        f"📍 {pharmacy_city()}\n"
        f"🕒 العمل: {pharmacy_hours()}\n"
        f"🚚 {pharmacy_delivery()}\n\n"
        "يمكنك إرسال اسم الدواء أو المنتج، وسأعرض لك السعر والتوفر مباشرة.\n\n"
        "أمثلة:\n"
        "• بنادول\n"
        "• بروفين\n"
        "• أوجمنتين\n\n"
        "ملاحظة: البوت مخصص للاستعلام عن السعر والتوفر والحجز فقط، ولا يقدم وصفات أو جرعات طبية."
    )


def build_medical_safety_reply(text: str = "") -> str:
    return (
        f"{pharmacy_name()} 🌿\n\n"
        "حرصاً على السلامة، لا يقدم البوت تشخيصاً أو وصفات أو جرعات طبية.\n\n"
        "يمكنني مساعدتك في:\n"
        "• معرفة توفر دواء محدد\n"
        "• عرض السعر\n"
        "• تسجيل الحجز\n\n"
        "اكتب اسم المنتج فقط للاستعلام عنه."
    )



def build_product_reply(item: dict, original_text: str = "") -> str:
    name = str(item.get("name", "")).strip()
    available = str(item.get("available", "") or "متوفر").strip()
    price = str(item.get("price", "")).strip()
    notes = str(item.get("notes", "")).strip()
    meta = []
    for label, field in [("الشكل", "form"), ("التركيز", "strength"), ("العبوة", "pack"), ("الشركة", "company")]:
        val = str(item.get(field, "")).strip()
        if val:
            meta.append(f"{label}: {val}")
    lines = [
        f"{pharmacy_name()} 🌿",
        "",
        f"✅ المنتج: {name}",
    ]
    if meta:
        lines.extend(meta)
    lines.append(f"📦 الحالة: {available}")
    if price:
        lines.append(f"💰 السعر: {price}")
    if notes:
        lines.append(f"📝 ملاحظة: {notes}")
    image = str(item.get("image", "")).strip()
    if image:
        lines.append(f"🖼️ صورة المنتج: {image}")

    related = [p for p in related_same_ingredient(item) if normalize(p.get("name", "")) != normalize(item.get("name", ""))]
    if related:
        lines += ["", "بدائل بنفس المادة/التركيز المتوفرة في ملف الصيدلية:"]
        for p in related[:4]:
            lines.append(f"• {product_label(p)}")
        lines.append("ملاحظة: اختيار البديل المناسب يتم مع الصيدلي.")

    lines += ["", "للحجز اكتب: نعم"]
    return "\n".join(lines)


def build_options_reply(items: List[dict], reason: str = "variants") -> str:
    if reason == "same_ingredient":
        intro = "وجدت أكثر من شركة/بديل بنفس المادة أو التركيز المسجل."
    else:
        intro = "وجدت أكثر من نوع قريب من طلبك."
    lines = [
        f"{pharmacy_name()} 🌿",
        "",
        intro,
        "اختر المطلوب بكتابة الرقم أو اسم النوع:",
        "",
    ]
    for i, item in enumerate(items[:8], 1):
        price = str(item.get("price", "")).strip()
        available = str(item.get("available", "") or "متوفر").strip()
        suffix = f" — {available}"
        if price:
            suffix += f" — {price}"
        lines.append(f"{i}. {product_label(item)}{suffix}")
        image = str(item.get("image", "")).strip()
        if image:
            lines.append(f"   صورة: {image}")
    lines += ["", "ملاحظة: البوت يعرض السعر والتوفر فقط ولا يحدد الجرعات أو العلاج."]
    return "\n".join(lines)


def build_suggestion_question(item: dict, alternatives=None) -> str:
    alternatives = alternatives or []
    if alternatives:
        return build_options_reply([item] + list(alternatives), "variants")
    name = str(item.get("name", "")).strip()
    return "\n".join([
        f"{pharmacy_name()} 🌿",
        "",
        f"هل تقصد: {name}؟",
        "",
        "إذا نعم اكتب: نعم",
        "وإذا لا، أرسل اسم المنتج بشكل أوضح.",
    ])


def build_suggestions_reply(items, from_number: str = "", reason: str = "variants") -> str:
    if not items:
        return build_not_found_reply("")
    # Multiple options: wait for number/name. One option: wait for yes.
    if len(items) > 1:
        if from_number:
            PENDING_OPTIONS[from_number] = {"items": items[:8], "reason": reason}
            PENDING_SUGGESTION.pop(from_number, None)
        return build_options_reply(items[:8], reason)
    first = items[0]
    if from_number:
        PENDING_SUGGESTION[from_number] = first
    return build_suggestion_question(first, [])



def build_not_found_reply(text: str) -> str:
    if asks_medical_advice(text):
        return build_medical_safety_reply(text)
    return (
        f"{pharmacy_name()} 🌿\n\n"
        "عذراً، لم أتمكن من معرفة اسم المنتج بدقة.\n"
        "اكتب اسم الدواء أو المنتج كما هو مكتوب على العلبة.\n\n"
        "مثال:\n"
        "• بنادول\n"
        "• بروفين\n"
        "• أوجمنتين"
    )


def build_image_under_review_reply() -> str:
    return (
        f"{pharmacy_name()} 🌿\n\n"
        "تم استلام الصورة وتحويلها للصيدلي للتأكيد.\n"
        "سيتواصل معك الموظف عند مراجعتها."
    )






def is_available_catalog_request(text: str) -> bool:
    q = normalize(text or "")
    if not q:
        return False
    phrases = [
        "شنو عندكم", "شن عندكم", "شنو المتوفر", "شن المتوفر", "شن الموجود",
        "قائمه المنتجات", "قائمة المنتجات", "المنتجات المتوفره", "المنتجات المتوفرة",
        "كل المنتجات", "الادويه المتوفره", "الأدوية المتوفرة", "ادويه متوفره",
        "ارسل القائمة", "ابعث القائمة", "catalog", "list products"
    ]
    return any(normalize(p) in q for p in phrases)


def build_available_products_reply(limit: int = 12) -> str:
    products = [p for p in load_products() if normalize(p.get("available", "متوفر")) not in {"غير متوفر", "غير موجود", "نفذ", "ناقص"}]
    if not products:
        return f"{pharmacy_name()} 🌿\n\nحالياً لا توجد منتجات متوفرة مسجلة في النظام."
    lines = [
        f"{pharmacy_name()} 🌿",
        "",
        f"هذه بعض المنتجات المتوفرة حالياً ({min(limit, len(products))} من {len(products)}):",
        "",
    ]
    for i, item in enumerate(products[:limit], 1):
        price = str(item.get("price", "")).strip()
        meta = []
        if item.get("strength"):
            meta.append(str(item.get("strength")))
        if item.get("form"):
            meta.append(str(item.get("form")))
        suffix = ""
        if meta:
            suffix += " — " + " / ".join(meta)
        if price:
            suffix += f" — {price}"
        lines.append(f"{i}. {item.get('name','')}{suffix}")
    lines += [
        "",
        "للبحث أسرع: اكتب اسم الدواء أو المنتج المطلوب فقط."
    ]
    return "\n".join(lines)


def build_order_success_reply(item: dict) -> str:
    name = str(item.get("name", "")).strip()
    price = str(item.get("price", "")).strip()
    lines = [
        "✅ تم استلام طلبك مبدئياً",
        "",
        f"المنتج: {name}",
    ]
    if price:
        lines.append(f"السعر: {price}")
    lines += [
        "",
        "الطلب بانتظار تأكيد الصيدلية. ستصلك رسالة تأكيد أو اعتذار على واتساب.",
        f"شكراً لتواصلكم مع {pharmacy_name()}.",
    ]
    return "\n".join(lines)


def customer_order_confirmed_message(row: dict) -> str:
    lines = [
        f"{pharmacy_name()} 🌿",
        "",
        "✅ تم تأكيد طلبك من الصيدلية.",
        f"المنتج: {row.get('product','')}",
    ]
    if row.get("price"):
        lines.append(f"السعر: {row.get('price')}")
    lines += ["", "يمكنك الحضور أو انتظار تواصل الموظف حسب طريقة الاتفاق."]
    return "\n".join(lines)


def customer_order_rejected_message(row: dict) -> str:
    return (
        f"{pharmacy_name()} 🌿\n\n"
        "نعتذر، لم يتم تأكيد طلبك حالياً بعد مراجعة الصيدلية.\n"
        f"المنتج: {row.get('product','')}\n\n"
        "يمكنك إرسال اسم منتج آخر للاستعلام عن التوفر والسعر."
    )


def build_daily_report_message() -> str:
    today = datetime.now().strftime("%Y-%m-%d")
    orders = read_orders()
    todays = [o for o in orders if str(o.get("time", "")).startswith(today)]
    pending = [o for o in todays if o.get("status", "") in {"new", "pending", "review"}]
    confirmed = [o for o in todays if o.get("status") == "confirmed"]
    rejected = [o for o in todays if o.get("status") == "rejected"]
    done = [o for o in todays if o.get("status") == "done"]
    lines = [
        f"📊 تقرير يومي - {pharmacy_name()}",
        f"التاريخ: {today}",
        "",
        f"إجمالي طلبات اليوم: {len(todays)}",
        f"بانتظار المراجعة: {len(pending)}",
        f"مؤكدة: {len(confirmed)}",
        f"مرفوضة: {len(rejected)}",
        f"منفذة: {len(done)}",
    ]
    if pending[:5]:
        lines += ["", "أحدث الطلبات المعلقة:"]
        for o in pending[-5:]:
            lines.append(f"• {o.get('product','')} — {o.get('phone','')}")
    lines += ["", "افتح لوحة الطلبات لمتابعة التفاصيل."]
    return "\n".join(lines)


def build_reply(text: str, from_number: str = "") -> str:
    raw_text = text or ""
    query = normalize(raw_text)
    yes_words = [normalize(x) for x in ["نعم", "اي", "تمام", "yes", "ok", "اوكي"]]
    reserve_words = [normalize(x) for x in ["حجز", "احجز", "نبي حجز", "اريد حجز", "أريد حجز"]]
    no_words = [normalize(x) for x in ["لا", "الغاء", "إلغاء", "cancel", "no"]]

    # Customer is choosing from a numbered/variant list.
    if from_number and from_number in PENDING_OPTIONS:
        pending = PENDING_OPTIONS[from_number]
        items = pending.get("items", [])
        if is_negative_response_query(query):
            PENDING_OPTIONS.pop(from_number, None)
            return f"{pharmacy_name()} 🌿\n\nتمام. أرسل اسم المنتج بشكل أوضح."
        chosen = option_choice_from_text(raw_text, items)
        if chosen:
            PENDING_OPTIONS.pop(from_number, None)
            LAST_PRODUCT[from_number] = chosen
            return build_product_reply(chosen, raw_text)
        if is_positive_response_query(query) and len(items) > 1:
            return f"{pharmacy_name()} 🌿\n\nاختر النوع المطلوب بكتابة الرقم من القائمة، مثال: 1"

    # Customer confirmed a single suggestion.
    if from_number and from_number in PENDING_SUGGESTION and is_positive_response_query(query):
        item = PENDING_SUGGESTION.pop(from_number)
        LAST_PRODUCT[from_number] = item
        return build_product_reply(item, raw_text)

    if from_number and from_number in PENDING_SUGGESTION and is_negative_response_query(query):
        PENDING_SUGGESTION.pop(from_number, None)
        return f"{pharmacy_name()} 🌿\n\nتمام. أرسل اسم المنتج بشكل أوضح."

    # Confirm reservation after product details were shown.
    if from_number and from_number in LAST_PRODUCT and is_negative_response_query(query):
        LAST_PRODUCT.pop(from_number, None)
        return f"{pharmacy_name()} 🌿\n\nتم إلغاء الحجز المؤقت. يمكنك كتابة اسم منتج آخر للاستعلام."

    if from_number and from_number in LAST_PRODUCT and (
        is_positive_response_query(query) or any(w in query for w in reserve_words)
    ):
        item = LAST_PRODUCT[from_number]
        save_order(from_number, item, raw_text)
        LAST_PRODUCT.pop(from_number, None)
        PENDING_SUGGESTION.pop(from_number, None)
        PENDING_OPTIONS.pop(from_number, None)
        return build_order_success_reply(item)

    if is_greeting(raw_text):
        return build_welcome_reply()

    if is_available_catalog_request(raw_text):
        return build_available_products_reply()

    # Persistent memory/corrections: if this wording was learned before, avoid AI.
    corrected_item = memory_get_admin_correction(raw_text)
    if corrected_item:
        if from_number:
            LAST_PRODUCT[from_number] = corrected_item
            PENDING_SUGGESTION.pop(from_number, None)
            PENDING_OPTIONS.pop(from_number, None)
        return build_product_reply(corrected_item, raw_text)

    # If customer asks about symptoms/dose/treatment without a named product: block.
    if asks_medical_advice(raw_text):
        return build_medical_safety_reply(raw_text)

    # Search product variants/companies first.
    matches = query_product_matches(raw_text, limit=8)
    if matches:
        # If multiple products match a generic request, ask the customer to choose.
        if should_ask_options(raw_text, matches):
            return build_suggestions_reply(matches, from_number, "variants")

        selected = matches[0]
        related = related_same_ingredient(selected)
        if len(related) > 1:
            return build_suggestions_reply(related, from_number, "same_ingredient")

        if from_number:
            LAST_PRODUCT[from_number] = selected
            PENDING_SUGGESTION.pop(from_number, None)
            PENDING_OPTIONS.pop(from_number, None)
        return build_product_reply(selected, raw_text)

    cached = memory_get_query_response(raw_text)
    if cached:
        if cached.get("type") == "product":
            cached_item = find_product_by_name_or_alias(cached.get("product_name", ""))
            if cached_item:
                if from_number:
                    LAST_PRODUCT[from_number] = cached_item
                    PENDING_SUGGESTION.pop(from_number, None)
                    PENDING_OPTIONS.pop(from_number, None)
                return build_product_reply(cached_item, raw_text)
        if cached.get("type") == "options":
            cached_items = memory_items_from_names(cached.get("product_names", []))
            if cached_items:
                return build_suggestions_reply(cached_items, from_number, "variants")

    ai_data = call_ai_json(raw_text) if ai_enabled() else {}
    intent = str(ai_data.get("intent", "")).strip().lower() if ai_data else ""
    if intent == "greeting":
        return build_welcome_reply()
    if intent == "medical_advice":
        return build_medical_safety_reply(raw_text)

    ai_suggestions = suggestions_from_ai(ai_data, raw_text) if ai_data else []
    if ai_suggestions:
        if len(ai_suggestions) == 1:
            memory_remember_product_query(raw_text, ai_suggestions[0], "ai")
        else:
            memory_remember_options_query(raw_text, ai_suggestions, "ai")
        return build_suggestions_reply(ai_suggestions, from_number, "variants")

    suggestions = suggested_products(raw_text)
    if suggestions:
        memory_remember_options_query(raw_text, suggestions, "local_suggestion")
        return build_suggestions_reply(suggestions, from_number, "variants")
    return build_not_found_reply(raw_text)

def send_whatsapp_message(to_number: str, message: str) -> bool:
    token = get_access_token()
    phone_number_id = get_phone_number_id()
    if not token or not phone_number_id:
        print("SEND ERROR: Missing WhatsApp token or PHONE_NUMBER_ID", flush=True)
        return False

    url = f"https://graph.facebook.com/{GRAPH_API_VERSION}/{phone_number_id}/messages"
    payload = {
        "messaging_product": "whatsapp",
        "to": to_number,
        "type": "text",
        "text": {"body": message},
    }
    req = urllib.request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        method="POST",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
    )

    try:
        with urllib.request.urlopen(req, timeout=20) as response:
            print("SEND OK:", response.read().decode("utf-8"), flush=True)
            return True
    except urllib.error.HTTPError as exc:
        print("SEND ERROR:", exc.code, exc.read().decode("utf-8", errors="ignore"), flush=True)
    except Exception as exc:
        print("SEND EXCEPTION:", str(exc), flush=True)
    return False




def notify_admin(message: str, customer_phone: str = "") -> bool:
    admin_phone = admin_notify_phone()
    customer_phone = normalize_phone(customer_phone)
    if not admin_phone:
        print("ADMIN NOTIFY SKIP: ADMIN_NOTIFY_PHONE not set", flush=True)
        return False
    if customer_phone and admin_phone == customer_phone:
        print("ADMIN NOTIFY SKIP: admin phone equals customer phone", flush=True)
        return False
    return send_whatsapp_message(admin_phone, message)


def build_admin_order_message(phone: str, item: dict, message: str) -> str:
    return (
        f"🔔 طلب حجز بانتظار التأكيد - {pharmacy_name()}\n\n"
        f"رقم الزبون: {phone}\n"
        f"المنتج: {item.get('name', '')}\n"
        f"السعر: {item.get('price', '')}\n"
        f"التوفر: {item.get('available', '')}\n"
        f"ملاحظة: {item.get('notes', '')}\n"
        f"رسالة الزبون: {message or '-'}\n"
        f"الوقت: {now_str()}\n\n"
        "افتح لوحة الطلبات لتأكيده أو رفضه."
    )


def save_review_order(phone: str, title: str, message: str, media_url: str = "") -> None:
    item = {
        "name": title,
        "price": "",
        "available": "بانتظار مراجعة الصيدلي",
        "notes": "صورة/روشتة تحتاج تأكيد" if media_url else "تحتاج مراجعة",
    }
    save_order(phone, item, (message or "") + (f"\nرابط الصورة: {media_url}" if media_url else ""))


def get_whatsapp_media_info(media_id: str) -> dict:
    token = get_access_token()
    if not token or not media_id:
        return {}
    url = f"https://graph.facebook.com/{GRAPH_API_VERSION}/{urllib.parse.quote(media_id)}"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    try:
        with urllib.request.urlopen(req, timeout=20) as response:
            return json.loads(response.read().decode("utf-8", errors="ignore"))
    except Exception as exc:
        print("MEDIA INFO ERROR:", str(exc), flush=True)
        return {}


def download_whatsapp_media(media_id: str) -> Tuple[bytes, str, str]:
    info = get_whatsapp_media_info(media_id)
    media_url = info.get("url", "")
    mime_type = info.get("mime_type", "image/jpeg") or "image/jpeg"
    if not media_url:
        return b"", mime_type, ""
    token = get_access_token()
    req = urllib.request.Request(media_url, headers={"Authorization": f"Bearer {token}"})
    try:
        with urllib.request.urlopen(req, timeout=30) as response:
            return response.read(), mime_type, info.get("sha256", "")
    except Exception as exc:
        print("MEDIA DOWNLOAD ERROR:", str(exc), flush=True)
        return b"", mime_type, info.get("sha256", "")


def save_incoming_media(media_bytes: bytes, mime_type: str, media_id: str) -> str:
    MEDIA_DIR.mkdir(parents=True, exist_ok=True)
    ext = ".jpg"
    if "png" in mime_type:
        ext = ".png"
    elif "webp" in mime_type:
        ext = ".webp"
    elif "pdf" in mime_type:
        ext = ".pdf"
    safe_id = re.sub(r"[^A-Za-z0-9_-]+", "", media_id or datetime.now().strftime("%Y%m%d%H%M%S"))[:80]
    filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_id}{ext}"
    path = MEDIA_DIR / filename
    path.write_bytes(media_bytes)
    return filename


def admin_media_url(filename: str) -> str:
    return f"{public_base_url()}/admin/media/{urllib.parse.quote(filename)}?key={urllib.parse.quote(ADMIN_KEY)}"


def image_ai_enabled() -> bool:
    return bool(ai_keys_for_provider("gemini"))


def call_gemini_vision_json(media_bytes: bytes, mime_type: str) -> dict:
    keys = ai_keys_for_provider("gemini")
    if not keys or not media_bytes:
        return {}
    model = ai_model_for("gemini") or "gemini-2.5-flash-lite"
    catalog = compact_catalog_for_ai("صورة منتج", limit=120)
    prompt = (
        "أنت مساعد صيدلية. حلل الصورة وأرجع JSON فقط. "
        "لا تقدم نصائح طبية. إذا كانت الصورة روشتة/وصفة طبية/ورقة بخط يد/أدوية كثيرة غير واضحة، اجعل image_type=prescription_or_unclear وrequires_admin_review=true. "
        "إذا كانت صورة علبة منتج دوائي أو منتج صيدلية واضح، اجعل image_type=product_packaging واستخرج أسماء المنتجات الظاهرة. "
        "استخدم الكتالوج فقط للمطابقة ولا تخترع منتجات. "
        "الشكل المطلوب: {\"image_type\":\"product_packaging|prescription_or_unclear|unknown\",\"product_names\":[],\"matched_product_names\":[],\"requires_admin_review\":true|false,\"confidence\":0.0} "
        f"\nCatalog: {json.dumps(catalog, ensure_ascii=False)}"
    )
    b64 = base64.b64encode(media_bytes).decode("ascii")
    for index, key in enumerate(keys, 1):
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{urllib.parse.quote(model, safe='')}:generateContent?key={urllib.parse.quote(key)}"
        payload = {
            "contents": [{
                "role": "user",
                "parts": [
                    {"text": prompt},
                    {"inline_data": {"mime_type": mime_type or "image/jpeg", "data": b64}},
                ],
            }],
            "generationConfig": {"temperature": 0.05, "maxOutputTokens": 500},
        }
        req = urllib.request.Request(url, data=json.dumps(payload).encode("utf-8"), method="POST", headers={"Content-Type": "application/json"})
        try:
            with urllib.request.urlopen(req, timeout=30) as response:
                raw = json.loads(response.read().decode("utf-8", errors="ignore"))
                content = ""
                for cand in raw.get("candidates", []):
                    for part in cand.get("content", {}).get("parts", []):
                        content += part.get("text", "") or ""
                data = json_from_model_text(content)
                if data:
                    print(f"VISION AI OK: gemini key_index={index} parsed={data}", flush=True)
                    memory_record_ai_usage("image", "gemini")
                    return data
        except urllib.error.HTTPError as exc:
            print("VISION AI HTTP ERROR:", exc.code, exc.read().decode("utf-8", errors="ignore")[:600], flush=True)
        except Exception as exc:
            print("VISION AI EXCEPTION:", str(exc), flush=True)
    return {}


def product_from_vision(data: dict) -> Optional[dict]:
    names: List[str] = []
    for key in ["matched_product_names", "product_names"]:
        for name in data.get(key) or []:
            if isinstance(name, str) and name.strip():
                names.append(name.strip())
    for name in names:
        item = find_product(name)
        if item:
            return item
        n = normalize(name)
        for product in load_products():
            keys = [product.get("name", "")] + split_aliases(product.get("aliases", ""))
            if any(normalize(k) == n for k in keys):
                return product
    return None


def process_image_message(msg: dict, from_number: str) -> str:
    image = msg.get("image", {}) or {}
    media_id = image.get("id", "")
    caption = image.get("caption", "") or ""
    media_bytes, mime_type, media_sha = download_whatsapp_media(media_id)
    if not media_sha and media_bytes:
        media_sha = hashlib.sha256(media_bytes).hexdigest()
    filename = save_incoming_media(media_bytes, mime_type, media_id) if media_bytes else ""
    link = admin_media_url(filename) if filename else ""

    if not media_bytes:
        save_review_order(from_number, "صورة تحتاج مراجعة", "تعذر تحميل الصورة من واتساب", link)
        notify_admin(
            f"📷 صورة تحتاج مراجعة - {pharmacy_name()}\n\nرقم الزبون: {from_number}\nسبب التحويل: تعذر تحميل الصورة من واتساب\nالوقت: {now_str()}",
            from_number,
        )
        return build_image_under_review_reply()

    cached_image = memory_lookup_image(media_sha)
    if cached_image:
        if cached_image.get("type") == "product":
            cached_item = find_product_by_name_or_alias(cached_image.get("product_name", ""))
            if cached_item:
                LAST_PRODUCT[from_number] = cached_item
                return build_product_reply(cached_item, caption)
        if cached_image.get("type") == "review":
            save_review_order(from_number, "روشتة/صورة تحتاج مراجعة", f"نتيجة محفوظة من الذاكرة: {cached_image.get('reason', 'review')}", link)
            notify_admin(
                f"📷 صورة تحتاج مراجعة - {pharmacy_name()}\n\nرقم الزبون: {from_number}\nالسبب: نتيجة محفوظة من الذاكرة\nرابط الصورة: {link or '-'}\nالوقت: {now_str()}",
                from_number,
            )
            return build_image_under_review_reply()

    vision = call_gemini_vision_json(media_bytes, mime_type) if image_ai_enabled() else {}
    image_type = str(vision.get("image_type", "")).lower()
    requires_review = bool(vision.get("requires_admin_review"))

    if image_type == "product_packaging" and not requires_review:
        item = product_from_vision(vision)
        if item:
            memory_remember_image_product(media_sha, item, "gemini_vision")
            LAST_PRODUCT[from_number] = item
            return build_product_reply(item, caption)
        # If vision saw names but no catalog match, ask admin to confirm instead of guessing.
        product_names = ", ".join([str(x) for x in (vision.get("product_names") or vision.get("matched_product_names") or []) if x])
        memory_remember_image_review(media_sha, "product_not_in_catalog")
        save_review_order(from_number, "صورة منتج غير موجودة بالكتالوج", f"الأسماء المقروءة: {product_names or '-'}", link)
        notify_admin(
            f"📷 صورة منتج تحتاج إضافة/تأكيد - {pharmacy_name()}\n\nرقم الزبون: {from_number}\nالأسماء المقروءة: {product_names or '-'}\nرابط الصورة: {link or '-'}\nالوقت: {now_str()}",
            from_number,
        )
        return (
            f"{pharmacy_name()} 🌿\n\n"
            "وصلت صورة المنتج، لكن لم أجد مطابقة مؤكدة في قائمة المنتجات.\n"
            "تم تحويلها للصيدلي للتأكيد."
        )

    reason = "روشتة أو صورة غير واضحة" if image_type in {"prescription_or_unclear", "unknown", ""} else image_type
    memory_remember_image_review(media_sha, reason)
    save_review_order(from_number, "روشتة/صورة تحتاج مراجعة", f"نوع الصورة: {reason}", link)
    notify_admin(
        f"📷 روشتة/صورة تحتاج مراجعة - {pharmacy_name()}\n\nرقم الزبون: {from_number}\nنوع الصورة: {reason}\nرابط الصورة: {link or '-'}\nالوقت: {now_str()}\n\nيرجى مراجعتها من لوحة الطلبات أو التواصل مع الزبون.",
        from_number,
    )
    return build_image_under_review_reply()


def save_order(phone: str, item: dict, message: str, status: str = "pending") -> None:
    ensure_sqlite_db()
    row = {
        "time": now_str(),
        "phone": phone,
        "product": item.get("name", ""),
        "price": item.get("price", ""),
        "available": item.get("available", ""),
        "notes": item.get("notes", ""),
        "message": message,
        "status": status,
    }
    with db_connect() as conn:
        conn.execute(
            f"INSERT INTO orders ({','.join(ORDER_FIELDS)}) VALUES ({','.join(['?'] * len(ORDER_FIELDS))})",
            tuple(row.get(field, "") for field in ORDER_FIELDS),
        )
        conn.commit()
    # Keep CSV mirror for easy download/backups.
    mirror_orders_to_csv(read_orders())
    notify_admin(build_admin_order_message(phone, item, message), phone)


def read_orders() -> List[dict]:
    ensure_sqlite_db()
    migrate_legacy_files_to_db()
    with db_connect() as conn:
        rows = conn.execute(f"SELECT {','.join(ORDER_FIELDS)} FROM orders ORDER BY id ASC").fetchall()
    return [{field: str(row[field] or "") for field in ORDER_FIELDS} for row in rows]


def write_orders(rows: List[dict]) -> None:
    ensure_sqlite_db()
    with db_connect() as conn:
        conn.execute("DELETE FROM orders")
        insert_order_rows(conn, rows)
        conn.commit()
    mirror_orders_to_csv(rows)


def check_admin(key: str) -> bool:
    return bool(key) and key == ADMIN_KEY


def safe_redirect(url: str) -> RedirectResponse:
    return RedirectResponse(url, status_code=303)


def e(value: object) -> str:
    return html.escape(str(value or ""), quote=True)


BASE_CSS = """
:root{
  --bg:#f6f8fb; --card:#ffffff; --text:#172026; --muted:#667085; --line:#e6e9ef;
  --brand:#0f766e; --brand2:#115e59; --soft:#ecfdf5; --danger:#b42318; --ok:#15803d;
  --warn:#b45309; --shadow:0 10px 26px rgba(16,24,40,.08); --radius:18px;
}
*{box-sizing:border-box} body{margin:0;padding:18px;background:linear-gradient(180deg,#eef7f5 0,#f7f8fb 220px);color:var(--text);font-family:Arial,Tahoma,sans-serif;direction:rtl}
a{color:var(--brand);text-decoration:none} h1{font-size:28px;margin:8px 0 12px} h2{font-size:22px;margin:4px 0 14px}.container{max-width:1180px;margin:0 auto}.hero{background:linear-gradient(135deg,#0f766e,#134e4a);color:#fff;border-radius:24px;padding:22px;box-shadow:var(--shadow);margin:6px 0 14px}.hero h1{margin:0 0 8px}.hero p{margin:0;color:#d8fffb}.box,.card{background:var(--card);border:1px solid var(--line);border-radius:var(--radius);padding:18px;margin:14px 0;box-shadow:var(--shadow)}.nav{display:grid;grid-template-columns:repeat(6,minmax(0,1fr));gap:10px;margin:14px 0}.btn,a.btn,button,input[type=submit]{display:inline-flex;align-items:center;justify-content:center;gap:6px;border:0;border-radius:12px;padding:12px 15px;font-size:16px;font-weight:700;background:var(--brand);color:#fff;cursor:pointer;text-align:center}.btn.secondary,a.btn.secondary{background:#fff;color:#0f5f59;border:1px solid #b9ddd8}.btn.danger,a.btn.danger{background:#fff1f0;color:var(--danger);border:1px solid #ffcbc5}.btn.ok,a.btn.ok{background:#ecfdf3;color:var(--ok);border:1px solid #bbf7d0}.btn.warn{background:#fffbeb;color:#92400e;border:1px solid #fde68a}.msg{color:var(--ok);font-weight:800;text-align:center}.notice{color:var(--muted);font-size:14px;line-height:1.9}.form-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}.field{display:flex;flex-direction:column;gap:7px}label{font-weight:800}input,textarea,select{width:100%;border:1px solid #cfd6dd;border-radius:12px;padding:12px;font-size:16px;background:#fff}textarea{min-height:180px;line-height:1.7}.stats{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px}.stat{background:#fff;border:1px solid var(--line);border-radius:18px;padding:16px;box-shadow:var(--shadow)}.stat b{font-size:28px;color:var(--brand)}.stat span{display:block;color:var(--muted);margin-top:4px}.product-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(285px,1fr));gap:14px}.product-title{font-size:20px;font-weight:900;margin-bottom:8px}.product-meta{color:var(--muted);font-size:14px;margin:4px 0 12px}.actions{display:flex;flex-wrap:wrap;gap:9px;align-items:center;margin-top:10px}.card-form{display:grid;gap:10px}.order-card{position:relative;display:grid;gap:8px}.badge{display:inline-flex;align-items:center;border-radius:999px;padding:5px 10px;font-weight:800;font-size:13px}.badge.new{background:#fff7ed;color:#c2410c}.badge.done{background:#ecfdf3;color:#15803d}.order-head{display:flex;justify-content:space-between;gap:10px;align-items:center;border-bottom:1px solid var(--line);padding-bottom:10px}.order-title{font-weight:900;font-size:19px}.order-row{display:flex;justify-content:space-between;gap:12px;border-bottom:1px dashed var(--line);padding:6px 0}.order-row strong{white-space:nowrap;color:#344054}.status-new{color:var(--warn);font-weight:900}.status-done{color:var(--ok);font-weight:900}.table-wrap{overflow-x:auto}table{width:100%;border-collapse:collapse;background:#fff}th,td{border:1px solid var(--line);padding:10px;text-align:center}th{background:#f0f2f4}.admin-hint{background:#f8fafc;border:1px dashed #cbd5e1;border-radius:14px;padding:12px;color:#475569;line-height:1.8}pre{background:#0b1220;color:#e5e7eb;border-radius:14px;padding:14px;overflow:auto}
@media(max-width:900px){.nav{grid-template-columns:repeat(2,minmax(0,1fr))}.stats{grid-template-columns:repeat(2,minmax(0,1fr))}.form-grid{grid-template-columns:1fr}}
@media(max-width:600px){body{padding:10px}h1{font-size:24px}.hero{padding:17px;border-radius:18px}.box,.card{padding:14px;border-radius:15px}.nav{grid-template-columns:1fr 1fr}.product-grid{grid-template-columns:1fr}.actions .btn,.actions button{flex:1 1 auto}.stats{grid-template-columns:1fr 1fr}.stat b{font-size:22px}}
"""


def page_layout(title: str, body: str) -> str:
    return f"""<!doctype html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{e(title)}</title>
  <style>{BASE_CSS}</style>
</head>
<body>
  <div class="container">
    {body}
  </div>
</body>
</html>"""



def admin_nav(key: str) -> str:
    key_q = urllib.parse.quote(key)
    return f"""
    <div class="nav">
      <a class="btn secondary" href="/admin?key={key_q}">📦 المنتجات</a>
      <a class="btn secondary" href="/admin/orders?key={key_q}">🧾 الطلبات</a>
      <a class="btn secondary" href="/admin/upload?key={key_q}">⬆️ رفع ملف</a>
      <a class="btn secondary" href="/admin/bulk?key={key_q}">➕ إدخال بالجملة</a>
      <a class="btn secondary" href="/admin/settings?key={key_q}">⚙️ الإعدادات</a>
      <a class="btn secondary" href="/admin/memory?key={key_q}">🧠 الذاكرة</a>
      <a class="btn secondary" href="/admin/database?key={key_q}">🗄️ قاعدة البيانات</a>
      <a class="btn secondary" href="/health">🟢 Health</a>
    </div>
    """

@app.get("/")
def home() -> dict:
    return {"status": "PriceBot WhatsApp bot is running", "version": "3.1.0", "business": business_name()}


@app.get("/health")
def health() -> dict:
    return {
        "ok": True,
        "version": "3.1.0",
        "business": business_name(),
        "city": business_city(),
        "hours": business_hours(),
        "delivery": delivery_text(),
        "products_count": len(load_products()),
        "orders_count": len(read_orders()),
        "phone_number_id_set": bool(get_phone_number_id()),
        "access_token_set": bool(get_access_token()),
        "ai_enabled": ai_enabled(),
        "ai_provider_order": ai_provider_order(),
        "ai_status": ai_status_text(),
        "ai_model": ai_model() or ai_default_model(),
        "database": db_stats(),
        "admin_notify_phone_set": bool(admin_notify_phone()),
    }


@app.get("/products")
def products_api() -> dict:
    return {"products": load_products()}



def update_env_values(values: Dict[str, str]) -> None:
    """Update .env atomically while preserving unrelated secrets/settings."""
    ENV_FILE.parent.mkdir(parents=True, exist_ok=True)
    lines = ENV_FILE.read_text(encoding="utf-8").splitlines() if ENV_FILE.exists() else []
    seen = set()
    out = []
    for line in lines:
        if line.strip() and not line.strip().startswith("#") and "=" in line:
            key, _ = line.split("=", 1)
            key = key.strip()
            if key in values:
                out.append(f"{key}={values[key]}")
                seen.add(key)
            else:
                out.append(line)
        else:
            out.append(line)
    for key, value in values.items():
        if key not in seen:
            out.append(f"{key}={value}")
    tmp = ENV_FILE.with_suffix(".env.tmp")
    tmp.write_text("\n".join(out).rstrip() + "\n", encoding="utf-8")
    tmp.replace(ENV_FILE)



@app.get("/admin/settings")
def admin_settings(key: str = "", msg: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    enabled_checked = "checked" if ai_enabled() else ""
    order_value = ",".join(ai_provider_order())
    admin_phone_value = admin_notify_phone()
    body = f"""
    <div class="hero"><h1>إعدادات {e(business_name())}</h1><p>بيانات الصيدلية، رقم تنبيهات الأدمن، والذكاء الاصطناعي.</p></div>
    {admin_nav(key)}
    <div class="box">
      <p class="msg">{e(msg)}</p>
      <form method="post" action="/admin/settings/save?key={urllib.parse.quote(key)}">
        <h2>بيانات الصيدلية</h2>
        <div class="form-grid">
          <div class="field"><label>اسم الصيدلية</label><input name="PHARMACY_NAME" value="{e(business_name())}"></div>
          <div class="field"><label>المدينة</label><input name="PHARMACY_CITY" value="{e(business_city())}"></div>
          <div class="field"><label>ساعات العمل</label><input name="PHARMACY_HOURS" value="{e(business_hours())}"></div>
          <div class="field"><label>نص التوصيل</label><input name="DELIVERY_TEXT" value="{e(delivery_text())}"></div>
          <div class="field"><label>رقم أدمن واتساب للتنبيهات</label><input name="ADMIN_NOTIFY_PHONE" value="{e(admin_phone_value)}" placeholder="2189XXXXXXXX"></div>
          <div class="field"><label>رابط الموقع العام</label><input name="PUBLIC_BASE_URL" value="{e(public_base_url())}" placeholder="https://46.101.148.246.sslip.io"></div>
        </div>
        <p><label><input type="checkbox" name="DELIVERY_AVAILABLE" value="yes" {'checked' if delivery_enabled() else ''}> التوصيل متوفر</label></p>
        <div class="admin-hint">رقم الأدمن يستقبل تنبيه واتساب عند تأكيد طلب أو وصول صورة/روشتة تحتاج مراجعة. لا تضع رقم الزبون هنا إلا إذا كنت تختبر فقط.</div>

        <h2>الذكاء الاصطناعي المجاني/الاحتياطي</h2>
        <p class="notice">الحالة الحالية: <b>{e(ai_status_text())}</b></p>
        <p><label><input type="checkbox" name="AI_ENABLED" value="yes" {enabled_checked}> تفعيل AI API</label></p>
        <div class="form-grid">
          <div class="field"><label>ترتيب المحاولة</label><input name="AI_PROVIDER_ORDER" value="{e(order_value)}" placeholder="gemini,openrouter,groq"></div>
          <div class="field"><label>OpenRouter model</label><input name="AI_OPENROUTER_MODEL" value="{e(ai_model_for('openrouter'))}" placeholder="openrouter/free"></div>
          <div class="field"><label>Gemini model</label><input name="AI_GEMINI_MODEL" value="{e(ai_model_for('gemini'))}" placeholder="gemini-2.5-flash-lite"></div>
          <div class="field"><label>Groq model</label><input name="AI_GROQ_MODEL" value="{e(ai_model_for('groq'))}" placeholder="llama-3.1-8b-instant"></div>
        </div>
        <p class="notice">الصق المفاتيح الجديدة فقط إذا تريد تغييرها. اترك الخانة فارغة للحفاظ على المفاتيح المحفوظة. يمكن وضع أكثر من مفتاح، كل مفتاح في سطر.</p>
        <div class="form-grid">
          <div class="field"><label>Gemini keys جديدة — المحفوظ: {e(masked_count_for('gemini'))}</label><textarea name="AI_GEMINI_KEYS_NEW" placeholder="AIza..."></textarea></div>
          <div class="field"><label>OpenRouter keys جديدة — المحفوظ: {e(masked_count_for('openrouter'))}</label><textarea name="AI_OPENROUTER_KEYS_NEW" placeholder="sk-or-v1-..."></textarea></div>
          <div class="field"><label>Groq keys جديدة — المحفوظ: {e(masked_count_for('groq'))}</label><textarea name="AI_GROQ_KEYS_NEW" placeholder="gsk_..."></textarea></div>
          <div class="field"><label>OpenAI keys جديدة — المحفوظ: {e(masked_count_for('openai'))}</label><textarea name="AI_OPENAI_KEYS_NEW" placeholder="اختياري - اتركه فارغاً"></textarea></div>
        </div>
        <div class="actions"><button type="submit">حفظ الإعدادات</button><a class="btn secondary" href="/admin/settings/test?key={urllib.parse.quote(key)}&q=كم سعر البروفين؟">اختبار AI</a></div>
      </form>
    </div>
    """
    return HTMLResponse(page_layout("إعدادات PriceBot", body))

@app.post("/admin/settings/save")
async def admin_settings_save(request: Request, key: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    form = await request.form()
    values = {
        "PHARMACY_NAME": str(form.get("PHARMACY_NAME", business_name())).strip() or business_name(),
        "PHARMACY_CITY": str(form.get("PHARMACY_CITY", business_city())).strip() or business_city(),
        "PHARMACY_HOURS": str(form.get("PHARMACY_HOURS", business_hours())).strip() or business_hours(),
        "DELIVERY_AVAILABLE": "yes" if form.get("DELIVERY_AVAILABLE") == "yes" else "no",
        "DELIVERY_TEXT": str(form.get("DELIVERY_TEXT", delivery_text())).strip() or delivery_text(),
        "ADMIN_NOTIFY_PHONE": normalize_phone(str(form.get("ADMIN_NOTIFY_PHONE", admin_notify_phone())).strip()),
        "PUBLIC_BASE_URL": str(form.get("PUBLIC_BASE_URL", public_base_url())).strip().rstrip("/") or public_base_url(),
        "AI_ENABLED": "yes" if form.get("AI_ENABLED") == "yes" else "no",
        "AI_PROVIDER_ORDER": str(form.get("AI_PROVIDER_ORDER", ",".join(ai_provider_order()))).strip() or "gemini,openrouter,groq",
        "AI_OPENROUTER_MODEL": str(form.get("AI_OPENROUTER_MODEL", ai_model_for("openrouter"))).strip() or ai_default_model_for("openrouter"),
        "AI_GEMINI_MODEL": str(form.get("AI_GEMINI_MODEL", ai_model_for("gemini"))).strip() or ai_default_model_for("gemini"),
        "AI_GROQ_MODEL": str(form.get("AI_GROQ_MODEL", ai_model_for("groq"))).strip() or ai_default_model_for("groq"),
        "AI_OPENAI_MODEL": str(form.get("AI_OPENAI_MODEL", ai_model_for("openai"))).strip() or ai_default_model_for("openai"),
        "AI_CUSTOM_BASE_URL": str(form.get("AI_CUSTOM_BASE_URL", get_config("AI_CUSTOM_BASE_URL", ""))).strip(),
    }
    key_fields = {
        "AI_OPENROUTER_KEYS_NEW": "AI_OPENROUTER_KEYS",
        "AI_GEMINI_KEYS_NEW": "AI_GEMINI_KEYS",
        "AI_GROQ_KEYS_NEW": "AI_GROQ_KEYS",
        "AI_OPENAI_KEYS_NEW": "AI_OPENAI_KEYS",
    }
    for form_name, env_name in key_fields.items():
        raw_keys = str(form.get(form_name, "")).strip()
        if raw_keys:
            values[env_name] = join_secret_list(raw_keys)
    update_env_values(values)
    # Refresh current process environment for immediate use without restart.
    for k, v in values.items():
        os.environ[k] = v
    return safe_redirect(f"/admin/settings?key={urllib.parse.quote(key)}&msg={urllib.parse.quote('تم حفظ الإعدادات')}")


@app.get("/admin/settings/test")
def admin_settings_test(key: str = "", q: str = "كم سعر البروفين؟"):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    local_reply = build_reply(q, "")
    ai_data = call_ai_json(q) if ai_enabled() else {"note": "AI غير مفعل أو لا يوجد مفتاح"}
    body = f"""
    <h1>اختبار الذكاء الاصطناعي</h1>
    {admin_nav(key)}
    <div class="box">
      <form method="get" action="/admin/settings/test">
        <input type="hidden" name="key" value="{e(key)}">
        <div class="field"><label>رسالة اختبار</label><input name="q" value="{e(q)}"></div>
        <div class="actions"><button type="submit">اختبار</button></div>
      </form>
    </div>
    <div class="box"><h2>رد البوت</h2><pre style="white-space:pre-wrap; direction:rtl; font-size:16px">{e(local_reply)}</pre></div>
    <div class="box"><h2>نتيجة AI JSON</h2><pre style="white-space:pre-wrap; direction:ltr; text-align:left">{e(json.dumps(ai_data, ensure_ascii=False, indent=2))}</pre></div>
    """
    return HTMLResponse(page_layout("اختبار AI", body))


@app.get("/admin")
def admin(key: str = "", msg: str = "", q: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)

    query = normalize(q)
    all_products = load_products()
    orders = read_orders()
    new_orders = [o for o in orders if (o.get("status") or "new") != "done"]
    products = all_products
    if query:
        products = [
            item for item in all_products
            if query in normalize(" ".join([item.get("name", ""), item.get("aliases", ""), item.get("notes", "")]))
        ]

    cards = ""
    for idx, item in enumerate(products):
        real_index = all_products.index(item) if item in all_products else idx
        cards += f"""
        <div class="card">
          <div class="product-title">{e(item.get('name'))}</div>
          <div class="product-meta">💰 <b>{e(item.get('price'))}</b> &nbsp; | &nbsp; 📦 <b>{e(item.get('available'))}</b> &nbsp; | &nbsp; 💊 {e(item.get('active_ingredient'))} &nbsp; | &nbsp; {e(item.get('form'))} {e(item.get('strength'))} &nbsp; | &nbsp; 🏷️ {e(item.get('company'))}</div>
          <form class="card-form" method="get" action="/admin/update">
            <input type="hidden" name="key" value="{e(key)}">
            <input type="hidden" name="idx" value="{real_index}">
            <div class="form-grid">
              <div class="field"><label>اسم المنتج</label><input name="name" value="{e(item.get('name'))}" required></div>
              <div class="field"><label>أسماء بديلة</label><input name="aliases" value="{e(item.get('aliases'))}" placeholder="بندول, panadol"></div>
              <div class="field"><label>السعر</label><input name="price" value="{e(item.get('price'))}"></div>
              <div class="field"><label>التوفر</label><input name="available" value="{e(item.get('available'))}"></div>
              <div class="field"><label>المادة الفعالة</label><input name="active_ingredient" value="{e(item.get('active_ingredient'))}" placeholder="Paracetamol"></div>
              <div class="field"><label>الشركة/المنشأ</label><input name="company" value="{e(item.get('company'))}" placeholder="إنجليزي / أردني"></div>
              <div class="field"><label>الشكل</label><input name="form" value="{e(item.get('form'))}" placeholder="أقراص / شراب / تحاميل"></div>
              <div class="field"><label>التركيز</label><input name="strength" value="{e(item.get('strength'))}" placeholder="500mg"></div>
              <div class="field"><label>العبوة</label><input name="pack" value="{e(item.get('pack'))}" placeholder="شريط / علبة"></div>
              <div class="field"><label>صورة المنتج</label><input name="image" value="{e(item.get('image'))}" placeholder="رابط صورة اختياري"></div>
              <div class="field"><label>ملاحظات</label><input name="notes" value="{e(item.get('notes'))}"></div>
            </div>
            <div class="actions"><button type="submit">حفظ</button><a class="btn danger" href="/admin/delete?key={urllib.parse.quote(key)}&idx={real_index}" onclick="return confirm('حذف المنتج؟')">حذف</a></div>
          </form>
        </div>
        """
    if not cards:
        cards = '<div class="box notice">لا توجد منتجات مطابقة. أضف منتجًا أو ارفع ملف Excel/CSV.</div>'

    body = f"""
    <div class="hero"><h1>لوحة إدارة {e(business_name())}</h1><p>إدارة المنتجات، الطلبات، الصور، والردود الآلية.</p></div>
    {admin_nav(key)}
    <div class="stats">
      <div class="stat"><b>{len(all_products)}</b><span>منتج</span></div>
      <div class="stat"><b>{len(new_orders)}</b><span>طلبات جديدة</span></div>
      <div class="stat"><b>{'نعم' if ai_enabled() else 'لا'}</b><span>AI مفعل</span></div>
      <div class="stat"><b>{'نعم' if admin_notify_phone() else 'لا'}</b><span>تنبيه الأدمن</span></div>
    </div>
    <div class="box">
      <p class="msg">{e(msg)}</p>
      <form method="get" action="/admin">
        <input type="hidden" name="key" value="{e(key)}">
        <div class="field"><label>بحث سريع</label><input name="q" value="{e(q)}" placeholder="ابحث باسم المنتج أو الاسم البديل"></div>
        <div class="actions"><button type="submit">بحث</button><a class="btn secondary" href="/admin?key={urllib.parse.quote(key)}">إظهار الكل</a><a class="btn ok" href="/admin/orders?key={urllib.parse.quote(key)}&status=new">الطلبات الجديدة</a></div>
      </form>
    </div>
    <div class="box">
      <h2>إضافة منتج جديد</h2>
      <form method="get" action="/admin/add">
        <input type="hidden" name="key" value="{e(key)}">
        <div class="form-grid">
          <div class="field"><label>اسم المنتج</label><input name="name" placeholder="بنادول" required></div>
          <div class="field"><label>أسماء بديلة</label><input name="aliases" placeholder="بندول, panadol"></div>
          <div class="field"><label>السعر</label><input name="price" placeholder="5 د.ل"></div>
          <div class="field"><label>التوفر</label><input name="available" value="متوفر"></div>
          <div class="field"><label>المادة الفعالة</label><input name="active_ingredient" placeholder="Paracetamol"></div>
          <div class="field"><label>الشركة/المنشأ</label><input name="company" placeholder="إنجليزي / أردني"></div>
          <div class="field"><label>الشكل</label><input name="form" placeholder="أقراص / شراب / تحاميل"></div>
          <div class="field"><label>التركيز</label><input name="strength" placeholder="500mg"></div>
          <div class="field"><label>العبوة</label><input name="pack" placeholder="شريط / علبة"></div>
          <div class="field"><label>صورة المنتج</label><input name="image" placeholder="رابط صورة اختياري"></div>
          <div class="field"><label>ملاحظات</label><input name="notes" placeholder="اختياري"></div>
        </div>
        <div class="actions"><button type="submit">إضافة المنتج</button></div>
      </form>
    </div>
    <h2>المنتجات الحالية ({len(products)})</h2>
    <div class="product-grid">{cards}</div>
    """
    return HTMLResponse(page_layout("لوحة PriceBot", body))

@app.get("/admin/add")
def admin_add(key: str = "", name: str = "", aliases: str = "", price: str = "", available: str = "متوفر", notes: str = "", active_ingredient: str = "", company: str = "", form: str = "", strength: str = "", pack: str = "", image: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    if name.strip():
        upsert_product({
            "name": name.strip(), "aliases": aliases.strip(), "active_ingredient": active_ingredient.strip(),
            "company": company.strip(), "form": form.strip(), "strength": strength.strip(), "pack": pack.strip(),
            "price": price.strip(), "available": available.strip() or "متوفر", "notes": notes.strip(), "image": image.strip()
        })
        return safe_redirect(f"/admin?key={urllib.parse.quote(key)}&msg={urllib.parse.quote('تمت الإضافة أو التحديث')}")
    return safe_redirect(f"/admin?key={urllib.parse.quote(key)}&msg={urllib.parse.quote('لم يتم إدخال اسم المنتج')}")


@app.get("/admin/update")
def admin_update(key: str = "", idx: int = -1, name: str = "", aliases: str = "", price: str = "", available: str = "", notes: str = "", active_ingredient: str = "", company: str = "", form: str = "", strength: str = "", pack: str = "", image: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    products = load_products()
    if 0 <= idx < len(products) and name.strip():
        products[idx] = {
            "name": name.strip(), "aliases": aliases.strip(), "active_ingredient": active_ingredient.strip(),
            "company": company.strip(), "form": form.strip(), "strength": strength.strip(), "pack": pack.strip(),
            "price": price.strip(), "available": available.strip() or "متوفر", "notes": notes.strip(), "image": image.strip()
        }
        save_products(products)
        msg = "تم الحفظ"
    else:
        msg = "لم يتم العثور على المنتج"
    return safe_redirect(f"/admin?key={urllib.parse.quote(key)}&msg={urllib.parse.quote(msg)}")


@app.get("/admin/delete")
def admin_delete(key: str = "", idx: int = -1, name: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    products = load_products()
    if 0 <= idx < len(products):
        products.pop(idx)
        save_products(products)
        msg = "تم الحذف"
    elif name:
        before = len(products)
        products = [p for p in products if p.get("name") != name]
        save_products(products)
        msg = "تم الحذف" if len(products) < before else "لم يتم العثور على المنتج"
    else:
        msg = "لم يتم تحديد المنتج"
    return safe_redirect(f"/admin?key={urllib.parse.quote(key)}&msg={urllib.parse.quote(msg)}")


@app.get("/admin/bulk")
def admin_bulk(key: str = "", msg: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    sample = "بنادول,بندول|panadol,5 د.ل,متوفر,شريط\nبروفين,brufen|ibuprofen,8,متوفر,400mg"
    body = f"""
    <h1>إدخال منتجات بالجملة</h1>
    {admin_nav(key)}
    <div class="box">
      <p class="msg">{e(msg)}</p>
      <p class="notice">كل سطر يمكن أن يكون بالترتيب القديم: الاسم, الأسماء البديلة, السعر, التوفر, ملاحظات. وللملفات الكبيرة الأفضل استخدام قالب CSV الجديد بأعمدة المادة والشكل والتركيز والشركة. يمكن فصل الأسماء البديلة بفاصلة أو علامة |.</p>
      <form method="post" action="/admin/bulk/save?key={urllib.parse.quote(key)}">
        <textarea name="data" placeholder="{e(sample)}"></textarea>
        <p><label><input type="checkbox" name="replace" value="1"> استبدال كل المنتجات الحالية</label></p>
        <button type="submit">استيراد المنتجات</button>
      </form>
    </div>
    """
    return HTMLResponse(page_layout("إدخال منتجات بالجملة", body))


@app.post("/admin/bulk/save")
async def admin_bulk_save(request: Request, key: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    form = await request.form()
    raw_data = str(form.get("data", ""))
    replace = form.get("replace") == "1"
    imported = parse_rows_from_text(raw_data)
    products = [] if replace else load_products()
    products.extend(imported)
    save_products(products)
    return safe_redirect(f"/admin/bulk?key={urllib.parse.quote(key)}&msg={urllib.parse.quote('تم استيراد ' + str(len(imported)) + ' منتج')}")


def parse_rows_from_text(raw_data: str) -> List[dict]:
    if not raw_data:
        return []
    rows: List[List[str]] = []
    if "\t" in raw_data:
        for line in raw_data.splitlines():
            rows.append([x.strip() for x in line.split("\t")])
    else:
        rows = [[x.strip() for x in row] for row in csv.reader(io.StringIO(raw_data)) if row]
    return rows_to_products(rows)



def header_map(headers: List[str]) -> Dict[str, int]:
    normalized = [normalize(h) for h in headers]
    variants = {
        "name": ["name", "product", "product name", "اسم", "الاسم", "اسم المنتج"],
        "aliases": ["aliases", "alias", "اسماء بديله", "اسامي بديله", "بدائل"],
        "active_ingredient": ["active ingredient", "active_ingredient", "ingredient", "generic", "ماده فعاله", "الماده الفعاله", "الماده", "مادة"],
        "brand": ["brand", "trade name", "trade_name", "اسم تجاري", "ماركه", "براند"],
        "company": ["company", "manufacturer", "origin", "country", "الشركه", "الشركة", "المنشا", "المنشأ", "بلد", "البلد"],
        "form": ["form", "dosage form", "dosage_form", "shape", "الشكل", "الشكل الدوائي", "نوع", "النوع"],
        "strength": ["strength", "concentration", "dose", "تركيز", "التركيز", "جرعه", "جرعة", "عيار"],
        "pack": ["pack", "package", "pack size", "pack_size", "عبوه", "العبوه", "التعبئه"],
        "price": ["price", "سعر", "السعر"],
        "available": ["available", "availability", "stock", "توفر", "التوفر", "الحاله", "متوفر"],
        "notes": ["notes", "note", "ملاحظات", "ملاحظه"],
        "image": ["image", "image url", "image_url", "photo", "صوره", "رابط الصوره"],
    }
    mapping: Dict[str, int] = {}
    for field, names in variants.items():
        normalized_names = [normalize(name) for name in names]
        for idx, h in enumerate(normalized):
            if h in normalized_names:
                mapping[field] = idx
                break
    return mapping


def rows_to_products(rows: List[List[str]]) -> List[dict]:
    if not rows:
        return []
    mapping = header_map(rows[0])
    start = 1 if "name" in mapping else 0
    imported = []
    for row in rows[start:]:
        if not row or not any(str(x).strip() for x in row):
            continue
        while len(row) < len(PRODUCT_FIELDS):
            row.append("")
        item = {}
        if mapping:
            for field in PRODUCT_FIELDS:
                idx = mapping.get(field)
                item[field] = row[idx].strip() if idx is not None and idx < len(row) else ""
        else:
            # Backward-compatible order: name, aliases, price, available, notes
            item = {field: "" for field in PRODUCT_FIELDS}
            item["name"] = row[0].strip() if len(row) > 0 else ""
            item["aliases"] = row[1].strip() if len(row) > 1 else ""
            item["price"] = row[2].strip() if len(row) > 2 else ""
            item["available"] = row[3].strip() if len(row) > 3 else ""
            item["notes"] = row[4].strip() if len(row) > 4 else ""
        if not item.get("name") or normalize(item.get("name", "")) in ["name", "اسم", "الاسم", "اسم المنتج"]:
            continue
        item["available"] = item.get("available") or "متوفر"
        imported.append(item)
    return imported


@app.get("/admin/upload")
def admin_upload_page(key: str = "", msg: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    body = f"""
    <h1>رفع ملف منتجات {e(business_name())}</h1>
    {admin_nav(key)}
    <div class="box">
      <p class="msg">{e(msg)}</p>
      <p class="notice">يقبل CSV أو Excel xlsx. الأعمدة المفضلة: name, aliases, active_ingredient, company, form, strength, pack, price, available, notes, image. يدعم أيضًا عناوين عربية مثل: الاسم، السعر، التوفر، ملاحظات.</p>
      <p><a class="btn secondary" href="/admin/template.csv?key={urllib.parse.quote(key)}">تحميل قالب CSV</a></p>
      <form method="post" action="/admin/upload/save?key={urllib.parse.quote(key)}" enctype="multipart/form-data">
        <div class="field"><label>اختر الملف</label><input type="file" name="file" accept=".csv,.xlsx,.txt" required></div>
        <p><label><input type="checkbox" name="replace" value="1"> استبدال كل المنتجات الحالية</label></p>
        <button type="submit">رفع واستيراد</button>
      </form>
    </div>
    """
    return HTMLResponse(page_layout("رفع ملف المنتجات", body))


@app.get("/admin/template.csv")
def admin_template_csv(key: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    content = "\ufeffname,aliases,price,available,notes\nبنادول,بندول|panadol,5 د.ل,متوفر,شريط\nبروفين,brufen|ibuprofen,8,متوفر,400mg\n"
    return PlainTextResponse(
        content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=pricebot_products_template.csv"},
    )


@app.post("/admin/upload/save")
async def admin_upload_save(request: Request, file: UploadFile = File(...), key: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    form = await request.form()
    replace = form.get("replace") == "1"
    content = await file.read()
    filename = (file.filename or "").lower()
    rows: List[List[str]] = []

    if filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            values = ["" if v is None else str(v).strip() for v in row]
            if any(values):
                rows.append(values)
    else:
        text = content.decode("utf-8-sig", errors="ignore")
        rows = [[x.strip() for x in row] for row in csv.reader(io.StringIO(text)) if row]

    imported = rows_to_products(rows)
    products = [] if replace else load_products()
    products.extend(imported)
    save_products(products)
    return safe_redirect(f"/admin/upload?key={urllib.parse.quote(key)}&msg={urllib.parse.quote('تم استيراد ' + str(len(imported)) + ' منتج من الملف')}")





@app.get("/admin/database")
def admin_database(key: str = "", msg: str = ""):
    if not check_admin(key):
        return HTMLResponse("Unauthorized", status_code=401)
    stats = db_stats()
    key_q = urllib.parse.quote(key)
    size_kb = round(stats.get("db_size", 0) / 1024, 1)
    body = f"""
    <div class="hero"><h1>🗄️ قاعدة البيانات</h1><p>تخزين المنتجات والطلبات والذاكرة والكاش داخل SQLite.</p></div>
    {admin_nav(key)}
    {f'<p class="msg">{e(msg)}</p>' if msg else ''}
    <div class="stats">
      <div class="stat"><b>{stats['products']}</b><span>منتجات في SQL</span></div>
      <div class="stat"><b>{stats['orders']}</b><span>طلبات في SQL</span></div>
      <div class="stat"><b>{stats['memory_entries']}</b><span>عناصر ذاكرة</span></div>
      <div class="stat"><b>{size_kb} KB</b><span>حجم قاعدة البيانات</span></div>
    </div>
    <div class="box">
      <h2>المسار</h2>
      <pre>{e(stats['db_file'])}</pre>
      <div class="admin-hint">
        قاعدة البيانات تعمل الآن محلياً على السيرفر. ملفات CSV و JSON تبقى كمرايا/نسخ احتياطية للتصدير وسهولة المراجعة، لكن التشغيل الأساسي صار من SQLite.
      </div>
      <div class="actions">
        <a class="btn secondary" href="/admin/orders/export?key={key_q}">تصدير الطلبات CSV</a>
        <a class="btn secondary" href="/admin/template.csv?key={key_q}">تحميل قالب المنتجات</a>
      </div>
    </div>
    """
    return HTMLResponse(page_layout("قاعدة بيانات PriceBot", body))

@app.get("/admin/memory")
def admin_memory(key: str = "", msg: str = ""):
    if not check_admin(key):
        return HTMLResponse("Unauthorized", status_code=401)
    stats = memory_stats()
    mem = load_memory()
    key_q = urllib.parse.quote(key)
    today_usage = "<br>".join([f"{e(k)}: {e(v)}" for k, v in stats.get("ai_usage_today", {}).items()]) or "لا يوجد استخدام AI اليوم"
    corrections_rows = ""
    for phrase, product in list(mem.get("admin_corrections", {}).items())[-80:]:
        corrections_rows += f"<tr><td>{e(phrase)}</td><td>{e(product)}</td><td><a class='btn danger' href='/admin/memory/delete?key={key_q}&phrase={urllib.parse.quote(phrase)}'>حذف</a></td></tr>"
    query_rows = ""
    for phrase, entry in list(mem.get("query_cache", {}).items())[-80:]:
        if not isinstance(entry, dict):
            continue
        target = entry.get("product_name") or ", ".join(entry.get("product_names", []))
        query_rows += f"<tr><td>{e(phrase)}</td><td>{e(entry.get('type',''))}</td><td>{e(target)}</td><td>{e(entry.get('source',''))}</td><td>{e(entry.get('hits',0))}</td></tr>"
    body = f"""
    <div class="hero"><h1>🧠 ذاكرة البوت</h1><p>تقلل استخدام AI وتحفظ التصحيحات المتكررة.</p></div>
    {admin_nav(key)}
    {f'<p class="msg">{e(msg)}</p>' if msg else ''}
    <div class="stats">
      <div class="stat"><b>{stats['query_cache']}</b><span>أسئلة محفوظة</span></div>
      <div class="stat"><b>{stats['image_cache']}</b><span>صور محفوظة</span></div>
      <div class="stat"><b>{stats['admin_corrections']}</b><span>تصحيحات أدمن</span></div>
      <div class="stat"><b>{stats['product_alias_memory']}</b><span>ذاكرة أسماء</span></div>
    </div>
    <div class="box">
      <h2>استخدام AI اليوم</h2>
      <div class="admin-hint">{today_usage}</div>
    </div>
    <div class="box">
      <h2>إضافة تصحيح يدوي</h2>
      <form method="post" action="/admin/memory/save?key={key_q}" class="form-grid">
        <div class="field"><label>عبارة الزبون</label><input name="phrase" placeholder="مثال: بندل"></div>
        <div class="field"><label>اسم المنتج الصحيح</label><input name="product" placeholder="مثال: بنادول"></div>
        <div class="field" style="grid-column:1/-1"><button type="submit">حفظ التصحيح</button></div>
      </form>
    </div>
    <div class="box">
      <h2>تصحيحات الأدمن</h2>
      <div class="table-wrap"><table><tr><th>العبارة</th><th>المنتج الصحيح</th><th>إجراء</th></tr>{corrections_rows or '<tr><td colspan="3">لا توجد تصحيحات</td></tr>'}</table></div>
    </div>
    <div class="box">
      <h2>آخر أسئلة محفوظة</h2>
      <div class="table-wrap"><table><tr><th>العبارة</th><th>النوع</th><th>النتيجة</th><th>المصدر</th><th>Hits</th></tr>{query_rows or '<tr><td colspan="5">لا توجد ذاكرة أسئلة</td></tr>'}</table></div>
    </div>
    <div class="box">
      <a class="btn danger" href="/admin/memory/clear?key={key_q}" onclick="return confirm('مسح ذاكرة الأسئلة والصور؟ التصحيحات اليدوية لا تُحذف.');">مسح الكاش فقط</a>
    </div>
    """
    return HTMLResponse(page_layout("ذاكرة البوت", body))


@app.post("/admin/memory/save")
async def admin_memory_save(request: Request, key: str = ""):
    if not check_admin(key):
        return HTMLResponse("Unauthorized", status_code=401)
    form = await request.form()
    phrase = memory_query_key(str(form.get("phrase", "")))
    product = str(form.get("product", "")).strip()
    if not phrase or not product:
        return safe_redirect(f"/admin/memory?key={urllib.parse.quote(key)}&msg={urllib.parse.quote('أدخل العبارة والمنتج')}")
    mem = load_memory()
    mem.setdefault("admin_corrections", {})[phrase] = product
    save_memory(mem)
    return safe_redirect(f"/admin/memory?key={urllib.parse.quote(key)}&msg={urllib.parse.quote('تم حفظ التصحيح')}")


@app.get("/admin/memory/delete")
def admin_memory_delete(key: str = "", phrase: str = ""):
    if not check_admin(key):
        return HTMLResponse("Unauthorized", status_code=401)
    mem = load_memory()
    mem.get("admin_corrections", {}).pop(phrase, None)
    save_memory(mem)
    return safe_redirect(f"/admin/memory?key={urllib.parse.quote(key)}&msg={urllib.parse.quote('تم حذف التصحيح')}")


@app.get("/admin/memory/clear")
def admin_memory_clear(key: str = ""):
    if not check_admin(key):
        return HTMLResponse("Unauthorized", status_code=401)
    mem = load_memory()
    mem["query_cache"] = {}
    mem["product_alias_memory"] = {}
    mem["image_cache"] = {}
    save_memory(mem)
    return safe_redirect(f"/admin/memory?key={urllib.parse.quote(key)}&msg={urllib.parse.quote('تم مسح كاش الذاكرة')}")



@app.get("/admin/orders")
def admin_orders(key: str = "", msg: str = "", status: str = "all", q: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    orders = read_orders()
    qn = normalize(q)

    def status_ar(s: str) -> str:
        return {
            "pending": "بانتظار التأكيد",
            "new": "جديد",
            "review": "يحتاج مراجعة",
            "confirmed": "مؤكد",
            "rejected": "مرفوض",
            "done": "منفذ",
        }.get(s or "pending", s or "بانتظار")

    def badge_class(s: str) -> str:
        if s in {"confirmed", "done"}:
            return "done"
        if s == "rejected":
            return "danger"
        return "new"

    filtered = []
    for row in orders:
        row_status = row.get("status") or "pending"
        if status != "all" and row_status != status:
            continue
        if qn:
            hay = normalize(" ".join([row.get("phone",""), row.get("product",""), row.get("notes",""), row.get("message","")]))
            if qn not in hay:
                continue
        filtered.append(row)

    total = len(orders)
    pending_count = sum(1 for r in orders if (r.get("status") or "pending") in {"pending", "new", "review"})
    confirmed_count = sum(1 for r in orders if r.get("status") == "confirmed")
    rejected_count = sum(1 for r in orders if r.get("status") == "rejected")
    done_count = sum(1 for r in orders if r.get("status") == "done")

    cards = ""
    # Display newest first; idx still points to original list index.
    for idx, row in sorted(list(enumerate(orders)), key=lambda x: x[1].get("time", ""), reverse=True):
        if row not in filtered:
            continue
        row_status = row.get("status") or "pending"
        phone = row.get("phone", "")
        phone_link = f"https://wa.me/{urllib.parse.quote(phone)}" if phone else "#"
        msg_text = row.get("message", "")
        image_link = ""
        m = re.search(r"https?://\S+", msg_text or "")
        if m:
            image_link = f'<a class="btn secondary" target="_blank" href="{e(m.group(0))}">فتح الصورة</a>'

        confirm_btn = ""
        reject_btn = ""
        done_btn = ""
        if row_status in {"pending", "new", "review"}:
            confirm_btn = f'<a class="btn ok" href="/admin/orders/confirm?key={urllib.parse.quote(key)}&idx={idx}" onclick="return confirm(\'تأكيد الطلب وإرسال رسالة للزبون؟\')">تأكيد للزبون</a>'
            reject_btn = f'<a class="btn danger" href="/admin/orders/reject?key={urllib.parse.quote(key)}&idx={idx}" onclick="return confirm(\'رفض الطلب وإرسال اعتذار للزبون؟\')">رفض</a>'
        if row_status in {"confirmed", "pending", "new", "review"}:
            done_btn = f'<a class="btn secondary" href="/admin/orders/done?key={urllib.parse.quote(key)}&idx={idx}">تم التنفيذ</a>'

        cards += f"""
        <div class="card order-card">
          <div class="order-head"><div class="order-title">{e(row.get('product'))}</div><span class="badge {badge_class(row_status)}">{status_ar(row_status)}</span></div>
          <div class="order-row"><strong>الوقت</strong><span>{e(row.get('time'))}</span></div>
          <div class="order-row"><strong>رقم الزبون</strong><a href="{phone_link}" target="_blank">{e(phone)}</a></div>
          <div class="order-row"><strong>السعر</strong><span>{e(row.get('price') or '-')}</span></div>
          <div class="order-row"><strong>التوفر</strong><span>{e(row.get('available') or '-')}</span></div>
          <div class="order-row"><strong>ملاحظة</strong><span>{e(row.get('notes') or '-')}</span></div>
          <div class="order-row"><strong>رسالة</strong><span>{e((msg_text or '-')[:180])}</span></div>
          <div class="actions">{confirm_btn}{reject_btn}{done_btn}<a class="btn secondary" target="_blank" href="{phone_link}">مراسلة الزبون</a>{image_link}<a class="btn danger" href="/admin/orders/delete?key={urllib.parse.quote(key)}&idx={idx}" onclick="return confirm('حذف الطلب؟')">حذف</a></div>
        </div>
        """
    if not cards:
        cards = '<div class="box notice">لا توجد طلبات في هذا القسم.</div>'

    key_q = urllib.parse.quote(key)
    body = f"""
    <div class="hero"><h1>طلبات الحجز</h1><p>{e(business_name())} — الطلب لا يعتبر مؤكداً حتى تضغط زر التأكيد.</p></div>
    {admin_nav(key)}
    <div class="stats">
      <div class="stat"><b>{total}</b><span>كل الطلبات</span></div>
      <div class="stat"><b>{pending_count}</b><span>بانتظار التأكيد</span></div>
      <div class="stat"><b>{confirmed_count}</b><span>مؤكدة</span></div>
      <div class="stat"><b>{rejected_count}</b><span>مرفوضة</span></div>
    </div>
    <div class="box">
      <p class="msg">{e(msg)}</p>
      <form method="get" action="/admin/orders">
        <input type="hidden" name="key" value="{e(key)}">
        <div class="form-grid">
          <div class="field"><label>بحث في الطلبات</label><input name="q" value="{e(q)}" placeholder="رقم زبون / منتج / ملاحظة"></div>
          <div class="field"><label>الحالة</label><select name="status">
            <option value="all" {'selected' if status=='all' else ''}>كل الطلبات</option>
            <option value="pending" {'selected' if status=='pending' else ''}>بانتظار التأكيد</option>
            <option value="review" {'selected' if status=='review' else ''}>يحتاج مراجعة</option>
            <option value="confirmed" {'selected' if status=='confirmed' else ''}>مؤكدة</option>
            <option value="rejected" {'selected' if status=='rejected' else ''}>مرفوضة</option>
            <option value="done" {'selected' if status=='done' else ''}>منفذة</option>
          </select></div>
        </div>
        <div class="actions"><button type="submit">تطبيق</button><a class="btn secondary" href="/admin/orders?key={key_q}&status=pending">بانتظار التأكيد</a><a class="btn secondary" href="/admin/orders?key={key_q}&status=review">مراجعة الصور</a><a class="btn ok" href="/admin/daily-report/send?key={key_q}">إرسال تقرير اليوم للأدمن</a><a class="btn secondary" href="/admin/orders/export?key={key_q}">تصدير CSV</a></div>
      </form>
    </div>
    <div class="product-grid">{cards}</div>
    """
    return HTMLResponse(page_layout("طلبات PriceBot", body))


@app.get("/admin/orders/confirm")
def admin_orders_confirm(key: str = "", idx: int = -1):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    orders = read_orders()
    if 0 <= idx < len(orders):
        orders[idx]["status"] = "confirmed"
        write_orders(orders)
        send_whatsapp_message(orders[idx].get("phone", ""), customer_order_confirmed_message(orders[idx]))
        msg = "تم تأكيد الطلب وإرسال رسالة للزبون"
    else:
        msg = "لم يتم العثور على الطلب"
    return safe_redirect(f"/admin/orders?key={urllib.parse.quote(key)}&msg={urllib.parse.quote(msg)}")


@app.get("/admin/orders/reject")
def admin_orders_reject(key: str = "", idx: int = -1):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    orders = read_orders()
    if 0 <= idx < len(orders):
        orders[idx]["status"] = "rejected"
        write_orders(orders)
        send_whatsapp_message(orders[idx].get("phone", ""), customer_order_rejected_message(orders[idx]))
        msg = "تم رفض الطلب وإرسال رسالة للزبون"
    else:
        msg = "لم يتم العثور على الطلب"
    return safe_redirect(f"/admin/orders?key={urllib.parse.quote(key)}&msg={urllib.parse.quote(msg)}")


@app.get("/admin/orders/done")
def admin_orders_done(key: str = "", idx: int = -1):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    orders = read_orders()
    if 0 <= idx < len(orders):
        orders[idx]["status"] = "done"
        write_orders(orders)
        msg = "تم تحديث الطلب كمنفذ"
    else:
        msg = "لم يتم العثور على الطلب"
    return safe_redirect(f"/admin/orders?key={urllib.parse.quote(key)}&msg={urllib.parse.quote(msg)}")


@app.get("/admin/orders/delete")
def admin_orders_delete(key: str = "", idx: int = -1):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    orders = read_orders()
    if 0 <= idx < len(orders):
        orders.pop(idx)
        write_orders(orders)
        msg = "تم حذف الطلب"
    else:
        msg = "لم يتم العثور على الطلب"
    return safe_redirect(f"/admin/orders?key={urllib.parse.quote(key)}&msg={urllib.parse.quote(msg)}")


@app.get("/admin/orders/export")
def admin_orders_export(key: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=ORDER_FIELDS)
    writer.writeheader()
    for row in read_orders():
        writer.writerow({field: row.get(field, "") for field in ORDER_FIELDS})
    return PlainTextResponse(
        "\ufeff" + output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=pricebot_orders.csv"},
    )


@app.get("/admin/daily-report/send")
def admin_daily_report_send(key: str = "", auto: str = "0"):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    ok = notify_admin(build_daily_report_message(), "")
    if auto == "1":
        return {"ok": bool(ok)}
    msg = "تم إرسال التقرير اليومي للأدمن" if ok else "لم يتم الإرسال: تأكد من رقم الأدمن"
    return safe_redirect(f"/admin/orders?key={urllib.parse.quote(key)}&msg={urllib.parse.quote(msg)}")


@app.get("/admin/media/{filename}")
def admin_media(filename: str, key: str = ""):
    if not check_admin(key):
        return PlainTextResponse("Forbidden", status_code=403)
    safe_name = Path(filename).name
    path = MEDIA_DIR / safe_name
    if not path.exists() or not path.is_file():
        return PlainTextResponse("Not found", status_code=404)
    return FileResponse(path)

@app.get("/privacy")
def privacy_policy():
    return PlainTextResponse(
        """
PriceBot Privacy Policy

PriceBot receives WhatsApp messages only to respond to customer product and price inquiries.
We do not sell user data.
We do not share customer messages with advertisers.
Messages may be processed to provide automated replies about product availability and prices.
Users can request deletion of their data by contacting the business owner.

Data deletion URL:
/delete-data
""".strip()
    )


@app.get("/delete-data")
def delete_data():
    return PlainTextResponse(
        """
Data Deletion Instructions

To request deletion of your WhatsApp messages or customer data, contact the business owner and provide your WhatsApp number.
The business owner will delete your related records from the system when applicable.
""".strip()
    )


@app.get("/webhook/whatsapp")
async def verify_webhook(request: Request):
    params = request.query_params
    if params.get("hub.mode") == "subscribe" and params.get("hub.verify_token") == VERIFY_TOKEN:
        return PlainTextResponse(params.get("hub.challenge", ""))
    return PlainTextResponse("Forbidden", status_code=403)



@app.post("/webhook/whatsapp")
async def whatsapp_webhook(request: Request):
    data = await request.json()
    try:
        for entry in data.get("entry", []):
            for change in entry.get("changes", []):
                value = change.get("value", {})
                if "statuses" in value and "messages" not in value:
                    return JSONResponse({"status": "status_event_ignored"})
                for msg in value.get("messages", []):
                    from_number = msg.get("from")
                    if not from_number:
                        continue
                    msg_type = msg.get("type")
                    if msg_type == "text":
                        text_msg = msg.get("text", {}).get("body", "")
                        reply = build_reply(text_msg, from_number)
                    elif msg_type == "image":
                        reply = process_image_message(msg, from_number)
                    elif msg_type in {"document", "audio", "video"}:
                        save_review_order(from_number, f"ملف {msg_type} يحتاج مراجعة", f"وصل ملف من نوع {msg_type}", "")
                        notify_admin(
                            f"📎 ملف يحتاج مراجعة - {pharmacy_name()}\n\nرقم الزبون: {from_number}\nنوع الملف: {msg_type}\nالوقت: {now_str()}",
                            from_number,
                        )
                        reply = build_image_under_review_reply()
                    else:
                        reply = (
                            f"{pharmacy_name()} 🌿\n\n"
                            "أستطيع الرد على أسماء المنتجات المكتوبة أو صور المنتجات.\n"
                            "للروشتات أو الصور غير الواضحة سيتم تحويلها للصيدلي."
                        )
                    send_whatsapp_message(from_number, reply)
    except Exception as exc:
        print("WEBHOOK PROCESS ERROR:", str(exc), flush=True)
    return JSONResponse({"status": "received"})

@app.get("/test")
def test_reply(q: str = ""):
    return {"reply": build_reply(q)}


if __name__ == "__main__":
    port = int(os.getenv("PORT", "8095"))
    uvicorn.run(app, host="0.0.0.0", port=port)
