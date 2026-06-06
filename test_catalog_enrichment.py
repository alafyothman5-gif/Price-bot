import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

import sys
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from tools.enrich_products_catalog import enrich_catalog, write_workbook, write_report, build_substitution_suggestions


class CatalogEnrichmentTests(unittest.TestCase):
    def make_input(self, rows):
        tmp = tempfile.TemporaryDirectory()
        path = Path(tmp.name) / "PriceList.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "price", "quantity", "barcode"])
        for row in rows:
            ws.append(row)
        wb.save(path)
        return tmp, path

    def test_examples_are_extracted_without_ai(self):
        tmp, path = self.make_input([
            ["CeraVe Hydrating Cleanser 236ml", "35 LYD", 3, ""],
            ["Flagyl Syrup 125mg/5ml", "7.5", 10, ""],
            ["Rilastil Xerolact PB Balm", "45", 2, ""],
            ["Amoclan Syrup 457", "12", 5, ""],
        ])
        self.addCleanup(tmp.cleanup)
        rows, extra = enrich_catalog(path, ROOT / "data", default_available="blank")
        by_name = {r["name"]: r for r in rows}

        cerave = by_name["CeraVe Hydrating Cleanser 236ml"]
        self.assertEqual(cerave["brand"], "CeraVe")
        self.assertEqual(cerave["category"], "cosmetic")
        self.assertEqual(cerave["form"], "cleanser")
        self.assertEqual(cerave["size"], "236ml")
        self.assertIn("cerave", cerave["aliases"])
        self.assertIn("cleanser", cerave["ocr_keywords"])
        self.assertEqual(cerave["review_status"], "ready")

        flagyl = by_name["Flagyl Syrup 125mg/5ml"]
        self.assertEqual(flagyl["brand"], "Flagyl")
        self.assertEqual(flagyl["category"], "medicine")
        self.assertEqual(flagyl["active_ingredient"], "metronidazole")
        self.assertEqual(flagyl["form"], "syrup")
        self.assertEqual(flagyl["strength"], "125mg/5ml")
        self.assertEqual(flagyl["review_status"], "ready")

        rilastil = by_name["Rilastil Xerolact PB Balm"]
        self.assertEqual(rilastil["brand"], "Rilastil")
        self.assertEqual(rilastil["category"], "cosmetic")
        self.assertEqual(rilastil["form"], "balm")
        self.assertIn("xerolact", rilastil["ocr_keywords"])

        amoclan = by_name["Amoclan Syrup 457"]
        self.assertEqual(amoclan["brand"], "Amoclan")
        self.assertEqual(amoclan["category"], "medicine")
        self.assertEqual(amoclan["active_ingredient"], "amoxicillin + clavulanic acid")
        self.assertEqual(amoclan["form"], "syrup")
        self.assertEqual(amoclan["review_status"], "needs_review")
        self.assertIn("strength unit unclear", amoclan["review_notes"])

    def test_outputs_have_required_headers(self):
        tmp, path = self.make_input([["CeraVe Hydrating Cleanser 236ml", "35", 1, ""]])
        self.addCleanup(tmp.cleanup)
        rows, extra = enrich_catalog(path, ROOT / "data", default_available="true")
        out = Path(tmp.name) / "ready.xlsx"
        report = Path(tmp.name) / "report.json"
        write_workbook(out, [r for r in rows if r["review_status"] == "ready"])
        rep = write_report(report, rows, extra, {"default_available": "true"})
        wb = load_workbook(out, read_only=True)
        headers = [c.value for c in next(wb.active.iter_rows(max_row=1))]
        for required in ["product_id", "name", "brand", "product_family", "category", "form", "price", "review_status", "review_notes"]:
            self.assertIn(required, headers)
        self.assertEqual(rep["total_products"], 1)
        self.assertTrue(report.exists())

    def test_suggested_substitutions_are_review_only(self):
        tmp, path = self.make_input([
            ["Panadol 500mg 24 Tabs", "10", 1, ""],
            ["Adol 500mg 24 Tabs", "8", 1, ""],
        ])
        self.addCleanup(tmp.cleanup)
        rows, _ = enrich_catalog(path, ROOT / "data", default_available="true")
        suggestions = build_substitution_suggestions(rows)
        self.assertTrue(suggestions)
        self.assertIn("pharmacist approval required", suggestions[0]["review_required"])


if __name__ == "__main__":
    unittest.main()
