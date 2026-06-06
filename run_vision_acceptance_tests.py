#!/usr/bin/env python3
"""Run offline Vision acceptance tests against the local matcher.

Input file: vision_test_cases.xlsx (or CSV)
Required/optional columns:
  image_path, vision_output_json, expected_brand, expected_product_family,
  expected_form, expected_strength, expected_size, expected_decision,
  expected_product_id, notes

This runner is intentionally local-safe: Vision AI extraction is tested through
vision_output_json. If no vision_output_json is supplied, the case fails with
missing_vision_output_json instead of guessing from the image file.
"""
from __future__ import annotations

import csv
import json
import sys
from pathlib import Path
from typing import Dict, List

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import database  # noqa: E402
import matcher  # noqa: E402


def read_cases(path: Path) -> List[Dict[str, str]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return [dict(r) for r in csv.DictReader(f)]
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {headers[i]: str(row[i] or "") for i in range(len(headers)) if headers[i]}
        if any(v.strip() for v in data.values()):
            rows.append(data)
    return rows


def write_report(path: Path, rows: List[Dict[str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "vision_report"
    headers = [
        "image_path", "expected_decision", "matcher_decision", "expected_product_id",
        "matched_product_id", "pass", "failure_reason", "vision_output", "notes",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(60, max(12, max(len(str(c.value or "")) for c in col) + 2))
    wb.save(path)


def main() -> None:
    input_path = Path(sys.argv[1] if len(sys.argv) > 1 else "vision_test_cases.xlsx")
    output_path = Path(sys.argv[2] if len(sys.argv) > 2 else "vision_test_report.xlsx")
    if not input_path.exists():
        raise SystemExit(f"Input not found: {input_path}")
    database.init_db()
    cases = read_cases(input_path)
    report = []
    for case in cases:
        expected = str(case.get("expected_decision") or "").strip().upper()
        vision_raw = str(case.get("vision_output_json") or "").strip()
        result = {
            "image_path": case.get("image_path", ""),
            "expected_decision": expected,
            "expected_product_id": case.get("expected_product_id", ""),
            "notes": case.get("notes", ""),
        }
        if not vision_raw:
            result.update({"matcher_decision": "", "matched_product_id": "", "pass": "FAIL", "failure_reason": "missing_vision_output_json", "vision_output": ""})
            report.append(result)
            continue
        try:
            ai_data = json.loads(vision_raw)
        except Exception as exc:
            result.update({"matcher_decision": "", "matched_product_id": "", "pass": "FAIL", "failure_reason": f"invalid_json:{exc}", "vision_output": vision_raw})
            report.append(result)
            continue
        decision = matcher.resolve_image_query_decision(ai_data)
        matched_id = ""
        if decision.product:
            matched_id = str(decision.product.get("id") or decision.product.get("product_id") or "")
        ok = bool(expected and decision.decision_type.name == expected)
        if case.get("expected_product_id"):
            ok = ok and matched_id == str(case.get("expected_product_id"))
        result.update({
            "matcher_decision": decision.decision_type.name,
            "matched_product_id": matched_id,
            "pass": "PASS" if ok else "FAIL",
            "failure_reason": "" if ok else f"expected {expected} got {decision.decision_type.name}; reason={decision.reason}",
            "vision_output": json.dumps(ai_data, ensure_ascii=False),
        })
        report.append(result)
    write_report(output_path, report)
    total = len(report)
    passed = sum(1 for r in report if r.get("pass") == "PASS")
    print(f"VISION_ACCEPTANCE_REPORT: {passed}/{total} passed -> {output_path}")
    if passed != total:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
